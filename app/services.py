from __future__ import annotations

import gc
import json
import logging
import os
import re
import threading
import time
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from sqlalchemy import text
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session
from app.config import settings
from app import wind_bulk, wind_sql
from app.bg_threads import ShutdownRequested, raise_if_shutting_down
from app.sql_dialect import (
    sql_date_compact_expr,
    sql_max_date_expr,
    sql_minutes_ago,
    sql_now,
    sql_order_date_asc,
    sql_order_date_desc,
)

_log = logging.getLogger(__name__)
_job_running = False
_progress_log_lock = threading.Lock()
_progress_log_last: dict[str, float] = {}
# 单批越小，Wind SQL Server 上「多表 OUTER APPLY」越不易被对端掐断(10054)；过大会整批失败
_WIND_QUOTE_CHUNK = 30
_SID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_-]{0,63}$")


def _process_rss_mb() -> float | None:
    """Best-effort peak RSS for logs; no psutil dependency."""
    try:
        import resource

        rss = float(resource.getrusage(resource.RUSAGE_SELF).ru_maxrss)
        if rss <= 0:
            return None
        return rss / (1024 * 1024) if rss > 10_000_000 else rss / 1024
    except Exception:
        return None


def _runtime_suffix() -> str:
    parts: list[str] = []
    rss = _process_rss_mb()
    if rss is not None:
        parts.append(f"rss_peak={rss:.1f}MB")
    try:
        g0, g1, g2 = gc.get_count()
        parts.append(f"gc={g0}/{g1}/{g2}")
    except Exception:
        pass
    return f" ({', '.join(parts)})" if parts else ""


def _log_runtime_progress(
    key: str,
    message: str,
    *,
    min_interval: float = 30.0,
    force: bool = False,
) -> None:
    """Throttled stdout progress that avoids extra DB reads/writes."""
    now_m = time.monotonic()
    with _progress_log_lock:
        last = float(_progress_log_last.get(key) or 0.0)
        if not force and now_m - last < min_interval:
            return
        _progress_log_last[key] = now_m
    _log.info("%s%s", message, _runtime_suffix())


def _strategy_excel_path(file_dir: str | None, file_name: str) -> str:
    """策略 Excel 绝对路径（支持 SERVER_UPLOAD_ROOT 下 strategies/ 与本地 STRATEGY_ROOT_DIR）。"""
    from app.server_files import resolve_strategy_excel_path

    return resolve_strategy_excel_path(file_dir, file_name)


def _safe_return(a: float | None, b: float | None) -> float | None:
    if a is None or b is None:
        return None
    if a <= 0 or b <= 0:
        return None
    return a / b - 1.0


def _compact_date(v: object) -> str:
    if v is None:
        return ""
    d = _row_sql_date(v)
    if d is not None:
        return d.strftime("%Y%m%d")
    if isinstance(v, datetime):
        return v.strftime("%Y%m%d")
    s = str(v).strip().replace("-", "")
    return s[:8] if len(s) >= 8 else s.zfill(8)


_NAV_MAX_TRADE_DAY_GAP_CALENDAR_DAYS = 12


def _nav_large_trade_day_gap(
    trade_days: list[str],
    *,
    max_calendar_days: int = _NAV_MAX_TRADE_DAY_GAP_CALENDAR_DAYS,
) -> tuple[str, str, int] | None:
    prev: date | None = None
    prev_c = ""
    for raw in trade_days:
        c = str(raw or "").strip()[:8]
        if len(c) < 8:
            continue
        try:
            cur = datetime.strptime(c, "%Y%m%d").date()
        except ValueError:
            continue
        if prev is not None:
            gap_days = (cur - prev).days
            if gap_days > max_calendar_days:
                return prev_c, c, gap_days
        prev = cur
        prev_c = c
    return None


def _nav_trade_days_continuous(
    sid: str,
    trade_days: list[str],
    *,
    label: str,
    sync_job_id: int | None = None,
    progress_cb: Callable[[str], None] | None = None,
    db: Session | None = None,
    turso_remote: bool = False,
) -> bool:
    gap = _nav_large_trade_day_gap(trade_days)
    if gap is None:
        return True
    prev_c, next_c, gap_days = gap
    msg = (
        f"{sid}：行情交易日列表异常断档 {prev_c}→{next_c}"
        f"（{gap_days} 自然日，{label}），净值重建已停止"
    )
    _log.warning("nav %s: %s", sid, msg)
    if sync_job_id is not None or progress_cb:
        _nav_progress_touch(
            msg,
            sync_job_id=sync_job_id,
            progress_cb=progress_cb,
            db=None if turso_remote else db,
            turso_remote=turso_remote,
        )
    return False


def _wind_code_key(code) -> str:
    """与 Wind 行情 dict 查找统一（大小写、首尾空格）。"""
    if code is None:
        return ""
    return str(code).strip().upper()


def _row_sql_date(v: object) -> date | None:
    """库表 DATE/TEXT 等转为 date，便于比较调仓日与快照日（Turso 常返回 str）。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    compact = s.replace("-", "").replace("/", "")[:8]
    if len(compact) == 8 and compact.isdigit():
        try:
            return datetime.strptime(compact, "%Y%m%d").date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(s[:10]).date()
    except ValueError:
        return None


def format_sql_date_display(v: object) -> str:
    """库内日期 → YYYY-MM-DD 展示；无效则返回空串。"""
    d = _row_sql_date(v)
    return d.isoformat() if d else ""


def latest_rebalance_date_by_strategy(db: Session) -> dict[str, str]:
    """各策略在 strategy_positions 中已导入的最大调仓日（最新一期）。"""
    rows = db.execute(
        text(
            f"""
            SELECT strategy_id, {sql_max_date_expr("rebalance_date")} AS latest_rb
            FROM strategy_positions
            GROUP BY strategy_id
            """
        )
    ).mappings().all()
    out: dict[str, str] = {}
    for r in rows:
        sid = str(r.get("strategy_id") or "").strip()
        if not sid:
            continue
        disp = format_sql_date_display(r.get("latest_rb"))
        if disp:
            out[sid] = disp
    return out


def _adj_close_td_ff(
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    sc: str,
    td: str,
    last_close_fill: dict[str, float],
) -> float | None:
    """当日复权收盘；无则沿用最近一次有效收盘（前向填充）。"""
    row = day_map.get(sc, {}).get(td)
    if row:
        cl, _ = row
        if cl is not None and not (isinstance(cl, float) and cl != cl) and cl > 0:
            last_close_fill[sc] = float(cl)
            return float(cl)
    pv = last_close_fill.get(sc)
    return pv if pv is not None and pv > 0 else None


def _strategy_nav_notional_capital() -> float:
    """固定股数法名义本金（元），nav_unit = 收盘市值 / 该值；收益序列与取值无关。"""
    v = float(settings.strategy_nav_initial_capital)
    return v if v > 0 else 100_000_000.0


def _import_progress_touch(
    db: Session,
    message: str,
    *,
    sync_job_id: int | None = None,
    strategy_import_job_id: int | None = None,
    do_commit: bool = False,
) -> None:
    """阶段1 Excel 导入：同步写入 admin_sync / strategy_import 任务说明。"""
    msg = (message or "")[:6000]
    if sync_job_id is not None:
        _admin_sync_job_touch(sync_job_id, "import", msg, db=db, do_commit=do_commit)
    if strategy_import_job_id is not None:
        _strategy_import_job_touch(
            db, strategy_import_job_id, message=msg, do_commit=do_commit
        )


def _nav_progress_touch(
    message: str,
    *,
    sync_job_id: int | None,
    progress_cb: Callable[[str], None] | None,
    db: Session | None = None,
    turso_remote: bool = False,
) -> None:
    msg = (message or "")[:6000]
    if sync_job_id is not None:
        _admin_sync_job_touch(
            sync_job_id,
            "nav",
            msg,
            db=None if turso_remote else db,
            do_commit=True,
        )
    elif progress_cb:
        progress_cb(msg)


def _job_progress(
    db: Session,
    job_id: int,
    message: str,
    do_commit: bool = True,
    *,
    sync_job_id: int | None = None,
) -> None:
    """更新数据更新任务进度（立即 commit，刷新页面可见）。"""
    msg = (message or "")[:6000]
    db.execute(
        text(
            f"""
            UPDATE strategy_update_jobs
            SET message=:m, progress_at={sql_now()}
            WHERE id=:id AND status='RUNNING'
            """
        ),
        {"m": msg, "id": job_id},
    )
    if sync_job_id is not None:
        _admin_sync_job_touch(
            sync_job_id, "holding_update", msg, db=db, do_commit=False
        )
    if do_commit:
        db.commit()


def _excel_meta_strategy_labels(df: pd.DataFrame) -> dict[str, str]:
    """从持仓 Excel 读取策略级「分类」「调仓频率」（列存在且首个非空才写入）。"""
    out: dict[str, str] = {}
    if "分类" in df.columns or "策略分类" in df.columns:
        val = ""
        for col in ("分类", "策略分类"):
            if col not in df.columns:
                continue
            for x in df[col].tolist():
                if x is None or (isinstance(x, float) and pd.isna(x)):
                    continue
                t = str(x).strip()
                if t:
                    val = t[:128]
                    break
        if val:
            out["strategy_category"] = val
    if "调仓频率" in df.columns:
        val = ""
        for x in df["调仓频率"].tolist():
            if x is None or (isinstance(x, float) and pd.isna(x)):
                continue
            t = str(x).strip()
            if t:
                val = t[:128]
                break
        if val:
            out["rebalance_frequency"] = val
    return out


def _first_nonempty_meta_from_openpyxl(
    ws: Any, col_index: dict[str, int], col_names: tuple[str, ...], *, max_rows: int = 3000
) -> str:
    from app.text_encoding import normalize_unicode_text

    idx = next((col_index[c] for c in col_names if c in col_index), None)
    if idx is None:
        return ""
    max_c = _strategy_excel_read_max_col()
    n = 0
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=max_c, values_only=True):
        n += 1
        if n > max_rows:
            break
        if not row or idx >= len(row):
            continue
        cell = row[idx]
        if cell is None or (isinstance(cell, float) and pd.isna(cell)):
            continue
        t = normalize_unicode_text(cell, max_len=128)
        if t:
            return t
    return ""


def _read_strategy_excel_label_meta_openpyxl(file_path: str) -> dict[str, str]:
    """与流式导入同一套 openpyxl 读表头/元数据，避免 pandas 引擎差异导致中文列名或值乱码。"""
    from openpyxl import load_workbook

    out: dict[str, str] = {}
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        max_c = _strategy_excel_read_max_col()
        header_row = next(
            ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_c, values_only=True),
            None,
        )
        if not header_row:
            return out
        col_index: dict[str, int] = {}
        for i, c in enumerate(header_row):
            name = str(c).strip() if c is not None else ""
            if name and name not in col_index:
                col_index[name] = i
        cat = _first_nonempty_meta_from_openpyxl(ws, col_index, ("分类", "策略分类"))
        if cat:
            out["strategy_category"] = cat
        freq = _first_nonempty_meta_from_openpyxl(ws, col_index, ("调仓频率",))
        if freq:
            out["rebalance_frequency"] = freq
    finally:
        wb.close()
    return out


_STRATEGY_EXCEL_REQUIRED_COLS = ("调整日期", "证券代码")
_STRATEGY_EXCEL_OPTIONAL_COLS = frozenset(
    {"持仓权重", "行业中性权重", "分类", "策略分类", "调仓频率"}
)
_STRATEGY_EXCEL_STREAM_SUFFIXES = (".xlsx", ".xlsm")


def _strategy_excel_read_max_col() -> int:
    """策略模板有效数据在前 N 列（默认 A～E）；忽略右侧格式/空列，避免宽表 OOM。"""
    n = int(getattr(settings, "strategy_excel_read_max_col", 5) or 5)
    return max(2, min(n, 20))


def _strategy_excel_pandas_usecols() -> list[int]:
    return list(range(_strategy_excel_read_max_col()))


def _strategy_excel_use_streaming(file_path: str) -> bool:
    if not bool(getattr(settings, "strategy_excel_streaming_import", True)):
        return False
    suf = Path(file_path).suffix.lower()
    if suf not in _STRATEGY_EXCEL_STREAM_SUFFIXES:
        return False
    min_mb = int(getattr(settings, "strategy_excel_streaming_min_mb", 0))
    try:
        size = os.path.getsize(file_path)
    except OSError:
        return True
    if min_mb <= 0:
        return True
    return size >= min_mb * 1024 * 1024


def _read_strategy_excel_label_meta(file_path: str) -> dict[str, str]:
    """只读分类/调仓频率列的前若干行，避免为 meta 加载整表。"""
    suf = Path(file_path).suffix.lower()
    if suf in _STRATEGY_EXCEL_STREAM_SUFFIXES:
        try:
            return _read_strategy_excel_label_meta_openpyxl(file_path)
        except Exception:
            _log.warning("openpyxl label meta failed, fallback pandas: %s", file_path, exc_info=True)
    meta_names = {"分类", "策略分类", "调仓频率"}
    try:
        usecols = _strategy_excel_pandas_usecols()
        header = pd.read_excel(file_path, sheet_name=0, nrows=0, usecols=usecols)
        cols = [str(c).strip() for c in header.columns]
        pick = [c for c in cols if c in meta_names]
        if not pick:
            return {}
        kw: dict[str, Any] = {"sheet_name": 0, "usecols": usecols, "nrows": 3000}
        try:
            df = pd.read_excel(file_path, engine="calamine", **kw)
        except Exception:
            df = pd.read_excel(file_path, **kw)
        return _excel_meta_strategy_labels(df)
    except Exception:
        return {}


def _excel_cell_rebalance_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            ts = pd.Timestamp("1899-12-30") + pd.Timedelta(days=float(v))
            return ts.date()
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    if not s:
        return None
    try:
        ts = pd.to_datetime(s, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.date() if hasattr(ts, "date") else pd.Timestamp(ts).date()
    except Exception:
        return None


def _excel_cell_weight(v: Any) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _strategy_excel_row_batch_size() -> int:
    n = int(getattr(settings, "strategy_excel_import_row_batch", 600) or 600)
    if _wind_low_memory_mode():
        return max(80, min(n, 120))
    return max(200, min(n, 1500))


def _strategy_import_progress_interval() -> int:
    return max(1, int(getattr(settings, "strategy_import_progress_every_batches", 8) or 8))


def _strategy_import_batch_retry_max() -> int:
    return max(3, int(getattr(settings, "strategy_import_batch_retry", 5) or 5))


def _strategy_import_position_batch_size() -> int:
    n = int(getattr(settings, "strategy_import_position_batch_size", 500) or 500)
    if _wind_low_memory_mode():
        return max(40, min(n, 100))
    return max(50, n)


def _strategy_excel_parse_header(header_row: tuple[Any, ...] | None) -> dict[str, int]:
    if not header_row:
        raise ValueError("Excel 表为空")
    max_c = _strategy_excel_read_max_col()
    idx: dict[str, int] = {}
    for i, c in enumerate(header_row[:max_c]):
        name = str(c).strip() if c is not None else ""
        if name and name not in idx:
            idx[name] = i
    for need in _STRATEGY_EXCEL_REQUIRED_COLS:
        if need not in idx:
            raise ValueError(f"缺少列: {need}（须在前 {max_c} 列内）")
    return idx


def _strategy_excel_fill_meta_from_cells(
    cells: tuple[Any, ...],
    col_index: dict[str, int],
    meta_out: dict[str, str],
    *,
    category_done: list[bool],
    freq_done: list[bool],
) -> None:
    from app.text_encoding import normalize_unicode_text

    if not category_done[0]:
        for name, key in (("分类", "strategy_category"), ("策略分类", "strategy_category")):
            li = col_index.get(name)
            if li is None or li >= len(cells):
                continue
            cell = cells[li]
            if cell is None or (isinstance(cell, float) and pd.isna(cell)):
                continue
            t = normalize_unicode_text(cell, max_len=128)
            if t:
                meta_out[key] = t
                category_done[0] = True
                break
    if not freq_done[0]:
        li = col_index.get("调仓频率")
        if li is not None and li < len(cells):
            cell = cells[li]
            if cell is not None and not (isinstance(cell, float) and pd.isna(cell)):
                t = normalize_unicode_text(cell, max_len=128)
                if t:
                    meta_out["rebalance_frequency"] = t
                    freq_done[0] = True


def _iter_strategy_holdings_excel_batches(
    file_path: str,
    *,
    meta_out: dict[str, str] | None = None,
) -> Any:
    """
    openpyxl 流式按批产出持仓行；固定只读前 N 列（默认 A～E），单遍扫描。
    """
    from openpyxl import load_workbook

    batch_size = _strategy_excel_row_batch_size()
    max_c = _strategy_excel_read_max_col()
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row = next(
            ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=max_c, values_only=True),
            None,
        )
        col_index = _strategy_excel_parse_header(header_row)
        i_dt = col_index["调整日期"]
        i_code = col_index["证券代码"]
        i_hw = col_index.get("持仓权重")
        i_iw = col_index.get("行业中性权重")
        batch: list[tuple[date, str, float | None, float | None]] = []
        last_rb: date | None = None
        category_done = [bool(meta_out and meta_out.get("strategy_category"))]
        freq_done = [bool(meta_out and meta_out.get("rebalance_frequency"))]
        meta_rows = 0
        row_it = ws.iter_rows(
            min_row=2,
            min_col=1,
            max_col=max_c,
            values_only=True,
        )
        for row in row_it:
            if not row:
                continue
            cells = tuple(row)
            if meta_out is not None and meta_rows < 3000:
                meta_rows += 1
                _strategy_excel_fill_meta_from_cells(
                    cells, col_index, meta_out, category_done=category_done, freq_done=freq_done
                )
            rb = _excel_cell_rebalance_date(cells[i_dt] if i_dt < len(cells) else None)
            raw_code = cells[i_code] if i_code < len(cells) else None
            if raw_code is None or (isinstance(raw_code, float) and pd.isna(raw_code)):
                continue
            if rb is None:
                rb = last_rb
            if rb is None:
                continue
            last_rb = rb
            try:
                code = normalize_code(raw_code)
            except ValueError:
                continue
            hw = _excel_cell_weight(cells[i_hw]) if i_hw is not None and i_hw < len(cells) else None
            iw = _excel_cell_weight(cells[i_iw]) if i_iw is not None and i_iw < len(cells) else None
            batch.append((rb, code, hw, iw))
            if len(batch) >= batch_size:
                yield batch
                batch = []
                gc.collect()
        if batch:
            yield batch
    finally:
        wb.close()


def _strategy_positions_max_rebalance_date(db: Session, sid: str) -> date | None:
    max_row = db.execute(
        text(
            f"""
            SELECT {sql_max_date_expr("rebalance_date")} AS d
            FROM strategy_positions
            WHERE strategy_id=:sid
            """
        ),
        {"sid": sid},
    ).mappings().first()
    return _row_sql_date(max_row["d"]) if max_row and max_row.get("d") else None


def _delete_strategy_positions_rebalance_period(
    db: Session, sid: str, rebalance: date
) -> None:
    rb_cmp = _compact_date(rebalance)
    db.execute(
        text(
            f"""
            DELETE FROM strategy_positions
            WHERE strategy_id=:sid
              AND {sql_date_compact_expr("rebalance_date")} = :rb_cmp
            """
        ),
        {"sid": sid, "rb_cmp": rb_cmp},
    )


def _mismatch_rebalance_dates(
    ex_counts: dict[str, int], db_counts: dict[str, int]
) -> set[str]:
    keys = set(ex_counts) | set(db_counts)
    return {d for d in keys if ex_counts.get(d, 0) != db_counts.get(d, 0)}


def _mismatch_rebalance_dates_in_excel_only(
    ex_counts: dict[str, int], db_counts: dict[str, int]
) -> set[str]:
    """增量切片：只比对 Excel 中出现的调仓日，不触碰/不要求库内其它日期。"""
    return {d for d in ex_counts if ex_counts.get(d, 0) != db_counts.get(d, 0)}


def _prepare_strategy_import_resume_from_db_max(
    db: Session, sid: str
) -> tuple[date | None, int]:
    """
    续传：查库内该策略 MAX(调仓日)，删除该日全部记录，再从 Excel 导入 >= 该日。
    返回 (cutoff_rb, 删后库内行数基线)。仅 1 次聚合查询 + 1 次 DELETE。
    """
    rows_before, max_rb = _strategy_positions_import_stats(db, sid)
    if max_rb is None:
        _log.info("import resume %s: 库内无持仓，从 Excel 首行导入", sid)
        return None, 0
    from app.db import run_under_turso_stream_lock

    def _del_last() -> None:
        _delete_strategy_positions_rebalance_period(db, sid, max_rb)
        db.commit()

    run_under_turso_stream_lock(_del_last)
    rows_after, max_after = _strategy_positions_import_stats(db, sid)
    _log.info(
        "import resume %s: 库内最新调仓日 %s（删前 %s 行）→ 已删该日全部记录；"
        "删后 %s 行 MAX=%s；从 Excel >=%s 重导",
        sid,
        max_rb.isoformat(),
        rows_before,
        rows_after,
        max_after.isoformat() if max_after else "—",
        max_rb.isoformat(),
    )
    return max_rb, rows_after


def _prepare_strategy_import_resume_incremental(
    db: Session, sid: str, file_path: str
) -> tuple[date | None, int]:
    """
    增量续传：仅删 Excel 文件内 MAX(调仓日) 在库中的该日记录，再从 Excel 写入 >= 该日。
    不删库内全局最新调仓日（避免动到非增量历史）。
    """
    resume_rb, ex_total = _excel_max_rebalance_date_and_total(file_path)
    if ex_total <= 0 or resume_rb is None:
        _log.info("import incremental resume %s: Excel 无有效行", sid)
        return None, _strategy_positions_row_count(db, sid)
    rows_before, db_max = _strategy_positions_import_stats(db, sid)
    from app.db import run_under_turso_stream_lock

    def _del_tail() -> None:
        _delete_strategy_positions_rebalance_period(db, sid, resume_rb)
        db.commit()

    run_under_turso_stream_lock(_del_tail)
    rows_after, max_after = _strategy_positions_import_stats(db, sid)
    _log.info(
        "import incremental resume %s: Excel 末调仓日 %s（删前 %s 行，库 MAX=%s）"
        "→ 已删该日全部记录；删后 %s 行 MAX=%s；从 Excel >=%s 重导",
        sid,
        resume_rb.isoformat(),
        rows_before,
        db_max.isoformat() if db_max else "—",
        rows_after,
        max_after.isoformat() if max_after else "—",
        resume_rb.isoformat(),
    )
    return resume_rb, rows_after


def _import_job_message_indicates_import_done(message: str | None) -> bool:
    """导入阶段已写完数据且失败数为 0，但 SUCCESS 终态尚未落库（如 SIGTERM 打断 finalize）。"""
    m = str(message or "")
    if "失败 0" in m and "已完成" in m:
        return True
    if "阶段1/3" in m and "导入已完成" in m and "失败 0" in m:
        return True
    return False


def _strategy_import_job_row_ready_to_succeed(row: dict[str, Any]) -> bool:
    """根据任务行判断是否可安全标 SUCCESS（避免数据已齐却标 FAILED）。"""
    ids = {str(x).strip() for x in _json_str_list(row.get("strategy_ids_json")) if str(x).strip()}
    done = {str(x).strip() for x in _json_str_list(row.get("completed_strategy_ids_json")) if str(x).strip()}
    if not ids or not ids <= done:
        return False
    ej = str(row.get("errors_json") or "[]").strip()
    if ej not in ("[]", "", "null"):
        return False
    if int(row.get("failed_count") or 0) > 0:
        return False
    msg = str(row.get("message") or "")
    if _import_job_message_indicates_import_done(msg):
        return True
    if "阶段1/3" in msg and "已完成" in msg and len(done) >= len(ids):
        return True
    return False


def _finalize_strategy_import_job(
    db: Session,
    job_id: int,
    ret: dict[str, Any],
    import_mode: str,
    strategy_ids: list[str],
) -> None:
    """根据 import_strategy_files 返回值与校验结果写入任务终态（SUCCESS/FAILED）。"""
    all_ids = {str(x).strip() for x in strategy_ids if str(x or "").strip()}
    done = set(ret.get("completed_strategy_ids") or [])
    verify_errors: list[str] = []
    row_checks: list[str] = []
    pre_verified = set(ret.get("verified_strategy_ids") or [])
    verify_rows = dict(ret.get("verify_rows") or {})
    if all_ids <= done and int(ret.get("failed") or 0) == 0:
        for sid in sorted(done):
            if sid in pre_verified:
                hint = verify_rows.get(sid)
                if hint:
                    row_checks.append(f"{sid}={hint}")
                else:
                    row_checks.append(f"{sid}={_strategy_positions_row_count(db, sid)}")
                continue
            try:
                fp = _strategy_config_excel_path(db, sid)
                db_n, ex_n = _verify_strategy_full_import_row_count(
                    db, sid, fp, import_mode=import_mode
                )
                if ex_n > 0:
                    row_checks.append(f"{sid}={db_n}/{ex_n}")
                else:
                    row_checks.append(f"{sid}={db_n}")
            except ValueError as ve:
                done.discard(sid)
                verify_errors.append(str(ve))
    for sid in sorted(all_ids):
        if sid not in done and not any(sid in e for e in verify_errors):
            cnt = _strategy_positions_row_count(db, sid)
            row_checks.append(f"{sid}={cnt}")
    errors_out = list(ret.get("errors") or []) + verify_errors
    ok = all_ids <= done and not errors_out
    st = "SUCCESS" if ok else "FAILED"
    rows_hint = "；".join(row_checks) if row_checks else ""
    msg = (
        f"完成：成功 {len(done)}/{len(all_ids)} 个策略（{rows_hint}）"
        if ok
        else f"未完成 {len(done)}/{len(all_ids)}，失败 {max(len(errors_out), int(ret.get('failed') or 0))}，可点「续传」或全量清库。{rows_hint}"
    )
    db.execute(
        text(
            f"""
            UPDATE strategy_import_jobs
            SET status=:st, finished_at={sql_now()}, message=:msg,
                imported_count=:ic, failed_count=:fc, errors_json=:ej,
                completed_strategy_ids_json=:cj
            WHERE id=:id AND status <> 'ABANDONED'
            """
        ),
        {
            "st": st,
            "msg": msg[:6000],
            "ic": len(done),
            "fc": max(len(errors_out), int(ret.get("failed") or 0)),
            "ej": json.dumps(errors_out[:50], ensure_ascii=False),
            "cj": json.dumps(sorted(done), ensure_ascii=False),
            "id": job_id,
        },
    )


def mark_running_strategy_import_jobs_interrupted(*, reason: str) -> None:
    """
    进程关闭前处理仍 RUNNING 的导入任务。
    若进度已显示「已完成 … 失败 0」而线程未及写 SUCCESS，补标成功而非误标失败。
    """
    from app.db import SessionLocalFactory

    if SessionLocalFactory is None:
        return
    db = SessionLocalFactory()
    try:
        rows = db.execute(
            text(
                """
                SELECT id, message, errors_json, strategy_ids_json, import_mode,
                       completed_strategy_ids_json, failed_count
                FROM strategy_import_jobs
                WHERE status='RUNNING'
                """
            )
        ).mappings().all()
        suffix_fail = f"（{reason}）"
        for row in rows:
            jid = int(row["id"])
            if not _strategy_import_job_row_ready_to_succeed(row):
                jid_mark = int(row["id"])
                db.execute(
                    text(
                        f"""
                        UPDATE strategy_import_jobs
                        SET status='FAILED', finished_at={sql_now()},
                            message=COALESCE(message, '') || :suf
                        WHERE id=:id AND status='RUNNING'
                        """
                    ),
                    {"suf": suffix_fail[:500], "id": jid_mark},
                )
                continue
            ids = _json_str_list(row.get("strategy_ids_json"))
            done = _json_str_list(row.get("completed_strategy_ids_json")) or ids
            try:
                _finalize_strategy_import_job(
                    db,
                    jid,
                    {
                        "completed_strategy_ids": sorted(set(done)),
                        "verified_strategy_ids": sorted(set(done)),
                        "verify_rows": {},
                        "failed": 0,
                        "errors": [],
                        "resumable": False,
                    },
                    str(row.get("import_mode") or "full"),
                    ids,
                )
                _log.info(
                    "strategy_import job %s: import done, promoted to SUCCESS on shutdown",
                    jid,
                )
            except Exception:
                _log.exception(
                    "strategy_import job %s: promote SUCCESS on shutdown failed",
                    jid,
                )
                db.execute(
                    text(
                        f"""
                        UPDATE strategy_import_jobs
                        SET status='FAILED', finished_at={sql_now()},
                            message=COALESCE(message, '') || :suf
                        WHERE id=:id AND status='RUNNING'
                        """
                    ),
                    {"suf": suffix_fail[:500], "id": jid},
                )
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def _import_after_batch_commit(db: Session) -> None:
    """每批落库后释放 ORM 身份映射，减轻长时流式导入内存爬升。"""
    try:
        db.expire_all()
    except Exception:
        pass
    gc.collect()


def _import_incremental_row_allowed(
    rebalance: date,
    cutoff: date | None,
    *,
    inclusive: bool,
) -> bool:
    if cutoff is None:
        return True
    return rebalance >= cutoff if inclusive else rebalance > cutoff


def _import_cutoff_plan(
    db: Session,
    sid: str,
    import_mode: str,
    *,
    repair_last_period: bool,
    file_path: str = "",
) -> tuple[date | None, bool, bool, int]:
    """
    返回 (cutoff_rb, cutoff_inclusive, delete_all, rows_baseline_after_cutoff)。
    首次全量：清空后整表重导。
    全量续传：库内 MAX(调仓日) 删该日全部记录，Excel 导入 >= 该日。
    首次增量：只追加 > MAX(rebalance_date)；库内更早调仓日不写不改。
    增量续传：删 Excel 末调仓日，从 Excel 导入 >= 该日（文件可仅含新增调仓）。
    """
    rows_baseline = 0
    mode = str(import_mode or "").lower()
    if mode == "full" and not repair_last_period:
        return None, False, True, rows_baseline
    cutoff_rb: date | None = None
    cutoff_inclusive = False
    if repair_last_period:
        if mode == "incremental" and file_path and os.path.isfile(file_path):
            cutoff_rb, rows_baseline = _prepare_strategy_import_resume_incremental(
                db, sid, file_path
            )
        else:
            cutoff_rb, rows_baseline = _prepare_strategy_import_resume_from_db_max(
                db, sid
            )
        cutoff_inclusive = cutoff_rb is not None
    elif mode == "incremental":
        cutoff_rb = _strategy_positions_max_rebalance_date(db, sid)
        rows_baseline = _strategy_positions_row_count(db, sid)
    return cutoff_rb, cutoff_inclusive, False, rows_baseline


def _strategy_positions_import_stats(db: Session, sid: str) -> tuple[int, date | None]:
    row = db.execute(
        text(
            f"""
            SELECT COUNT(*) AS n, {sql_max_date_expr("rebalance_date")} AS max_rb
            FROM strategy_positions WHERE strategy_id=:sid
            """
        ),
        {"sid": sid},
    ).mappings().first()
    if not row:
        return 0, None
    n = int(row.get("n") or 0)
    max_rb = _row_sql_date(row.get("max_rb"))
    return n, max_rb


def _strategy_positions_row_count(db: Session, sid: str) -> int:
    row = db.execute(
        text("SELECT COUNT(*) AS n FROM strategy_positions WHERE strategy_id=:sid"),
        {"sid": sid},
    ).mappings().first()
    return int(row.get("n") or 0) if row else 0


def _count_strategy_excel_data_rows(file_path: str) -> int:
    """统计 Excel 有效行数（与流式导入相同规则）。"""
    n = 0
    for chunk in _iter_strategy_holdings_excel_batches(file_path):
        n += len(chunk)
        chunk.clear()
    return n


def _excel_period_row_counts(file_path: str) -> tuple[dict[str, int], int]:
    """按调仓日统计 Excel 有效行数（与导入解析规则一致）。"""
    counts: dict[str, int] = defaultdict(int)
    total = 0
    for chunk in _iter_strategy_holdings_excel_batches(file_path):
        for rb, _code, _hw, _iw in chunk:
            k = rb.isoformat()
            counts[k] += 1
            total += 1
        chunk.clear()
    gc.collect()
    return dict(counts), total


def _strategy_import_heavy_row_threshold() -> int:
    return max(
        500,
        int(getattr(settings, "strategy_import_heavy_row_threshold", 8000) or 8000),
    )


def _import_is_heavy(*, excel_rows: int = 0, db_rows: int = 0) -> bool:
    th = _strategy_import_heavy_row_threshold()
    return int(excel_rows) >= th or int(db_rows) >= th


def _excel_max_rebalance_date_and_total(file_path: str) -> tuple[date | None, int]:
    """单次扫描：仅求末调仓日与总行数（增量续传删尾日，避免构建整表期次 dict）。"""
    max_rb: date | None = None
    total = 0
    for chunk in _iter_strategy_holdings_excel_batches(file_path):
        for rb, _code, _hw, _iw in chunk:
            total += 1
            if max_rb is None or rb > max_rb:
                max_rb = rb
        chunk.clear()
    gc.collect()
    return max_rb, total


def _excel_period_counts_from_stats(
    stats: dict[str, int | str] | None,
) -> tuple[dict[str, int], int] | None:
    if not stats:
        return None
    raw = stats.get("excel_period_counts")
    if not isinstance(raw, dict) or not raw:
        return None
    ex_counts = {str(k): int(v) for k, v in raw.items()}
    ex_total = int(stats.get("expected_file_rows") or 0) or sum(ex_counts.values())
    return ex_counts, ex_total


def _accumulate_excel_period_counts(
    stats: dict[str, int | str],
    chunk: list[tuple[date, str, float | None, float | None]],
) -> None:
    """导入扫描时累计各调仓日行数，供结束校验复用，避免再整文件扫一遍。"""
    if not chunk:
        return
    pc = stats.get("excel_period_counts")
    if pc is None:
        pc = defaultdict(int)
        stats["excel_period_counts"] = pc
    elif not isinstance(pc, defaultdict):
        pc = defaultdict(int, {str(k): int(v) for k, v in pc.items()})
        stats["excel_period_counts"] = pc
    for rb, _code, _hw, _iw in chunk:
        pc[rb.isoformat()] += 1
    stats["expected_file_rows"] = int(stats.get("expected_file_rows") or 0) + len(chunk)


def _finalize_excel_period_counts_in_stats(stats: dict[str, int | str]) -> None:
    pc = stats.get("excel_period_counts")
    if isinstance(pc, defaultdict):
        stats["excel_period_counts"] = dict(pc)


def _db_period_row_counts_for_dates(
    db: Session, sid: str, dates: list[str]
) -> dict[str, int]:
    """仅查询指定调仓日的行数（大表校验切片时用，避免多余逻辑）。"""
    uniq = sorted({str(d) for d in dates if str(d).strip()})
    if not uniq:
        return {}
    counts: dict[str, int] = {d: 0 for d in uniq}
    batch = 60
    for i in range(0, len(uniq), batch):
        part = uniq[i : i + batch]
        cmps = [_compact_date(date.fromisoformat(d)) for d in part]
        in_list = ", ".join(f":c{j}" for j in range(len(cmps)))
        params: dict[str, Any] = {"sid": sid}
        for j, c in enumerate(cmps):
            params[f"c{j}"] = c
        rows = db.execute(
            text(
                f"""
                SELECT rebalance_date AS rb, COUNT(*) AS cnt
                FROM strategy_positions
                WHERE strategy_id=:sid
                  AND {sql_date_compact_expr("rebalance_date")} IN ({in_list})
                GROUP BY rebalance_date
                """
            ),
            params,
        ).mappings().all()
        for r in rows:
            d = _row_sql_date(r.get("rb"))
            if d:
                counts[d.isoformat()] = int(r.get("cnt") or 0)
    return counts


def _db_period_row_counts(db: Session, sid: str) -> tuple[dict[str, int], int]:
    rows = db.execute(
        text(
            """
            SELECT rebalance_date AS rb, COUNT(*) AS cnt
            FROM strategy_positions
            WHERE strategy_id=:sid
            GROUP BY rebalance_date
            """
        ),
        {"sid": sid},
    ).mappings().all()
    counts: dict[str, int] = {}
    total = 0
    for r in rows:
        d = _row_sql_date(r.get("rb"))
        if not d:
            continue
        k = d.isoformat()
        c = int(r.get("cnt") or 0)
        counts[k] = c
        total += c
    return counts, total


def _verify_strategy_positions_exact_match(
    db: Session,
    sid: str,
    file_path: str,
    *,
    ex_counts: dict[str, int] | None = None,
    ex_total: int | None = None,
) -> tuple[int, int]:
    """
    全量导入终态校验：总行数与每个调仓日行数均须与 Excel 完全一致（不多不少）。
    返回 (db_rows, excel_rows)；不达标抛错，禁止标 SUCCESS。
    """
    if not os.path.isfile(file_path):
        raise ValueError(f"{sid} Excel 不存在: {file_path}")
    if ex_counts is None or ex_total is None:
        ex_counts, ex_total = _excel_period_row_counts(file_path)
    else:
        ex_total = int(ex_total)
    if ex_total <= 0:
        raise ValueError(f"{sid} Excel 无有效持仓行")
    db_total = _strategy_positions_row_count(db, sid)
    if db_total != ex_total:
        raise ValueError(
            f"{sid} 行数不一致：Excel {ex_total} 行，库内 {db_total} 行（差 {db_total - ex_total:+d}）。"
            f"请全量导入并传 confirm_wipe:true 清库后重导。"
        )
    db_counts = _db_period_row_counts_for_dates(db, sid, list(ex_counts.keys()))
    bad: list[tuple[str, int, int]] = []
    for d in sorted(ex_counts):
        ec = ex_counts[d]
        dc = db_counts.get(d, 0)
        if ec != dc:
            bad.append((d, ec, dc))
    if bad:
        d0, e0, b0 = bad[0]
        extra = f"；另有 {len(bad) - 1} 个调仓日" if len(bad) > 1 else ""
        raise ValueError(
            f"{sid} 调仓日与 Excel 不一致：共 {len(bad)} 期{extra}。"
            f"首期差异 {d0}（Excel {e0} / 库内 {b0}）。"
            f"自动补写后仍不一致，请再点「续传」或 confirm_wipe:true 全量重导。"
        )
    return db_total, ex_total


def _verify_strategy_positions_incremental_match(
    db: Session,
    sid: str,
    file_path: str,
    *,
    ex_counts: dict[str, int] | None = None,
    ex_total: int | None = None,
) -> tuple[int, int]:
    """
    增量终态校验：Excel 文件中每个调仓日行数须与库内一致；不要求库总行数=Excel 总行数。
    返回 (库内该切片行数, Excel 行数)。
    """
    if not os.path.isfile(file_path):
        raise ValueError(f"{sid} Excel 不存在: {file_path}")
    if ex_counts is None or ex_total is None:
        ex_counts, ex_total = _excel_period_row_counts(file_path)
    else:
        ex_total = int(ex_total)
    if ex_total <= 0:
        raise ValueError(f"{sid} Excel 无有效持仓行")
    db_counts = _db_period_row_counts_for_dates(db, sid, list(ex_counts.keys()))
    bad: list[tuple[str, int, int]] = []
    slice_db = 0
    for d in sorted(ex_counts):
        ec = ex_counts[d]
        dc = db_counts.get(d, 0)
        slice_db += dc
        if ec != dc:
            bad.append((d, ec, dc))
    if bad:
        d0, e0, b0 = bad[0]
        extra = f"；另有 {len(bad) - 1} 个调仓日" if len(bad) > 1 else ""
        raise ValueError(
            f"{sid} 增量切片与 Excel 不一致：共 {len(bad)} 期{extra}。"
            f"首期差异 {d0}（Excel {e0} / 库内 {b0}）。"
            f"请点「续传」或检查增量文件是否仅含新增调仓。"
        )
    return slice_db, ex_total


def _verify_strategy_positions_incremental_match_with_auto_repair(
    db: Session,
    sid: str,
    file_path: str,
    *,
    ex_counts: dict[str, int] | None = None,
    ex_total: int | None = None,
    max_repair_rounds: int = 3,
    heavy: bool = False,
) -> tuple[int, int]:
    last_err: ValueError | None = None
    rounds = 1 if heavy else max(1, int(max_repair_rounds or 3))
    for rnd in range(rounds):
        try:
            return _verify_strategy_positions_incremental_match(
                db,
                sid,
                file_path,
                ex_counts=ex_counts,
                ex_total=ex_total,
            )
        except ValueError as ve:
            last_err = ve
            if ex_counts is None or ex_total is None:
                ex_counts, ex_total = _excel_period_row_counts(file_path)
            db_counts = _db_period_row_counts_for_dates(db, sid, list(ex_counts.keys()))
            bad = _mismatch_rebalance_dates_in_excel_only(ex_counts, db_counts)
            if not bad:
                raise
            from app.db import run_under_turso_stream_lock

            def _repair_once() -> int:
                return _repair_strategy_positions_against_excel(
                    db,
                    sid,
                    file_path,
                    bad_dates=bad,
                    ex_counts=ex_counts,
                )

            n = int(run_under_turso_stream_lock(_repair_once) or 0)
            if n <= 0:
                raise
            _log.info(
                "import incremental %s: 切片校验第 %s 轮不一致，已补写 %s 行",
                sid,
                rnd + 1,
                n,
            )
    if last_err is not None:
        raise last_err
    raise ValueError(f"{sid} 增量导入校验失败")


def _verify_strategy_positions_exact_match_with_auto_repair(
    db: Session,
    sid: str,
    file_path: str,
    *,
    ex_counts: dict[str, int] | None = None,
    ex_total: int | None = None,
    max_repair_rounds: int = 3,
    heavy: bool = False,
) -> tuple[int, int]:
    """终态校验；不一致时按调仓日自动补写后重试，避免中断残留导致误失败。"""
    last_err: ValueError | None = None
    rounds = 1 if heavy else max(1, int(max_repair_rounds or 3))
    for rnd in range(rounds):
        try:
            return _verify_strategy_positions_exact_match(
                db,
                sid,
                file_path,
                ex_counts=ex_counts,
                ex_total=ex_total,
            )
        except ValueError as ve:
            last_err = ve
            if ex_counts is None or ex_total is None:
                ex_counts, ex_total = _excel_period_row_counts(file_path)
            db_counts = _db_period_row_counts_for_dates(db, sid, list(ex_counts.keys()))
            bad = _mismatch_rebalance_dates_in_excel_only(ex_counts, db_counts)
            if not bad:
                bad = _mismatch_rebalance_dates(ex_counts, db_counts)
            from app.db import run_under_turso_stream_lock

            def _repair_once() -> int:
                return _repair_strategy_positions_against_excel(
                    db,
                    sid,
                    file_path,
                    bad_dates=bad,
                    ex_counts=ex_counts,
                )

            n = int(run_under_turso_stream_lock(_repair_once) or 0)
            if n <= 0:
                raise
            _log.info(
                "import %s: 校验第 %s 轮不一致，已自动补写 %s 行后重试",
                sid,
                rnd + 1,
                n,
            )
    if last_err is not None:
        raise last_err
    raise ValueError(f"{sid} 导入校验失败")


def _heal_positions_gaps_from_excel(
    db: Session,
    sid: str,
    file_path: str,
    *,
    stats: dict[str, int | str] | None = None,
    sync_job_id: int | None = None,
    strategy_import_job_id: int | None = None,
    incremental_slice: bool = False,
    ex_counts: dict[str, int] | None = None,
    ex_total: int | None = None,
) -> int:
    """
    对比 Excel/库内各调仓日行数，对不一致的日期整期 UPSERT（可修中间残缺，不依赖 confirm_wipe）。
    incremental_slice=True 时仅处理 Excel 中出现的调仓日。
    大行数时跳过续传前 heal，改由删尾重导 + 结束校验补写，避免多遍读 Excel OOM。
    """
    from app.db import run_under_turso_stream_lock

    db_rows = _strategy_positions_row_count(db, sid)
    cached = _excel_period_counts_from_stats(stats) if stats else None
    if cached:
        ex_counts, ex_total = cached
    elif ex_counts is None:
        if _import_is_heavy(excel_rows=0, db_rows=db_rows):
            _log.info(
                "import %s: skip pre-heal (db %s rows >= threshold %s)",
                sid,
                db_rows,
                _strategy_import_heavy_row_threshold(),
            )
            return 0
        ex_counts, ex_total = _excel_period_row_counts(file_path)
    else:
        ex_total = int(ex_total or 0) or sum(ex_counts.values())
    if _import_is_heavy(excel_rows=ex_total, db_rows=db_rows):
        if stats is not None and ex_total > 0:
            stats["expected_file_rows"] = ex_total
            if ex_counts:
                stats["excel_period_counts"] = ex_counts
        _log.info(
            "import %s: skip pre-heal (%s excel / %s db rows >= threshold %s)",
            sid,
            ex_total,
            db_rows,
            _strategy_import_heavy_row_threshold(),
        )
        return 0
    if stats is not None and ex_total > 0:
        stats["expected_file_rows"] = ex_total
        if ex_counts:
            stats["excel_period_counts"] = ex_counts
    if ex_total <= 0:
        return 0
    db_counts = _db_period_row_counts_for_dates(db, sid, list(ex_counts.keys()))
    if incremental_slice:
        bad = _mismatch_rebalance_dates_in_excel_only(ex_counts, db_counts)
    else:
        bad = _mismatch_rebalance_dates(ex_counts, db_counts)
    if not bad:
        return 0

    def _do_heal() -> int:
        return _repair_strategy_positions_against_excel(
            db, sid, file_path, bad_dates=bad, ex_counts=ex_counts
        )

    n = int(run_under_turso_stream_lock(_do_heal) or 0)
    if n > 0:
        msg = (
            f"阶段1/3 {sid}：检测到 {len(bad)} 个调仓日与 Excel 不一致，"
            f"已自动补写 {n} 行…"
        )
        _import_progress_touch(
            db,
            msg,
            sync_job_id=sync_job_id,
            strategy_import_job_id=strategy_import_job_id,
            do_commit=True,
        )
        _log.info("import %s: pre-import heal %s periods, %s rows", sid, len(bad), n)
    return n


def _verify_strategy_full_import_row_count(
    db: Session,
    sid: str,
    file_path: str,
    *,
    import_mode: str,
    ex_counts: dict[str, int] | None = None,
    ex_total: int | None = None,
    heavy: bool = False,
) -> tuple[int, int]:
    """全量：整表与 Excel 一致；增量：仅 Excel 切片逐期一致。"""
    mode = str(import_mode or "").lower()
    kw = {
        "ex_counts": ex_counts,
        "ex_total": ex_total,
        "heavy": heavy,
    }
    if mode == "full":
        return _verify_strategy_positions_exact_match_with_auto_repair(
            db, sid, file_path, **kw
        )
    return _verify_strategy_positions_incremental_match_with_auto_repair(
        db, sid, file_path, **kw
    )


def _import_verify_positions_after_write(
    db: Session,
    sid: str,
    file_path: str,
    import_mode: str,
    stats: dict[str, int | str],
) -> tuple[int, int]:
    """导入写入结束后校验；大行数复用 stats 内 excel_period_counts，避免再扫 Excel。"""
    cached = _excel_period_counts_from_stats(stats)
    ex_counts = cached[0] if cached else None
    ex_total = cached[1] if cached else None
    heavy = _import_is_heavy(
        excel_rows=int(ex_total or stats.get("excel_rows") or 0),
        db_rows=int(stats.get("rows_after") or 0),
    )
    db_n, ex_n = _verify_strategy_full_import_row_count(
        db,
        sid,
        file_path,
        import_mode=import_mode,
        ex_counts=ex_counts,
        ex_total=ex_total,
        heavy=heavy,
    )
    stats["verify_slice_db"] = db_n
    stats["verify_slice_ex"] = ex_n
    return db_n, ex_n


def _strategy_config_excel_path(db: Session, sid: str) -> str:
    row = db.execute(
        text(
            "SELECT file_dir, file_name FROM strategy_configs WHERE strategy_id=:sid LIMIT 1"
        ),
        {"sid": sid},
    ).mappings().first()
    if not row:
        raise ValueError(f"{sid} 配置不存在")
    return _strategy_excel_path(row.get("file_dir"), row["file_name"])


def _validate_strategy_import_result(
    sid: str,
    stats: dict[str, int],
    *,
    delete_all: bool,
    repair_last_period: bool,
    import_mode: str,
) -> None:
    """导入结束校验：以库内行数为准；续传完成时须接近 Excel 文件总行数（非本次扫描累计）。"""
    rows_before = int(stats.get("rows_before") or 0)
    rows_after = int(stats.get("rows_after") or 0)
    imported = int(stats.get("imported_rows") or 0)
    excel_rows = int(stats.get("excel_rows") or 0)
    expected_total = int(stats.get("expected_file_rows") or 0) or excel_rows
    mode = str(import_mode or "").lower()
    need_complete = delete_all or (repair_last_period and mode == "full")
    if need_complete and expected_total > 0:
        if rows_after != expected_total:
            gap = expected_total - rows_after
            scanned_all = excel_rows >= expected_total
            if repair_last_period and scanned_all and abs(gap) > max(50, int(expected_total * 0.01)):
                raise ValueError(
                    f"{sid} 已扫完 Excel 但库内 {rows_after} 行 ≠ Excel {expected_total} 行，"
                    f"中间多期不完整，续传无法补全；请 confirm_wipe:true 全量清库重导。"
                )
            raise ValueError(
                f"{sid} 行数不一致：Excel {expected_total} 行，库内 {rows_after} 行（差 {gap:+d}）。"
                f"{'进程可能在 Render 重启前未写完，请再点「续传」。' if not scanned_all else '请 confirm_wipe:true 全量清库重导。'}"
            )
    if delete_all:
        if rows_after <= 0:
            raise ValueError(f"{sid} 全量导入后库内 0 行")
        return
    if repair_last_period:
        if mode == "incremental":
            if rows_after + 200 < rows_before:
                raise ValueError(
                    f"{sid} 增量续传后总行数异常减少 {rows_before} → {rows_after}，"
                    f"请再点「续传」；历史调仓日应未被改动"
                )
            return
        exp = int(stats.get("expected_file_rows") or 0) or excel_rows
        if rows_before > 0 and imported <= 0 and exp > 0 and rows_after != exp:
            raise ValueError(
                f"{sid} 续传未写入新行且库内 {rows_after} 行 ≠ Excel {exp} 行，请全量清库重导"
            )
        if rows_after + 200 < rows_before:
            raise ValueError(
                f"{sid} 续传后行数异常减少 {rows_before} → {rows_after}，请再点「续传」或全量重来"
            )


def _import_strategy_holdings_from_excel(
    db: Session,
    sid: str,
    file_path: str,
    import_mode: str,
    *,
    sync_job_id: int | None = None,
    strategy_import_job_id: int | None = None,
    repair_last_period: bool = False,
) -> tuple[dict[str, str], dict[str, int]]:
    """导入单策略 Excel 持仓；大文件走流式。返回 (label_meta, import_stats)。"""
    label_meta: dict[str, str] = {}
    stats: dict[str, int | str] = {
        "rows_before": 0,
        "excel_rows": 0,
        "imported_rows": 0,
        "skipped_rows": 0,
        "rows_after": 0,
    }
    delete_all = False
    if _strategy_excel_use_streaming(file_path):
        resume_cutoff: date | None = None
        cutoff_rb: date | None = None
        cutoff_inclusive = False
        delete_all = False
        rows_baseline = 0
        inc_slice = str(import_mode or "").lower() == "incremental"
        if repair_last_period:
            healed_pre = _heal_positions_gaps_from_excel(
                db,
                sid,
                file_path,
                stats=stats,
                sync_job_id=sync_job_id,
                strategy_import_job_id=strategy_import_job_id,
                incremental_slice=inc_slice,
            )
            if healed_pre:
                stats["healed_rows"] = int(stats.get("healed_rows") or 0) + healed_pre
            cutoff_rb, cutoff_inclusive, delete_all, rows_baseline = _import_cutoff_plan(
                db,
                sid,
                import_mode,
                repair_last_period=True,
                file_path=file_path,
            )
            resume_cutoff = cutoff_rb
            stats["rows_before"] = rows_baseline
            resume_from = resume_cutoff
            if inc_slice:
                resume_msg = (
                    f"阶段1/3 {sid}：增量续传 — 已补写 Excel 内不一致调仓日"
                    f"{f'（{healed_pre} 行）' if healed_pre else ''}；"
                    f"已删 Excel 末调仓日 {resume_from.isoformat() if resume_from else '（无）'}"
                    f" 全部记录，从 Excel 写入 >= 该日（库内更早调仓日不动）；"
                    f"删后基线 {stats['rows_before']} 行…"
                )
            else:
                resume_msg = (
                    f"阶段1/3 {sid}：续传 — 已自动补写残缺调仓日"
                    f"{f'（{healed_pre} 行）' if healed_pre else ''}；"
                    f"库内最新调仓日 {resume_from.isoformat() if resume_from else '（无）'}"
                    f"，已删该日全部记录，从 Excel 写入 >= 该日；"
                    f"删后基线 {stats['rows_before']} 行…"
                )
            _import_progress_touch(
                db,
                resume_msg,
                sync_job_id=sync_job_id,
                strategy_import_job_id=strategy_import_job_id,
                do_commit=False,
            )
        else:
            cutoff_rb, cutoff_inclusive, delete_all, rows_baseline = _import_cutoff_plan(
                db,
                sid,
                import_mode,
                repair_last_period=False,
                file_path=file_path,
            )
            if inc_slice:
                stats["rows_before"] = rows_baseline
        if delete_all:
            stage_token = _strategy_import_stage_token(
                sid, strategy_import_job_id, sync_job_id
            )
            _import_progress_touch(
                db,
                f"阶段1/3 {sid}：全量首次导入，清空旧持仓后从 Excel 首行写入…",
                sync_job_id=sync_job_id,
                strategy_import_job_id=strategy_import_job_id,
                do_commit=False,
            )
            from app.db import run_under_turso_stream_lock

            def _wipe() -> None:
                _cleanup_strategy_import_stage(db, sid)
                db.commit()

            run_under_turso_stream_lock(_wipe)
            _import_after_batch_commit(db)
            stats["rows_before"] = 0
        else:
            stage_token = None
        imported_rows = 0
        batch_no = 0
        skipped_rows = 0
        scan_batch_no = 0
        excel_rows = 0
        # Pre-heal may populate Excel counts before the real import scan. Rebuild
        # them from this pass so resume/full validation does not double-count.
        stats["expected_file_rows"] = 0
        stats["excel_period_counts"] = defaultdict(int)
        _log_runtime_progress(
            f"import:{sid}",
            f"import {sid}: streaming start mode={import_mode} file={Path(file_path).name}",
            force=True,
        )
        for chunk in _iter_strategy_holdings_excel_batches(file_path, meta_out=label_meta):
            if batch_no % 8 == 0:
                raise_if_shutting_down()
            excel_rows += len(chunk)
            _accumulate_excel_period_counts(stats, chunk)
            if cutoff_rb is not None:
                before = len(chunk)
                chunk = [
                    x
                    for x in chunk
                    if _import_incremental_row_allowed(
                        x[0], cutoff_rb, inclusive=cutoff_inclusive
                    )
                ]
                skipped_rows += before - len(chunk)
            if not chunk:
                scan_batch_no += 1
                if scan_batch_no % 48 == 0 and cutoff_rb is not None:
                    est = int(stats.get("rows_before") or 0) + imported_rows
                    _import_progress_touch(
                        db,
                        f"阶段1/3 {sid}：续传扫描 Excel… 已跳过 {skipped_rows} 行"
                        f"{f'，本段已 commit {imported_rows} 行' if imported_rows else ''}；"
                        f"库内约 {est} 行（按批累计，非全表 COUNT）…",
                        sync_job_id=sync_job_id,
                        strategy_import_job_id=strategy_import_job_id,
                        do_commit=False,
                    )
                continue
            batch_no += 1
            _commit_positions_chunk_resilient(
                db,
                sid,
                chunk,
                batch_no=batch_no,
                stage_token=stage_token,
            )
            imported_rows += len(chunk)
            _import_after_batch_commit(db)
            batch_max_rb = max((x[0] for x in chunk), default=None)
            if batch_max_rb is not None:
                stats["last_committed_rebalance_date"] = batch_max_rb.isoformat()
            stats["imported_rows"] = imported_rows
            stats["skipped_rows"] = skipped_rows
            stats["excel_rows"] = excel_rows
            if stage_token:
                stats["rows_after"] = _strategy_import_stage_row_count(
                    db, sid, stage_token
                )
            else:
                stats["rows_after"] = int(stats.get("rows_before") or 0) + imported_rows
            _log_runtime_progress(
                f"import:{sid}",
                (
                    f"import {sid}: batch={batch_no} scanned={excel_rows} "
                    f"imported={imported_rows} skipped={skipped_rows} "
                    f"est_rows={stats['rows_after']}"
                ),
            )
            prog_iv = _strategy_import_progress_interval()
            if batch_no % prog_iv == 0:
                from app.db import run_under_turso_stream_lock

                lrb = stats.get("last_committed_rebalance_date")

                def _touch_progress() -> None:
                    if strategy_import_job_id is not None:
                        _strategy_import_job_save_checkpoint(
                            db,
                            int(strategy_import_job_id),
                            sid,
                            stats=stats,
                            batch_no=batch_no,
                            do_commit=False,
                        )
                    _import_progress_touch(
                        db,
                        f"阶段1/3 {sid}：第 {batch_no} 批已 commit +{len(chunk)} 行"
                        f"，累计约 {stats['rows_after']} 行"
                        f"{f'，末调仓日 {lrb}' if lrb else ''}"
                        f"{f'（Excel已扫 {excel_rows}，跳过 {skipped_rows}）' if cutoff_rb else ''}…",
                        sync_job_id=sync_job_id,
                        strategy_import_job_id=strategy_import_job_id,
                        do_commit=True,
                    )

                run_under_turso_stream_lock(_touch_progress)
            chunk.clear()
        stats["excel_rows"] = excel_rows
        stats["imported_rows"] = imported_rows
        stats["skipped_rows"] = skipped_rows
        _finalize_excel_period_counts_in_stats(stats)
        heavy_import = _import_is_heavy(
            excel_rows=excel_rows,
            db_rows=int(stats.get("rows_before") or 0) + imported_rows,
        )
        from app.db import run_under_turso_stream_lock

        if stage_token:

            def _finish_stage_import() -> None:
                cached = _excel_period_counts_from_stats(stats)
                if not cached:
                    raise ValueError(f"{sid} staging 缺少 Excel 期数缓存，无法安全替换")
                ex_counts, ex_total = cached
                promoted = _verify_strategy_import_stage_exact(
                    db, sid, stage_token, ex_counts, ex_total
                )
                stats["rows_after"] = _promote_strategy_import_stage(
                    db, sid, stage_token
                )
                if int(stats["rows_after"]) != promoted:
                    raise ValueError(
                        f"{sid} staging 替换异常：暂存 {promoted} 行，正式 {stats['rows_after']} 行"
                    )

            run_under_turso_stream_lock(_finish_stage_import)
        elif str(import_mode or "").lower() == "full" and not heavy_import:

            def _finish_import() -> None:
                nonlocal imported_rows
                cached = _excel_period_counts_from_stats(stats)
                repaired = _repair_strategy_positions_against_excel(
                    db,
                    sid,
                    file_path,
                    ex_counts=cached[0] if cached else None,
                )
                if repaired:
                    stats["repaired_rows"] = int(stats.get("repaired_rows") or 0) + repaired
                    imported_rows += repaired
                    stats["imported_rows"] = imported_rows
                stats["rows_after"] = _strategy_positions_row_count(db, sid)

            run_under_turso_stream_lock(_finish_import)
        else:
            stats["rows_after"] = _strategy_positions_row_count(db, sid)
        if repair_last_period and import_mode == "full":
            exp = int(stats.get("expected_file_rows") or 0)
            _log.info(
                "import resume %s: 结束 库内=%s 行，Excel 文件约 %s 行",
                sid,
                stats["rows_after"],
                exp or "?",
            )
        if imported_rows == 0 and import_mode == "full" and not repair_last_period:
            raise ValueError("无有效行（请检查「调整日期」「证券代码」是否为空）")
        _validate_strategy_import_result(
            sid,
            stats,
            delete_all=delete_all,
            repair_last_period=repair_last_period,
            import_mode=import_mode,
        )
        _import_verify_positions_after_write(
            db, sid, file_path, import_mode, stats
        )
        stats["positions_verified"] = 1
        _log.info(
            "import_strategy %s streaming done excel=%s imported=%s db=%s skipped=%s mode=%s repair=%s",
            sid,
            excel_rows,
            imported_rows,
            stats["rows_after"],
            skipped_rows,
            import_mode,
            repair_last_period,
        )
        _log_runtime_progress(
            f"import:{sid}",
            (
                f"import {sid}: done excel={excel_rows} imported={imported_rows} "
                f"db={stats['rows_after']} skipped={skipped_rows}"
            ),
            force=True,
        )
    else:
        # .xls 等：pandas 仅读必要列；行数过多时仍可能 OOM，建议另存为 .xlsx
        _import_progress_touch(
            db,
            f"阶段1/3 {sid}：读取 Excel（前 {_strategy_excel_read_max_col()} 列）…",
            sync_job_id=sync_job_id,
            strategy_import_job_id=strategy_import_job_id,
            do_commit=False,
        )
        label_meta = _read_strategy_excel_label_meta(file_path)
        label_meta2, rows_to_write = _read_strategy_holdings_excel(file_path)
        _import_progress_touch(
            db,
            f"阶段1/3 {sid}：已解析 {len(rows_to_write)} 行，写入数据库…",
            sync_job_id=sync_job_id,
            strategy_import_job_id=strategy_import_job_id,
            do_commit=False,
        )
        label_meta.update({k: v for k, v in (label_meta2 or {}).items() if v})
        stats["excel_rows"] = len(rows_to_write)
        pc: dict[str, int] = defaultdict(int)
        for rb, _code, _hw, _iw in rows_to_write:
            pc[rb.isoformat()] += 1
        stats["excel_period_counts"] = dict(pc)
        stats["expected_file_rows"] = len(rows_to_write)
        delete_all = import_mode == "full" and not repair_last_period
        _import_write_positions_one_strategy(
            db,
            sid,
            rows_to_write,
            label_meta,
            import_mode,
            repair_last_period=repair_last_period,
            file_path=file_path,
        )
        stats["rows_after"] = _strategy_positions_row_count(db, sid)
        stats["imported_rows"] = stats["rows_after"] if delete_all else max(
            0, stats["rows_after"] - stats["rows_before"]
        )
        _validate_strategy_import_result(
            sid,
            stats,
            delete_all=delete_all,
            repair_last_period=repair_last_period,
            import_mode=import_mode,
        )
        _import_verify_positions_after_write(
            db, sid, file_path, import_mode, stats
        )
        stats["positions_verified"] = 1
        rows_to_write.clear()
        del rows_to_write
        gc.collect()
        return label_meta, stats
    if label_meta:
        keys = [k for k, v in label_meta.items() if str(v or "").strip()]
        if keys:
            sets = ", ".join(f"{k}=:{k}" for k in keys)
            params = {k: label_meta[k] for k in keys}
            params["sid"] = sid
            db.execute(
                text(f"UPDATE strategy_configs SET {sets} WHERE strategy_id=:sid"),
                params,
            )
    return label_meta, stats

_POSITION_UPSERT_SQL = text(
    """
    INSERT INTO strategy_positions
    (strategy_id, rebalance_date, stock_code, holding_weight, industry_neutral_weight)
    VALUES (:sid, :rdate, :scode, :hw, :iw)
    ON CONFLICT(strategy_id, rebalance_date, stock_code) DO UPDATE SET
      holding_weight=excluded.holding_weight,
      industry_neutral_weight=excluded.industry_neutral_weight
    """
)

_POSITION_STAGE_UPSERT_SQL = text(
    """
    INSERT INTO strategy_positions_import_stage
    (import_token, strategy_id, rebalance_date, stock_code, holding_weight, industry_neutral_weight)
    VALUES (:token, :sid, :rdate, :scode, :hw, :iw)
    ON CONFLICT(import_token, strategy_id, rebalance_date, stock_code) DO UPDATE SET
      holding_weight=excluded.holding_weight,
      industry_neutral_weight=excluded.industry_neutral_weight
    """
)


def _read_strategy_holdings_excel(
    file_path: str,
) -> tuple[dict[str, str], list[tuple[date, str, float | None, float | None]]]:
    """
    只读前 N 列（默认 A～E），避免宽表/格式列撑爆内存；解析后释放 DataFrame。
    """
    max_c = _strategy_excel_read_max_col()
    usecols = _strategy_excel_pandas_usecols()
    header = pd.read_excel(file_path, sheet_name=0, nrows=0, usecols=usecols)
    all_cols = [str(c).strip() for c in header.columns]
    for need in _STRATEGY_EXCEL_REQUIRED_COLS:
        if need not in all_cols:
            raise ValueError(f"缺少列: {need}（须在前 {max_c} 列内）")
    read_kw: dict[str, Any] = {"sheet_name": 0, "dtype": object, "usecols": usecols}
    try:
        df = pd.read_excel(file_path, engine="calamine", **read_kw)
    except Exception:
        df = pd.read_excel(file_path, **read_kw)
    label_meta = _excel_meta_strategy_labels(df)
    dt = pd.to_datetime(df["调整日期"], errors="coerce").ffill().bfill()
    valid = dt.notna() & df["证券代码"].notna()
    if not bool(valid.any()):
        raise ValueError("无有效行（请检查「调整日期」「证券代码」是否为空）")
    sub = df.loc[valid]
    rows: list[tuple[date, str, float | None, float | None]] = []
    has_hw = "持仓权重" in sub.columns
    has_iw = "行业中性权重" in sub.columns
    for idx in sub.index:
        ts = dt.loc[idx]
        rebalance = ts.date() if hasattr(ts, "date") else pd.Timestamp(ts).date()
        code = normalize_code(sub.at[idx, "证券代码"])
        holding = None
        if has_hw:
            v = sub.at[idx, "持仓权重"]
            if pd.notna(v):
                holding = float(v)
        industry_w = None
        if has_iw:
            v = sub.at[idx, "行业中性权重"]
            if pd.notna(v):
                industry_w = float(v)
        rows.append((rebalance, code, holding, industry_w))
    del df, sub, dt, valid, header
    return label_meta, rows


def _import_positions_batch(
    db: Session,
    sid: str,
    write_rows: list[tuple[date, str, float | None, float | None]],
) -> None:
    batch = _strategy_import_position_batch_size()
    for i in range(0, len(write_rows), batch):
        chunk = write_rows[i : i + batch]
        db.execute(
            _POSITION_UPSERT_SQL,
            [
                {
                    "sid": sid,
                    "rdate": rebalance,
                    "scode": code,
                    "hw": holding,
                    "iw": industry_w,
                }
                for rebalance, code, holding, industry_w in chunk
            ],
        )


def _import_positions_stage_batch(
    db: Session,
    token: str,
    sid: str,
    write_rows: list[tuple[date, str, float | None, float | None]],
) -> None:
    batch = min(_strategy_import_position_batch_size(), 60)
    for i in range(0, len(write_rows), batch):
        chunk = write_rows[i : i + batch]
        values: list[str] = []
        params: dict[str, Any] = {}
        for j, (rebalance, code, holding, industry_w) in enumerate(chunk):
            values.append(
                f"(:token{j}, :sid{j}, :rdate{j}, :scode{j}, :hw{j}, :iw{j})"
            )
            params[f"token{j}"] = token
            params[f"sid{j}"] = sid
            params[f"rdate{j}"] = rebalance
            params[f"scode{j}"] = code
            params[f"hw{j}"] = holding
            params[f"iw{j}"] = industry_w
        db.execute(
            text(
                f"""
                INSERT INTO strategy_positions_import_stage
                (import_token, strategy_id, rebalance_date, stock_code,
                 holding_weight, industry_neutral_weight)
                VALUES {", ".join(values)}
                ON CONFLICT(import_token, strategy_id, rebalance_date, stock_code)
                DO UPDATE SET
                  holding_weight=excluded.holding_weight,
                  industry_neutral_weight=excluded.industry_neutral_weight
                """
            ),
            params,
        )


def _strategy_import_stage_token(
    sid: str, strategy_import_job_id: int | None, sync_job_id: int | None
) -> str:
    base = strategy_import_job_id if strategy_import_job_id is not None else f"sync{sync_job_id or 0}"
    return f"{sid}:{base}:{int(time.time() * 1000)}"


def _cleanup_strategy_import_stage(db: Session, sid: str, token: str | None = None) -> None:
    if token:
        db.execute(
            text(
                """
                DELETE FROM strategy_positions_import_stage
                WHERE import_token=:token AND strategy_id=:sid
                """
            ),
            {"token": token, "sid": sid},
        )
        return
    db.execute(
        text("DELETE FROM strategy_positions_import_stage WHERE strategy_id=:sid"),
        {"sid": sid},
    )


def _strategy_import_stage_row_count(db: Session, sid: str, token: str) -> int:
    row = db.execute(
        text(
            """
            SELECT COUNT(*) AS n
            FROM strategy_positions_import_stage
            WHERE import_token=:token AND strategy_id=:sid
            """
        ),
        {"token": token, "sid": sid},
    ).mappings().first()
    return int(row.get("n") or 0) if row else 0


def _stage_period_row_counts(db: Session, sid: str, token: str) -> tuple[dict[str, int], int]:
    rows = db.execute(
        text(
            """
            SELECT rebalance_date AS rb, COUNT(*) AS cnt
            FROM strategy_positions_import_stage
            WHERE import_token=:token AND strategy_id=:sid
            GROUP BY rebalance_date
            """
        ),
        {"token": token, "sid": sid},
    ).mappings().all()
    counts: dict[str, int] = {}
    total = 0
    for r in rows:
        d = _row_sql_date(r.get("rb"))
        if not d:
            continue
        k = d.isoformat()
        c = int(r.get("cnt") or 0)
        counts[k] = c
        total += c
    return counts, total


def _stage_existing_keys_for_dates(
    db: Session, sid: str, token: str, dates: list[date]
) -> set[tuple[str, str]]:
    uniq = sorted({d.isoformat() for d in dates})
    if not uniq:
        return set()
    out: set[tuple[str, str]] = set()
    for i in range(0, len(uniq), 60):
        part = uniq[i : i + 60]
        in_list = ", ".join(f":d{j}" for j in range(len(part)))
        params: dict[str, Any] = {"token": token, "sid": sid}
        for j, d in enumerate(part):
            params[f"d{j}"] = d
        rows = db.execute(
            text(
                f"""
                SELECT rebalance_date AS rb, stock_code
                FROM strategy_positions_import_stage
                WHERE import_token=:token
                  AND strategy_id=:sid
                  AND rebalance_date IN ({in_list})
                """
            ),
            params,
        ).mappings().all()
        for r in rows:
            d = _row_sql_date(r.get("rb"))
            code = str(r.get("stock_code") or "").strip()
            if d and code:
                out.add((d.isoformat(), code))
    return out


def _stage_missing_rows(
    db: Session,
    sid: str,
    token: str,
    rows: list[tuple[date, str, float | None, float | None]],
) -> list[tuple[date, str, float | None, float | None]]:
    existing = _stage_existing_keys_for_dates(db, sid, token, [r[0] for r in rows])
    return [r for r in rows if (r[0].isoformat(), r[1]) not in existing]


def _verify_strategy_import_stage_exact(
    db: Session,
    sid: str,
    token: str,
    ex_counts: dict[str, int],
    ex_total: int,
) -> int:
    stage_counts, stage_total = _stage_period_row_counts(db, sid, token)
    bad = _mismatch_rebalance_dates(ex_counts, stage_counts)
    if stage_total != ex_total or bad:
        sample = ", ".join(sorted(bad)[:8])
        raise ValueError(
            f"{sid} staging 校验失败：Excel {ex_total} 行，暂存 {stage_total} 行"
            + (f"，不一致调仓日 {len(bad)} 个：{sample}" if bad else "")
        )
    return stage_total


def _promote_strategy_import_stage(db: Session, sid: str, token: str) -> int:
    db.execute(text("DELETE FROM strategy_positions WHERE strategy_id=:sid"), {"sid": sid})
    db.execute(
        text(
            """
            INSERT INTO strategy_positions
            (strategy_id, rebalance_date, stock_code, holding_weight, industry_neutral_weight)
            SELECT strategy_id, rebalance_date, stock_code, holding_weight, industry_neutral_weight
            FROM strategy_positions_import_stage
            WHERE import_token=:token AND strategy_id=:sid
            """
        ),
        {"token": token, "sid": sid},
    )
    promoted = _strategy_import_stage_row_count(db, sid, token)
    _cleanup_strategy_import_stage(db, sid, token)
    db.commit()
    return promoted


def _write_stage_chunk_verified(
    db: Session,
    sid: str,
    token: str,
    chunk: list[tuple[date, str, float | None, float | None]],
    *,
    batch_no: int,
    max_attempts: int,
) -> None:
    pending = list(chunk)
    for attempt in range(max_attempts):
        _import_positions_stage_batch(db, token, sid, pending)
        db.commit()
        pending = _stage_missing_rows(db, sid, token, chunk)
        if not pending:
            return
        _log.warning(
            "import %s stage batch %s: %s/%s rows missing after commit, retry %s/%s",
            sid,
            batch_no,
            len(pending),
            len(chunk),
            attempt + 1,
            max_attempts,
        )
        time.sleep(0.2 * (2**attempt))
    sample = ", ".join(f"{r[0].isoformat()}|{r[1]}" for r in pending[:8])
    raise ValueError(
        f"{sid} stage batch {batch_no} 写入确认失败：仍缺 {len(pending)} 行"
        + (f"（示例 {sample}）" if sample else "")
    )


def _commit_positions_chunk_resilient(
    db: Session,
    sid: str,
    chunk: list[tuple[date, str, float | None, float | None]],
    *,
    batch_no: int = 0,
    stage_token: str | None = None,
) -> None:
    """
    写入并 commit（短持 Turso 流锁）。不在每批 COUNT；失败仅按 OperationalError 重试。
    """
    from app.db import run_under_turso_stream_lock

    need = len(chunk)
    if need <= 0:
        return
    max_attempts = _strategy_import_batch_retry_max()

    def _write_once() -> None:
        if stage_token:
            _write_stage_chunk_verified(
                db,
                sid,
                stage_token,
                chunk,
                batch_no=batch_no,
                max_attempts=max_attempts,
            )
        else:
            _import_positions_batch(db, sid, chunk)
            db.commit()

    for attempt in range(max_attempts):
        try:
            run_under_turso_stream_lock(_write_once)
            return
        except OperationalError as oe:
            try:
                db.rollback()
            except Exception:
                pass
            if _mysql_lock_contention(oe) and attempt < max_attempts - 1:
                time.sleep(0.3 * (2**attempt))
                continue
            raise
    _log.warning(
        "import %s batch %s: 重试 %s 次后仍失败，留待结束补写",
        sid,
        batch_no,
        max_attempts,
    )


def _repair_strategy_positions_against_excel(
    db: Session,
    sid: str,
    file_path: str,
    *,
    bad_dates: set[str] | None = None,
    ex_counts: dict[str, int] | None = None,
) -> int:
    """
    按调仓日对比 Excel/库内行数，对不一致的日期整期 UPSERT 补写。
    返回补写行数。
    """
    if bad_dates is None:
        if ex_counts is None:
            ex_counts, _ = _excel_period_row_counts(file_path)
        db_total = _strategy_positions_row_count(db, sid)
        ex_total = sum(ex_counts.values())
        if db_total != ex_total:
            db_counts, _ = _db_period_row_counts(db, sid)
            bad_dates = _mismatch_rebalance_dates(ex_counts, db_counts)
        else:
            db_counts = _db_period_row_counts_for_dates(db, sid, list(ex_counts.keys()))
            bad_dates = _mismatch_rebalance_dates_in_excel_only(ex_counts, db_counts)
    if not bad_dates:
        return 0
    _log.info(
        "import %s: 自动补写 %s 个调仓日（与 Excel 行数不一致）",
        sid,
        len(bad_dates),
    )
    buf: list[tuple[date, str, float | None, float | None]] = []
    repaired = 0
    sub_batch = max(80, _strategy_import_position_batch_size())
    for chunk in _iter_strategy_holdings_excel_batches(file_path):
        for row in chunk:
            if row[0].isoformat() in bad_dates:
                buf.append(row)
                if len(buf) >= sub_batch:
                    _commit_positions_chunk_resilient(db, sid, buf, batch_no=0)
                    repaired += len(buf)
                    buf.clear()
        chunk.clear()
    if buf:
        _commit_positions_chunk_resilient(db, sid, buf, batch_no=0)
        repaired += len(buf)
    if ex_counts is None:
        ex_counts, _ = _excel_period_row_counts(file_path)
    db_counts2 = _db_period_row_counts_for_dates(db, sid, list(bad_dates))
    still = [
        d for d in bad_dates if db_counts2.get(d, 0) != ex_counts.get(d, 0)
    ]
    if still:
        _log.warning(
            "import %s: 补写后仍有 %s 个调仓日与 Excel 不一致: %s…",
            sid,
            len(still),
            ", ".join(sorted(still)[:5]),
        )
    return repaired


def _turso_stream_busy(exc: BaseException) -> bool:
    s = str(exc).lower()
    return "stream already in use" in s or "hrana" in s and "400" in s


def _mysql_lock_contention(exc: BaseException) -> bool:
    """库锁等待 / 死锁（MySQL InnoDB 或 SQLite database is locked），可短重试。"""
    if isinstance(exc, OperationalError):
        orig = getattr(exc, "orig", None)
        if orig is not None and getattr(orig, "args", ()):
            try:
                code = int(orig.args[0])
            except (TypeError, ValueError):
                code = None
            if code in (1205, 1213):
                return True
    s = str(exc).lower()
    return (
        "1205" in s
        or "1213" in s
        or "lock wait timeout" in s
        or "deadlock" in s
        or "database is locked" in s
        or _turso_stream_busy(exc)
    )


def _format_update_job_failure_message(exc: BaseException) -> str:
    """写入 strategy_update_jobs.message，附常见原因便于运维排查。"""
    base = str(exc)[:60000]
    low = base.lower()
    extra = ""
    if _mysql_lock_contention(exc):
        extra = (
            " 【排查】与「导入/同步」或其它会话同时写 strategy_positions、"
            "strategy_holding_daily 等表时易触发库锁等待；请错开执行。"
        )
    elif "gone away" in low or "lost connection to mysql" in low:
        extra = " 【排查】MySQL 连接中断，检查 max_allowed_packet / wait_timeout / 网络后重试。"
    elif "no trade date" in low:
        extra = " 【排查】Wind 库无交易日数据或无法连接 Wind SQL Server。"
    elif "wind sql server" in low or "wind" in low and "未初始化" in base:
        extra = " 【排查】Wind 远程库未配置或 ODBC 失败，见应用启动日志与 .env 中 WIND_SQLSERVER_*。"
    elif "10054" in base or "08s01" in low or "通讯链接失败" in base:
        extra = (
            " 【排查】Wind SQL Server 连接被中断(10054/08S01)，常见于单次查询股票过多、"
            "执行超时或网络不稳；程序已分批拉取，若仍失败请检查 VPN、防火墙与 SQL Server 超时设置。"
        )
    return (base + extra)[:65000]


def _mark_update_job_failed(db: Session, job_id: int, message: str, do_commit: bool) -> None:
    """将任务标为 FAILED；主 Session 失效时用新连接补写，避免界面长期卡在 RUNNING。"""
    params = {"m": message[:65000], "id": job_id}
    upd = text(
        f"UPDATE strategy_update_jobs SET status='FAILED', finished_at={sql_now()}, message=:m WHERE id=:id"
    )
    try:
        db.execute(upd, params)
        if do_commit:
            db.commit()
    except Exception:
        if do_commit:
            try:
                db.rollback()
            except Exception:
                pass
        try:
            from app.db import SessionLocalFactory

            db2 = SessionLocalFactory()
            try:
                db2.execute(upd, params)
                db2.commit()
            finally:
                db2.close()
        except Exception:
            pass


def _wind_sql_transient_disconnect(exc: BaseException) -> bool:
    """pyodbc / SQL Server 执行中断、连接被对端关闭等，可换连接重试。"""
    s = str(exc).lower()
    if "10054" in s or "08s01" in s:
        return True
    if "通讯链接失败" in str(exc) or "远程主机强迫关闭" in str(exc):
        return True
    return "connection" in s and ("forcibly closed" in s or "broken pipe" in s)


def _fetch_wind_quote_map_batched(
    db: Session,
    wind: Any,
    stock_codes: list[str],
    td_compact: object,
) -> tuple[Any, dict[str, Any]]:
    """
    sql_quote_batch 在数百只股票 + 多表 OUTER APPLY 时，单条 SQL 易触发 SQL Server 超时或断连(10054)。
    按批 IN 查询并合并；遇瞬断则关闭连接后从池取新连接重试。
    """
    quote_map: dict[str, Any] = {}
    td = str(td_compact).strip()
    w = wind
    chunk = _WIND_QUOTE_CHUNK
    for i in range(0, len(stock_codes), chunk):
        part = stock_codes[i : i + chunk]
        quoted = ",".join("'" + c.replace("'", "''") + "'" for c in part)
        if not quoted:
            continue
        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                rows = w.execute(
                    text(wind_sql.sql_quote_batch(quoted)),
                    {"td_compact": td},
                ).mappings().all()
                for row in rows:
                    quote_map[_wind_code_key(row["stock_code"])] = row
                break
            except Exception as ex:
                if attempt >= max_attempts - 1 or not _wind_sql_transient_disconnect(ex):
                    raise
                _log.warning(
                    "Wind sql_quote_batch 分批失败 sid_chunk=%s..%s attempt=%s: %s",
                    i,
                    min(i + chunk, len(stock_codes)),
                    attempt + 1,
                    ex,
                )
                time.sleep(0.5 * (2**attempt))
                try:
                    w.close()
                except Exception:
                    pass
                w = wind_sql.open_wind(db)
    return w, quote_map


_HOLDING_INSERT_CHUNK = 200
_HOLDING_INSERT_SQL = text(
    """
    INSERT INTO strategy_holding_daily(
      strategy_id, trade_date, rebalance_date, stock_code, stock_name,
      period_weight, latest_weight, latest_price, last_1d_pct, period_return,
      ret_5d, ret_20d, ret_60d, ret_ytd, market_cap, industry_name, pe, pb
    ) VALUES (
      :strategy_id, :trade_date, :rebalance_date, :stock_code, :stock_name,
      :period_weight, :latest_weight, :latest_price, :last_1d_pct, :period_return,
      :ret_5d, :ret_20d, :ret_60d, :ret_ytd, :market_cap, :industry_name, :pe, :pb
    )
    ON CONFLICT(strategy_id, trade_date, rebalance_date, stock_code) DO UPDATE SET
      stock_name=excluded.stock_name,
      period_weight=excluded.period_weight,
      latest_weight=excluded.latest_weight,
      latest_price=excluded.latest_price,
      last_1d_pct=excluded.last_1d_pct,
      period_return=excluded.period_return,
      ret_5d=excluded.ret_5d,
      ret_20d=excluded.ret_20d,
      ret_60d=excluded.ret_60d,
      ret_ytd=excluded.ret_ytd,
      market_cap=excluded.market_cap,
      industry_name=excluded.industry_name,
      pe=excluded.pe,
      pb=excluded.pb
    """
)


def _group_strategy_positions_by_rebalance(
    db: Session, sid: str
) -> tuple[list[tuple[date, list[dict[str, Any]]]], date | None, date | None]:
    """
    按日历调仓日合并 strategy_positions（避免 TEXT 格式混用导致 DISTINCT 重复、写入 UNIQUE 冲突）。
    返回 (rb_positions 升序, 最新调仓日, 最早调仓日)。
    """
    rows = db.execute(
        text(
            """
            SELECT rebalance_date, stock_code, holding_weight, industry_neutral_weight
            FROM strategy_positions
            WHERE strategy_id=:sid
            """
        ),
        {"sid": sid},
    ).mappings().all()
    by_rd: dict[date, dict[str, dict[str, Any]]] = {}
    for r in rows:
        rd = _row_sql_date(r["rebalance_date"])
        if rd is None:
            continue
        sc = str(r.get("stock_code") or "").strip()
        if not sc:
            continue
        wk = _wind_code_key(sc)
        by_rd.setdefault(rd, {})[wk] = {
            "stock_code": sc,
            "holding_weight": r.get("holding_weight"),
            "industry_neutral_weight": r.get("industry_neutral_weight"),
        }
    if not by_rd:
        return [], None, None
    rb_sorted = sorted(by_rd.keys())
    rb_positions = [(rd, list(by_rd[rd].values())) for rd in rb_sorted]
    return rb_positions, rb_sorted[-1], rb_sorted[0]


def _last_nav_compact_for_update(db: Session, sid: str) -> str | None:
    """日常增量用末净值日；若尾部存在尺度断裂则退回最后一个好交易日。"""
    return _nav_last_good_trade_compact(db, sid) or _strategy_nav_max_trade_compact(
        db, sid
    )


def _needs_holding_snapshot_before_nav_incremental(
    db: Session,
    sid: str,
    rb_positions: list[tuple[date, list[Any]]],
) -> bool:
    """
    末净值存在但缺「锚定调仓期」持仓快照时，须先写持仓再跑增量净值。
    否则 bootstrap 只能用 strategy_positions 静态权重，与已漂移的 nav_unit 尺度不一致。
    """
    last_nav_c = _last_nav_compact_for_update(db, sid)
    if not last_nav_c or len(last_nav_c) < 8:
        return False
    rb_sorted = [d for d, _ in rb_positions if d is not None]
    if not rb_sorted:
        return False
    last_nav_d = datetime.strptime(last_nav_c[:8], "%Y%m%d").date()
    _, anchor_rb = _nav_rb_idx_on_date(rb_sorted, last_nav_d)
    return not _nav_holding_snapshot_on_day(db, sid, anchor_rb, last_nav_c)


def _holding_incremental_scope(
    db: Session,
    sid: str,
    rb_positions: list[tuple[date, list[dict[str, Any]]]],
    full_refresh: bool,
) -> tuple[int, date | None, str | None, str]:
    """
    与净值增量同一约定：末净值日 → 调仓日≤该日的最近一期为锚；
    仅锚定及之后调仓期拉 Wind，更早期沿用上一行情日快照（不逐期重拉全历史）。
    返回 (hold_start_idx, anchor_rb, last_nav_c, 说明文案)。
    """
    n_rb = len(rb_positions)
    if full_refresh:
        return 0, (rb_positions[0][0] if rb_positions else None), None, "全量刷新"
    last_nav_c = _last_nav_compact_for_update(db, sid)
    if not last_nav_c or len(last_nav_c) < 8:
        return (
            0,
            None,
            None,
            "库中无末净值，无法锚定；持仓将逐期拉 Wind（与首建相同）",
        )
    rb_sorted = [rb for rb, _ in rb_positions]
    last_nav_d = datetime.strptime(last_nav_c[:8], "%Y%m%d").date()
    hold_start_idx, anchor_rb = _nav_rb_idx_on_date(rb_sorted, last_nav_d)
    note = (
        f"末净值 {last_nav_c} → 锚定调仓 {_compact_date(anchor_rb)}，"
        f"持仓仅开放调仓期拉 Wind（EOD 回溯约 "
        f"{wind_bulk.holding_eod_lookback_calendar_days()} 自然日）"
    )
    return hold_start_idx, anchor_rb, last_nav_c, note


def _holding_anchor_start_idx(
    db: Session,
    sid: str,
    rb_positions: list[tuple[date, list[dict[str, Any]]]],
    full_refresh: bool,
) -> tuple[int, date | None]:
    idx, anchor, _, _ = _holding_incremental_scope(db, sid, rb_positions, full_refresh)
    return idx, anchor


def _holding_rb_indices_need_wind(
    rb_positions: list[tuple[date, list[dict[str, Any]]]],
    trade_date: date,
    full_refresh: bool,
    hold_start_idx: int,
) -> list[int]:
    """增量：仅「锚定及之后、且截至行情日仍未结束」的调仓期拉 Wind；已结束期复制上一日。"""
    if full_refresh:
        return list(range(len(rb_positions)))
    td_cmp = _compact_date(trade_date)
    out: list[int] = []
    for i, (_rb, _pos) in enumerate(rb_positions):
        if i < hold_start_idx:
            continue
        if i + 1 < len(rb_positions):
            pe = _compact_date(rb_positions[i + 1][0])
            if pe <= td_cmp:
                continue
        out.append(i)
    return out


def _holding_union_codes_for_indices(
    rb_positions: list[tuple[date, list[dict[str, Any]]]],
    indices: list[int],
) -> list[str]:
    codes: set[str] = set()
    for i in indices:
        for p in rb_positions[i][1]:
            if p.get("stock_code"):
                codes.add(_wind_code_key(p["stock_code"]))
    return sorted(codes)


def _effective_rebalance_stock_count(
    rb_positions: list[tuple[date, list[dict[str, Any]]]],
    *,
    as_of: date | None = None,
) -> int:
    if not rb_positions:
        return 0
    chosen = rb_positions[-1]
    if as_of is not None:
        for item in reversed(rb_positions):
            if item[0] <= as_of:
                chosen = item
                break
    latest_positions = chosen[1] or []
    codes = {
        _wind_code_key(p["stock_code"])
        for p in latest_positions
        if p.get("stock_code")
    }
    return len(codes)


def _holding_eod_start_for_indices(
    trade_date: date,
    rb_positions: list[tuple[date, list[dict[str, Any]]]],
    indices: list[int],
    *,
    full_refresh: bool,
) -> str:
    """单策略持仓 Wind 区间起点（取各待拉调仓期起点的最早 compact）。"""
    if not indices:
        return _compact_date(trade_date)
    starts: list[str] = []
    for i in indices:
        rb = rb_positions[i][0]
        starts.append(
            wind_bulk.holding_eod_start_for_period(
                trade_date, rb, full_refresh=full_refresh
            )
        )
    return min(starts)


def _holding_prior_trade_date(db: Session, sid: str, trade_date: date) -> date | None:
    row = db.execute(
        text(
            f"""
            SELECT {sql_max_date_expr("trade_date")} AS d
            FROM strategy_holding_daily
            WHERE strategy_id=:sid
              AND {sql_date_compact_expr("trade_date")} < :td_cmp
            """
        ),
        {"sid": sid, "td_cmp": _compact_date(trade_date)},
    ).mappings().first()
    if not row or row.get("d") is None:
        return None
    return _row_sql_date(row["d"])


def _copy_holding_daily_rebalances(
    db: Session,
    *,
    sid: str,
    from_trade_date: date,
    to_trade_date: date,
    rebalance_dates: list[date],
    do_commit: bool,
) -> int:
    """将上一行情日、锚定调仓期之前的各期快照复制到本日（不调 Wind）。"""
    rb_cmps = [_compact_date(rd) for rd in rebalance_dates]
    rb_cmps = [c for c in rb_cmps if len(c) >= 8]
    if not rb_cmps:
        return 0
    in_list = ",".join("'" + c.replace("'", "''") + "'" for c in rb_cmps)
    from_cmp = _compact_date(from_trade_date)
    to_td = to_trade_date
    res = db.execute(
        text(
            f"""
            INSERT INTO strategy_holding_daily (
              strategy_id, trade_date, rebalance_date, stock_code, stock_name,
              period_weight, latest_weight, latest_price, last_1d_pct, period_return,
              ret_5d, ret_20d, ret_60d, ret_ytd, market_cap, industry_name, pe, pb
            )
            SELECT
              strategy_id, :to_td, rebalance_date, stock_code, stock_name,
              period_weight, latest_weight, latest_price, last_1d_pct, period_return,
              ret_5d, ret_20d, ret_60d, ret_ytd, market_cap, industry_name, pe, pb
            FROM strategy_holding_daily
            WHERE strategy_id=:sid
              AND {sql_date_compact_expr("trade_date")} = :from_cmp
              AND {sql_date_compact_expr("rebalance_date")} IN ({in_list})
            ON CONFLICT(strategy_id, trade_date, rebalance_date, stock_code) DO UPDATE SET
              stock_name=excluded.stock_name,
              period_weight=excluded.period_weight,
              latest_weight=excluded.latest_weight,
              latest_price=excluded.latest_price,
              last_1d_pct=excluded.last_1d_pct,
              period_return=excluded.period_return,
              ret_5d=excluded.ret_5d,
              ret_20d=excluded.ret_20d,
              ret_60d=excluded.ret_60d,
              ret_ytd=excluded.ret_ytd,
              market_cap=excluded.market_cap,
              industry_name=excluded.industry_name,
              pe=excluded.pe,
              pb=excluded.pb
            """
        ),
        {"sid": sid, "to_td": to_td, "from_cmp": from_cmp},
    )
    if do_commit:
        db.commit()
    return int(res.rowcount or 0)


def _holding_snapshot_complete(
    db: Session,
    *,
    sid: str,
    trade_date: date,
    rebalance,
    expected_codes: list[str] | set[str],
) -> bool:
    expected = {
        _wind_code_key(c)
        for c in expected_codes
        if str(c or "").strip()
    }
    if not expected:
        return False
    row = db.execute(
        text(
            f"""
            SELECT COUNT(DISTINCT stock_code) AS n
            FROM strategy_holding_daily
            WHERE strategy_id=:sid
              AND {sql_date_compact_expr("trade_date")} = :td
              AND {sql_date_compact_expr("rebalance_date")} = :rb
            """
        ),
        {
            "sid": sid,
            "td": _compact_date(trade_date),
            "rb": _compact_date(rebalance),
        },
    ).mappings().first()
    return int((row or {}).get("n") or 0) >= len(expected)


def _flush_strategy_holding_daily_batch(db: Session, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    for i in range(0, len(rows), _HOLDING_INSERT_CHUNK):
        chunk = rows[i : i + _HOLDING_INSERT_CHUNK]
        db.execute(_HOLDING_INSERT_SQL, chunk)


def _holding_period_end_compact(
    next_rebalance: date | None,
    latest_trade: str | date,
) -> str | None:
    """
    本期持有段末日（含）：有下一调仓日用其日历日（按该日收盘价换仓），否则 None 表示截至 latest_trade。
    收益与短周期指标取 [本期调仓日, 段末日] 闭区间内的后复权价。
    """
    if next_rebalance is not None:
        return _compact_date(next_rebalance)
    return None


def _holding_quote_td_compact(
    period_end_compact: str | None,
    latest_trade: str | date,
) -> str:
    """行情/PE/PB 截止日：历史调仓期用本期段末日，当前开放期用 Wind 最新交易日。"""
    pe = str(period_end_compact or "").strip().replace("-", "")[:8]
    if len(pe) == 8 and pe.isdigit():
        return pe
    return _compact_date(latest_trade)


def _build_holding_daily_row_from_wind(
    *,
    sid: str,
    trade_date: date,
    rebalance,
    p: dict[str, Any],
    quote_map: dict[str, Any],
    eod_series: list,
    latest_trade: str,
    period_end_compact: str | None = None,
    desc_max_bars: int | None = None,
) -> dict[str, Any]:
    """由 Wind 行情 + 单股 EOD 序列生成 strategy_holding_daily 一行（不含 latest_weight）。

    quote_map 须与 period_end_compact 对齐：历史调仓期按段末日拉取（含 PE/PB/市值/行业）。
    """
    desc_n = int(desc_max_bars) if desc_max_bars is not None else 280
    scode = p["stock_code"]
    period_weight = float(p["holding_weight"] or 0.0)
    wk = _wind_code_key(scode)
    quote = quote_map.get(wk)
    if not quote:
        return {
            "strategy_id": sid,
            "trade_date": trade_date,
            "rebalance_date": rebalance,
            "stock_code": scode,
            "stock_name": None,
            "period_weight": period_weight,
            "latest_price": None,
            "last_1d_pct": None,
            "period_return": None,
            "ret_5d": None,
            "ret_20d": None,
            "ret_60d": None,
            "ret_ytd": None,
            "market_cap": None,
            "industry_name": None,
            "pe": None,
            "pb": None,
        }
    latest_price = float(quote["latest_price"] or 0.0)
    prev_close = float(quote["prev_close"] or 0.0)
    lt_compact = str(latest_trade).strip().replace("-", "")[:8]
    rb_compact_this = _compact_date(rebalance)
    period_start_close = wind_bulk.first_close_on_or_after(eod_series, rb_compact_this)

    if period_end_compact:
        asof_c = period_end_compact
        seg = wind_bulk.series_on_or_before(eod_series, asof_c)
        period_end_px = wind_bulk.last_close_on_or_before(eod_series, asof_c)
        period_ret = _safe_return(period_end_px, period_start_close)
        day_ret = wind_bulk.day_return_adj_for_asof(eod_series, asof_c)
        desc_closes = wind_bulk.closes_desc_from_asc(seg, desc_n)
        px = (
            period_end_px
            if period_end_px is not None and period_end_px > 0
            else (desc_closes[0] if desc_closes else None)
        )
        end_year = int(asof_c[:4])
        ytd_close = wind_bulk.last_close_before_calendar_date(
            seg, f"{end_year}0101"
        )
        row_price = px
    else:
        asof_c = lt_compact
        day_ret = wind_bulk.day_return_adj_for_asof(eod_series, asof_c)
        if day_ret is None:
            day_ret = _safe_return(latest_price, prev_close)
        desc_closes = wind_bulk.closes_desc_from_asc(eod_series, desc_n)
        latest_adj = desc_closes[0] if desc_closes else None
        px = latest_adj if (latest_adj is not None and latest_adj > 0) else latest_price
        period_ret = _safe_return(px, period_start_close)
        ytd_close = wind_bulk.last_close_before_calendar_date(
            eod_series, f"{trade_date.year}0101"
        )
        row_price = latest_price if latest_price > 0 else px

    close_5 = wind_bulk.close_n_trading_days_ago(desc_closes, 5)
    close_20 = wind_bulk.close_n_trading_days_ago(desc_closes, 20)
    close_60 = wind_bulk.close_n_trading_days_ago(desc_closes, 60)

    return {
        "strategy_id": sid,
        "trade_date": trade_date,
        "rebalance_date": rebalance,
        "stock_code": scode,
        "stock_name": quote["stock_name"],
        "period_weight": period_weight,
        "latest_price": row_price if row_price and float(row_price) > 0 else None,
        "last_1d_pct": day_ret,
        "period_return": period_ret,
        "ret_5d": _safe_return(px, close_5),
        "ret_20d": _safe_return(px, close_20),
        "ret_60d": _safe_return(px, close_60),
        "ret_ytd": _safe_return(px, ytd_close),
        "market_cap": quote["market_cap"],
        "industry_name": quote["industry_name"],
        "pe": quote["pe"],
        "pb": quote["pb"],
    }


def _flush_rebalance_holding_period(
    db: Session,
    *,
    sid: str,
    trade_date: date,
    rebalance,
    i_rb: int,
    prepared_rows: list[dict[str, Any]],
    total_weight: float,
    do_commit: bool,
) -> int:
    """计算 drift/目标权重并写入本期持仓快照。返回写入条数。"""
    snap_target_weights = False
    rb_d = _row_sql_date(rebalance)
    rb_cmp = _compact_date(rebalance)
    td_cmp = _compact_date(trade_date)
    if rb_cmp and td_cmp:
        db.execute(
            text(
                f"""
                DELETE FROM strategy_holding_daily
                WHERE strategy_id = :sid
                  AND {sql_date_compact_expr("trade_date")} = :td_cmp
                  AND {sql_date_compact_expr("rebalance_date")} = :rb_cmp
                """
            ),
            {"sid": sid, "td_cmp": td_cmp, "rb_cmp": rb_cmp},
        )
    if i_rb == 1 and rb_d is not None and total_weight > 0:
        mx = db.execute(
            text(
                f"""
                SELECT {sql_max_date_expr("trade_date")} AS m
                FROM strategy_nav_daily
                WHERE strategy_id=:sid
                  AND {sql_date_compact_expr("trade_date")} < :td_cmp
                """
            ),
            {"sid": sid, "td_cmp": _compact_date(trade_date)},
        ).mappings().first()
        last_nt = _row_sql_date(mx["m"]) if mx and mx.get("m") is not None else None
        if trade_date == rb_d:
            snap_target_weights = True
        elif trade_date > rb_d and (last_nt is None or last_nt < rb_d):
            snap_target_weights = True
    if snap_target_weights:
        for row in prepared_rows:
            pw = max(row["period_weight"], 0.0)
            row["latest_weight"] = pw / total_weight
    else:
        drift_total = sum(
            max(row["period_weight"], 0.0)
            * (1.0 + (row["period_return"] if row["period_return"] is not None else 0.0))
            for row in prepared_rows
        )
        base = drift_total if drift_total > 0 else (total_weight if total_weight > 0 else 1.0)
        for row in prepared_rows:
            drift = max(row["period_weight"], 0.0) * (
                1.0 + (row["period_return"] if row["period_return"] is not None else 0.0)
            )
            row["latest_weight"] = drift / base
    deduped: dict[str, dict[str, Any]] = {}
    for row in prepared_rows:
        sc = str(row.get("stock_code") or "").strip()
        if sc:
            deduped[_wind_code_key(sc)] = dict(row)
    _flush_strategy_holding_daily_batch(db, list(deduped.values()))
    if do_commit:
        db.commit()
    return len(prepared_rows)


def _run_update_try_build_work_item(
    db: Session,
    cfg: Any,
    trade_date: date,
    full_refresh: bool,
    job_id: int,
    do_commit: bool,
    *,
    sync_job_id: int | None = None,
) -> dict[str, Any] | None:
    """
    若本策略本次 run_update 应跳过或无可处理持仓，返回 None（并已写进度）；
    否则返回后续计算所需的元数据（不再重复查 positions / rebalance 列表）。
    """
    sid = cfg["strategy_id"]
    bench_code_raw = cfg.get("benchmark_code")
    bench_code = str(bench_code_raw or "").strip().upper() if bench_code_raw else ""
    last_row = db.execute(
        text(
            f"""
            SELECT {sql_max_date_expr("trade_date")} AS d
            FROM strategy_holding_daily
            WHERE strategy_id=:sid
            """
        ),
        {"sid": sid},
    ).mappings().first()
    last_td = _row_sql_date(last_row["d"]) if last_row else None
    rb_positions, latest_rb, min_rb_date = _group_strategy_positions_by_rebalance(db, sid)
    if not rb_positions or latest_rb is None or min_rb_date is None:
        return None
    latest_rb_compact = _compact_date(latest_rb)
    code_keys: set[str] = set()
    for _, positions in rb_positions:
        for p in positions:
            if p.get("stock_code"):
                code_keys.add(_wind_code_key(p["stock_code"]))
    stock_codes = sorted(code_keys)
    if not stock_codes:
        return None
    hold_start_idx, anchor_rb_hold, last_nav_c_scope, hold_scope_note = (
        _holding_incremental_scope(db, sid, rb_positions, full_refresh)
    )
    wind_rb_indices = _holding_rb_indices_need_wind(
        rb_positions, trade_date, full_refresh, hold_start_idx
    )
    n_rb_wind = len(wind_rb_indices)
    wind_stock_codes = (
        _holding_union_codes_for_indices(rb_positions, wind_rb_indices)
        if wind_rb_indices
        else stock_codes
    )
    if full_refresh or hold_start_idx <= 0 or anchor_rb_hold is None:
        start_c = wind_bulk.bulk_eod_start_compact(trade_date, min_rb_date)
    else:
        start_c = _holding_eod_start_for_indices(
            trade_date, rb_positions, wind_rb_indices, full_refresh=False
        )
        for i in wind_rb_indices:
            rb = rb_positions[i][0]
            st = wind_bulk.holding_eod_start_for_period(
                trade_date, rb, full_refresh=False
            )
            if st < start_c:
                start_c = st
    skip_holdings = False
    if (not full_refresh) and last_td is not None and last_td >= trade_date:
        rb_pos_row = db.execute(
            text(
                f"SELECT {sql_max_date_expr('rebalance_date')} AS m "
                "FROM strategy_positions WHERE strategy_id=:sid"
            ),
            {"sid": sid},
        ).mappings().first()
        td_cmp = _compact_date(trade_date)
        rb_hold_row = db.execute(
            text(
                f"""
                SELECT {sql_max_date_expr("rebalance_date")} AS m
                FROM strategy_holding_daily
                WHERE strategy_id=:sid
                  AND {sql_date_compact_expr("trade_date")} = :td_cmp
                """
            ),
            {"sid": sid, "td_cmp": td_cmp},
        ).mappings().first()
        mx_pos = rb_pos_row.get("m") if rb_pos_row else None
        mx_hold = rb_hold_row.get("m") if rb_hold_row else None
        dp = _row_sql_date(mx_pos)
        dh = _row_sql_date(mx_hold)
        if dp is not None and dh is not None and dp <= dh:
            last_nav_c = _strategy_nav_max_trade_compact(db, sid)
            td_nav_cmp = _compact_date(trade_date)
            if (
                last_nav_c
                and len(last_nav_c) >= 8
                and last_nav_c >= td_nav_cmp
            ):
                _job_progress(
                    db,
                    job_id,
                    f"{sid}：增量跳过（行情日={trade_date} 持仓与净值均已齐），刷新列表指标快照…",
                    do_commit=do_commit,
                    sync_job_id=sync_job_id,
                )
                from app.strategy_list_metrics import refresh_strategy_list_metrics_safe

                refresh_strategy_list_metrics_safe(
                    db,
                    sid,
                    do_commit=do_commit,
                    stock_count_on_last_date=_effective_rebalance_stock_count(
                        rb_positions, as_of=trade_date
                    ),
                    last_trade_date=trade_date.isoformat(),
                )
                return None
            _job_progress(
                db,
                job_id,
                f"{sid}：持仓快照已齐，仅补算净值（行情日={trade_date}）",
                do_commit=do_commit,
                sync_job_id=sync_job_id,
            )
            skip_holdings = True
    return {
        "cfg": cfg,
        "sid": sid,
        "bench_code": bench_code,
        "trade_date": trade_date,
        "latest_rb": latest_rb,
        "latest_rb_compact": latest_rb_compact,
        "n_rb": len(rb_positions),
        "rb_positions": rb_positions,
        "stock_codes": stock_codes,
        "min_rb_date": min_rb_date,
        "start_c": start_c,
        "skip_holdings": skip_holdings,
        "hold_start_idx": hold_start_idx,
        "anchor_rb_hold": anchor_rb_hold,
        "last_nav_c_scope": last_nav_c_scope,
        "hold_scope_note": hold_scope_note,
        "n_rb_wind": n_rb_wind,
        "wind_rb_indices": wind_rb_indices,
        "wind_stock_codes": wind_stock_codes,
    }


def _wind_low_memory_mode() -> bool:
    return bool(getattr(settings, "wind_low_memory_mode", True))


def _release_wind_memory(*bundles: dict[str, Any] | None) -> None:
    for wb in bundles:
        if not wb:
            continue
        for key in ("eod", "idx", "quote", "td"):
            part = wb.get(key)
            if isinstance(part, dict):
                part.clear()
            elif isinstance(part, list):
                part.clear()
        wb.clear()
    gc.collect()


def _use_wind_merged_prefetch(n_strategies: int, n_union_codes: int) -> bool:
    """多策略合并预拉 EOD 会长期占用内存直至整轮结束，Render 上默认仅单策略启用。"""
    if _wind_low_memory_mode():
        return False
    max_s = max(1, int(getattr(settings, "wind_merged_prefetch_max_strategies", 1) or 1))
    if n_strategies > max_s:
        return False
    max_codes = int(getattr(settings, "wind_merged_prefetch_max_union_codes", 800) or 800)
    if max_codes > 0 and n_union_codes > max_codes:
        return False
    return True


def _release_run_update_strategy_memory(
    *,
    eod_by_code: dict | None = None,
    index_eod_by_code: dict | None = None,
    quote_map: dict | None = None,
    eod_local: dict | None = None,
) -> None:
    """单策略 run_update 循环末尾：断开对大 dict 的引用并建议 GC。"""
    for d in (eod_by_code, index_eod_by_code, quote_map, eod_local):
        if isinstance(d, dict):
            d.clear()
    gc.collect()


def _load_wind_bundle_for_nav_plan(
    wind: Any,
    db: Session,
    plan: dict[str, Any],
    latest_trade_c: str,
) -> tuple[Any, dict[str, Any]]:
    """单策略净值：仅拉本策略股票+基准+交易日历（低内存串行）。"""
    codes = sorted(plan["code_set"])
    start_c = str(plan["start_c"])
    bench = str(plan.get("bench_code") or "").strip().upper()
    lt = str(latest_trade_c).strip()
    wind, eod = wind_bulk.load_eod_by_code(wind, codes, start_c, lt, db)
    idx: dict[str, list] = {}
    if bench:
        wind, idx = wind_bulk.load_index_eod_by_code(wind, [bench], start_c, lt, db)
    wind, td = wind_bulk.fetch_trade_date_compacts(wind, db, start_c, lt)
    return wind, {"eod": eod, "idx": idx, "td": td}


def _run_update_prefetch_wind_merged(
    db: Session,
    wind: Any,
    work_items: list[dict[str, Any]],
    latest_trade_compact: str,
) -> tuple[Any, dict[str, Any] | None]:
    """对本轮待处理策略合并拉取 Wind EOD / 指数 / 行情，减少 SQL Server 往返（低内存模式不调用）。"""
    union_codes: set[str] = set()
    union_bench: set[str] = set()
    global_st: str | None = None
    for w in work_items:
        for c in w.get("wind_stock_codes") or w["stock_codes"]:
            union_codes.add(c)
        bc = str(w.get("bench_code") or "").strip().upper()
        if bc:
            union_bench.add(bc)
        st = w["start_c"]
        if global_st is None or st < global_st:
            global_st = st
    if not union_codes or not global_st:
        return wind, None
    lt = str(latest_trade_compact).strip()
    wind, eod_all = wind_bulk.load_eod_by_code(wind, sorted(union_codes), global_st, lt, db)
    idx_all: dict[str, list] = {}
    if union_bench:
        wind, idx_all = wind_bulk.load_index_eod_by_code(
            wind, sorted(union_bench), global_st, lt, db
        )
    wind, quote_all = _fetch_wind_quote_map_batched(db, wind, sorted(union_codes), lt)
    wind, td_all = wind_bulk.fetch_trade_date_compacts(wind, db, global_st, lt)
    return wind, {"eod": eod_all, "idx": idx_all, "quote": quote_all, "td": td_all}


def _import_write_positions_one_strategy(
    db: Session,
    sid: str,
    rows_to_write: list[tuple[date, str, float | None, float | None]],
    label_meta: dict[str, str],
    import_mode: str,
    *,
    repair_last_period: bool = False,
    file_path: str = "",
) -> None:
    """单策略：删/增 strategy_positions，可选更新 strategy_configs 元数据（不含 commit）。"""
    if repair_last_period and file_path and str(import_mode or "").lower() == "incremental":
        _heal_positions_gaps_from_excel(
            db,
            sid,
            file_path,
            incremental_slice=True,
        )
    cutoff_rb, cutoff_inclusive, delete_all, _baseline = _import_cutoff_plan(
        db,
        sid,
        import_mode,
        repair_last_period=repair_last_period,
        file_path=file_path,
    )
    if delete_all:
        db.execute(text("DELETE FROM strategy_positions WHERE strategy_id=:sid"), {"sid": sid})
        write_rows = rows_to_write
    else:
        write_rows = [
            x
            for x in rows_to_write
            if cutoff_rb is None
            or _import_incremental_row_allowed(
                x[0], cutoff_rb, inclusive=cutoff_inclusive
            )
        ]
    _import_positions_batch(db, sid, write_rows)
    if label_meta:
        keys = [k for k, v in label_meta.items() if str(v or "").strip()]
        if keys:
            sets = ", ".join(f"{k}=:{k}" for k in keys)
            params = {k: label_meta[k] for k in keys}
            params["sid"] = sid
            db.execute(
                text(f"UPDATE strategy_configs SET {sets} WHERE strategy_id=:sid"),
                params,
            )


def normalize_code(code) -> str:
    """证券代码：支持 Wind 风格字符串，以及 Excel 读出的 int/float（如 600000.0）。"""
    if code is None or (isinstance(code, float) and pd.isna(code)):
        raise ValueError("empty stock code")
    if isinstance(code, bool):
        raise ValueError("invalid stock code")
    if isinstance(code, int):
        text_code = str(code)
    elif isinstance(code, float):
        text_code = str(int(code)) if code.is_integer() else str(code).strip().upper()
    else:
        text_code = str(code).strip().upper()
        if len(text_code) > 2 and text_code[-2:] == ".0" and text_code[:-2].isdigit():
            text_code = text_code[:-2]
    if "." in text_code:
        return text_code
    if len(text_code) == 6 and text_code.startswith("6"):
        return f"{text_code}.SH"
    if len(text_code) == 6:
        return f"{text_code}.SZ"
    return text_code


def _json_str_list(raw: Any) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    try:
        v = json.loads(str(raw))
        if isinstance(v, list):
            return [str(x).strip() for x in v if str(x).strip()]
    except json.JSONDecodeError:
        pass
    return []


def _sync_load_checkpoint(raw: Any) -> dict[str, Any]:
    if not raw:
        return {
            "completed_import": [],
            "completed_nav": [],
            "completed_update_rb": [],
            "nav_progress": {},
            "stage": "import",
        }
    try:
        ck = json.loads(str(raw)) if not isinstance(raw, dict) else raw
        if not isinstance(ck, dict):
            return {
                "completed_import": [],
                "completed_nav": [],
                "completed_update_rb": [],
                "nav_progress": {},
                "stage": "import",
            }
        return {
            "completed_import": _json_str_list(ck.get("completed_import")),
            "completed_nav": _json_str_list(ck.get("completed_nav")),
            "completed_update_rb": _json_str_list(ck.get("completed_update_rb")),
            "nav_progress": ck.get("nav_progress") if isinstance(ck.get("nav_progress"), dict) else {},
            "stage": str(ck.get("stage") or "import"),
        }
    except json.JSONDecodeError:
        return {
            "completed_import": [],
            "completed_nav": [],
            "completed_update_rb": [],
            "nav_progress": {},
            "stage": "import",
        }


def _sync_save_checkpoint(
    db: Session,
    sync_job_id: int,
    *,
    completed_import: list[str],
    completed_nav: list[str],
    stage: str,
    completed_update_rb: list[str] | None = None,
    nav_progress: dict[str, Any] | None = None,
    do_commit: bool = True,
) -> None:
    payload: dict[str, Any] = {
        "completed_import": completed_import,
        "completed_nav": completed_nav,
        "stage": stage,
    }
    if completed_update_rb is not None:
        payload["completed_update_rb"] = completed_update_rb
    if nav_progress is not None:
        payload["nav_progress"] = nav_progress
    body = json.dumps(payload, ensure_ascii=False)
    db.execute(
        text("UPDATE admin_sync_jobs SET checkpoint_json=:c WHERE id=:id"),
        {"c": body, "id": sync_job_id},
    )
    if do_commit:
        db.commit()


def admin_sync_job_bootstrap_checkpoint(
    db: Session,
    sync_job_id: int,
    *,
    strategy_ids: list[str] | None = None,
    import_mode: str | None = None,
    skip_excel_import: bool = True,
    do_commit: bool = True,
) -> dict[str, Any]:
    """checkpoint_json 为空时写入初始断点，避免「可续传」但无断点记录。"""
    row = (
        db.execute(
            text(
                """
                SELECT checkpoint_json, strategy_ids_json, import_mode
                FROM admin_sync_jobs WHERE id=:id
                """
            ),
            {"id": sync_job_id},
        )
        .mappings()
        .first()
    )
    if not row:
        raise ValueError("sync job not found")
    raw_ck = row.get("checkpoint_json")
    if raw_ck is not None and str(raw_ck).strip():
        return _sync_load_checkpoint(raw_ck)
    ids = [str(x).strip() for x in (strategy_ids or []) if str(x).strip()]
    if not ids:
        ids = _json_str_list(row.get("strategy_ids_json"))
    mode = str(import_mode or row.get("import_mode") or "incremental").strip().lower()
    completed_import = list(ids) if skip_excel_import or mode == "full" else []
    stage = "nav" if skip_excel_import else "import"
    _sync_save_checkpoint(
        db,
        sync_job_id,
        completed_import=completed_import,
        completed_nav=[],
        completed_update_rb=[],
        stage=stage,
        nav_progress={},
        do_commit=do_commit,
    )
    return {
        "completed_import": completed_import,
        "completed_nav": [],
        "completed_update_rb": [],
        "nav_progress": {},
        "stage": stage,
    }


def _sync_nav_progress_map(raw: Any) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    if not isinstance(raw, dict):
        return out
    for sid_raw, val in raw.items():
        sid = str(sid_raw or "").strip()
        if not sid or not isinstance(val, dict):
            continue
        last_done = str(val.get("last_done") or "").strip().replace("-", "")[:8]
        start = str(val.get("start") or "").strip().replace("-", "")[:8]
        latest = str(val.get("latest") or "").strip().replace("-", "")[:8]
        if len(last_done) == 8 and last_done.isdigit():
            item: dict[str, str] = {"last_done": last_done}
            if len(start) == 8 and start.isdigit():
                item["start"] = start
            if len(latest) == 8 and latest.isdigit():
                item["latest"] = latest
            out[sid] = item
    return out


def _sync_nav_progress_for_strategy(
    sync_job_id: int | None,
    strategy_id: str,
    *,
    db: Session,
) -> dict[str, str]:
    if sync_job_id is None:
        return {}
    row = (
        db.execute(
            text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
            {"id": sync_job_id},
        )
        .mappings()
        .first()
    )
    cp = _sync_load_checkpoint(row.get("checkpoint_json") if row else None)
    progress = _sync_nav_progress_map(cp.get("nav_progress"))
    return progress.get(str(strategy_id or "").strip(), {})


def _sync_mark_nav_progress(
    sync_job_id: int | None,
    strategy_id: str,
    last_done: str | None,
    *,
    start_c: str,
    latest_c: str,
    db: Session,
    do_commit: bool = True,
) -> None:
    if sync_job_id is None or not last_done:
        return
    sid = str(strategy_id or "").strip()
    last_c = str(last_done or "").strip().replace("-", "")[:8]
    if not sid or len(last_c) != 8 or not last_c.isdigit():
        return
    row = (
        db.execute(
            text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
            {"id": sync_job_id},
        )
        .mappings()
        .first()
    )
    cp = _sync_load_checkpoint(row.get("checkpoint_json") if row else None)
    progress = _sync_nav_progress_map(cp.get("nav_progress"))
    progress[sid] = {
        "last_done": last_c,
        "start": str(start_c or "").strip().replace("-", "")[:8],
        "latest": str(latest_c or "").strip().replace("-", "")[:8],
    }
    _sync_save_checkpoint(
        db,
        sync_job_id,
        completed_import=list(cp.get("completed_import") or []),
        completed_nav=list(cp.get("completed_nav") or []),
        completed_update_rb=list(cp.get("completed_update_rb") or []),
        nav_progress=progress,
        stage="nav",
        do_commit=do_commit,
    )


def _sync_update_rb_key(strategy_id: str, rb_compact: str) -> str:
    sid = str(strategy_id or "").strip()
    rb_key = str(rb_compact or "").strip().replace("-", "")[:8]
    return f"{sid}:{rb_key}" if sid and rb_key else rb_key


def _sync_normalize_update_rb_done(values: Any) -> set[str]:
    done: set[str] = set()
    for raw in values or []:
        val = str(raw or "").strip()
        if not val:
            continue
        if ":" in val:
            sid, rb = val.split(":", 1)
            key = _sync_update_rb_key(sid, rb)
        else:
            key = val.replace("-", "")[:8]
        if key:
            done.add(key)
    return done


def _sync_update_rb_done_for_strategy(
    done: set[str] | None,
    strategy_id: str,
    *,
    allow_legacy_dates: bool,
) -> set[str]:
    sid = str(strategy_id or "").strip()
    prefix = f"{sid}:"
    out: set[str] = set()
    for raw in done or set():
        val = str(raw or "").strip()
        if not val:
            continue
        if ":" in val:
            if val.startswith(prefix):
                rb = val.split(":", 1)[1].replace("-", "")[:8]
                if rb:
                    out.add(rb)
        elif allow_legacy_dates:
            rb = val.replace("-", "")[:8]
            if rb:
                out.add(rb)
    return out


def _sync_mark_update_rb_done(
    sync_job_id: int,
    strategy_id: str,
    rb_compact: str,
    *,
    db: Session | None = None,
    do_commit: bool = True,
) -> None:
    """阶段3 每写完一个调仓期，落库断点，避免 OOM/重启后从第一期重拉 EOD。"""
    rb_key = str(rb_compact or "").strip().replace("-", "")[:8]
    if not rb_key:
        return
    done_key = _sync_update_rb_key(strategy_id, rb_key)

    def _apply(sess: Session) -> None:
        row = (
            sess.execute(
                text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
                {"id": sync_job_id},
            )
            .mappings()
            .first()
        )
        cp = _sync_load_checkpoint(row.get("checkpoint_json") if row else None)
        done = _sync_normalize_update_rb_done(cp.get("completed_update_rb") or [])
        done.add(done_key)
        _sync_save_checkpoint(
            sess,
            sync_job_id,
            completed_import=list(cp.get("completed_import") or []),
            completed_nav=list(cp.get("completed_nav") or []),
            completed_update_rb=sorted(done),
            nav_progress=_sync_nav_progress_map(cp.get("nav_progress")),
            stage="update",
            do_commit=do_commit,
        )

    if db is not None:
        _apply(db)
        return
    from app.db import SessionLocalFactory, turso_stream_lock

    with turso_stream_lock():
        own = SessionLocalFactory()
        try:
            _apply(own)
        finally:
            own.close()


def _effective_eod_chunk_size(n_stocks: int) -> int:
    base = wind_bulk.eod_stock_chunk_size()
    if n_stocks > 800:
        return min(base, 8)
    if n_stocks > 400:
        return min(base, 10)
    return base


def _strategy_import_job_save_checkpoint(
    db: Session,
    job_id: int,
    sid: str,
    *,
    stats: dict[str, int],
    batch_no: int,
    do_commit: bool = True,
) -> None:
    """每批 commit 后写 checkpoint（用内存累计行数，不再每批 COUNT 全表）。"""
    rows_after = int(stats.get("rows_after") or 0)
    max_rb_s = stats.get("last_committed_rebalance_date")
    payload = {
        "strategy_id": sid,
        "batch_no": batch_no,
        "excel_rows_scanned": int(stats.get("excel_rows") or 0),
        "imported_rows_session": int(stats.get("imported_rows") or 0),
        "skipped_rows": int(stats.get("skipped_rows") or 0),
        "rows_in_db": rows_after,
        "db_max_rebalance_date": str(max_rb_s) if max_rb_s else None,
        "expected_file_rows": int(stats.get("expected_file_rows") or 0) or None,
    }
    db.execute(
        text(
            f"""
            UPDATE strategy_import_jobs
            SET checkpoint_json=:cp, progress_at={sql_now()}
            WHERE id=:id
            """
        ),
        {"cp": json.dumps(payload, ensure_ascii=False), "id": job_id},
    )
    if do_commit:
        db.commit()


def _strategy_import_job_touch(
    db: Session,
    job_id: int,
    *,
    status: str | None = None,
    message: str | None = None,
    completed_ids: list[str] | None = None,
    imported: int | None = None,
    failed: int | None = None,
    errors: list[str] | None = None,
    do_commit: bool = True,
) -> None:
    sets: list[str] = []
    params: dict[str, Any] = {"id": job_id}
    if status is not None:
        sets.append("status=:st")
        params["st"] = status
    if message is not None:
        sets.append("message=:msg")
        params["msg"] = (message or "")[:6000]
    if completed_ids is not None:
        sets.append("completed_strategy_ids_json=:cj")
        params["cj"] = json.dumps(completed_ids, ensure_ascii=False)
    if imported is not None:
        sets.append("imported_count=:ic")
        params["ic"] = int(imported)
    if failed is not None:
        sets.append("failed_count=:fc")
        params["fc"] = int(failed)
    if errors is not None:
        sets.append("errors_json=:ej")
        params["ej"] = json.dumps(errors[:50], ensure_ascii=False)
    if not sets:
        return
    sets.append(f"progress_at={sql_now()}")
    db.execute(
        text(f"UPDATE strategy_import_jobs SET {', '.join(sets)} WHERE id=:id"),
        params,
    )
    if do_commit:
        db.commit()


def create_strategy_import_job(
    db: Session,
    *,
    strategy_ids: list[str],
    import_mode: str,
    triggered_by: str,
) -> int:
    ids = [str(x).strip() for x in strategy_ids if str(x).strip()]
    res = db.execute(
        text(
            """
            INSERT INTO strategy_import_jobs
            (status, import_mode, strategy_ids_json, completed_strategy_ids_json,
             imported_count, failed_count, message, triggered_by)
            VALUES ('QUEUED', :im, :sj, '[]', 0, 0, :msg, :by)
            """
        ),
        {
            "im": import_mode,
            "sj": json.dumps(ids, ensure_ascii=False),
            "msg": "任务已入队，等待后台执行",
            "by": triggered_by,
        },
    )
    from app.sql_dialect import executed_rowid

    job_id = executed_rowid(db, res)
    if not job_id:
        raise RuntimeError("创建策略导入任务失败")
    abandon_older_strategy_import_jobs(db, keep_job_id=job_id)
    return job_id


def get_strategy_import_job_row(db: Session, job_id: int) -> dict[str, Any] | None:
    row = (
        db.execute(
            text(
                """
                SELECT id, status, import_mode, strategy_ids_json, completed_strategy_ids_json,
                       imported_count, failed_count, errors_json, message, triggered_by,
                       created_at, started_at, finished_at, progress_at, checkpoint_json
                FROM strategy_import_jobs
                WHERE id = :id
                LIMIT 1
                """
            ),
            {"id": job_id},
        )
        .mappings()
        .first()
    )
    return dict(row) if row else None


def strategy_import_job_is_resumable(job: dict[str, Any]) -> bool:
    st = str(job.get("status") or "").upper()
    if st in ("SUCCESS", "ABANDONED"):
        return False
    if "Superseded by a newer" in str(job.get("message") or ""):
        return False
    return st in ("FAILED", "RUNNING", "QUEUED", "PARTIAL")


def abandon_older_strategy_import_jobs(db: Session, *, keep_job_id: int) -> None:
    db.execute(
        text(
            f"""
            UPDATE strategy_import_jobs
            SET status='FAILED',
                finished_at={sql_now()},
                checkpoint_json=NULL,
                message='Superseded by a newer strategy import job'
            WHERE id <> :keep
              AND status IN ('FAILED', 'PARTIAL')
            """
        ),
        {"keep": int(keep_job_id)},
    )


def abandon_older_admin_sync_jobs(db: Session, *, keep_job_id: int) -> None:
    db.execute(
        text(
            f"""
            UPDATE admin_sync_jobs
            SET status='FAILED',
                checkpoint_json=NULL,
                result_json=:result_json,
                finished_at={sql_now()},
                message='Superseded by a newer sync job'
            WHERE id <> :keep
              AND status='FAILED'
            """
        ),
        {
            "keep": int(keep_job_id),
            "result_json": json.dumps(
                {"resumable": False, "superseded": True},
                ensure_ascii=False,
                separators=(",", ":"),
            ),
        },
    )


def find_resumable_strategy_import_job(
    db: Session, strategy_ids: list[str]
) -> dict[str, Any] | None:
    """是否存在针对相同策略、可续传的失败/中断任务（避免重复全量清空）。"""
    want = {str(x).strip() for x in strategy_ids if str(x).strip()}
    if not want:
        return None
    rows = db.execute(
        text(
            """
            SELECT id, status, import_mode, strategy_ids_json, message, progress_at
            FROM strategy_import_jobs
            WHERE status IN ('FAILED', 'RUNNING', 'QUEUED', 'PARTIAL')
            ORDER BY id DESC
            LIMIT 30
            """
        )
    ).mappings().all()
    for row in rows:
        job = dict(row)
        if not strategy_import_job_is_resumable(job):
            continue
        ids = set(_json_str_list(job.get("strategy_ids_json")))
        if ids & want:
            return job
    return None


def count_strategy_positions_rows(db: Session, strategy_ids: list[str]) -> int:
    ids = [str(x).strip() for x in strategy_ids if str(x).strip()]
    if not ids:
        return 0
    total = 0
    for sid in ids:
        total += _strategy_positions_row_count(db, sid)
    return total


def admin_sync_job_is_resumable(row: dict[str, Any]) -> bool:
    st = str(row.get("status") or "").upper()
    if st == "ABANDONED":
        return False
    if st != "FAILED":
        return False
    try:
        rj = row.get("result_json")
        rj_obj = json.loads(rj) if isinstance(rj, str) and rj.strip() else {}
        return bool(rj_obj.get("resumable") or row.get("checkpoint_json"))
    except json.JSONDecodeError:
        return bool(row.get("checkpoint_json"))


def abandon_strategy_import_job(db: Session, job_id: int) -> None:
    job = get_strategy_import_job_row(db, job_id)
    if not job:
        raise ValueError("strategy import job not found")
    if not strategy_import_job_is_resumable(job):
        raise ValueError("该导入任务不可放弃")
    db.execute(
        text(
            f"""
            UPDATE strategy_import_jobs
            SET status='ABANDONED', finished_at={sql_now()}, message=:m
            WHERE id=:id
            """
        ),
        {"m": "已放弃，不可续传", "id": job_id},
    )


def abandon_admin_sync_job(db: Session, job_id: int) -> None:
    row = db.execute(
        text(
            """
            SELECT id, status, result_json, checkpoint_json
            FROM admin_sync_jobs WHERE id=:id
            """
        ),
        {"id": job_id},
    ).mappings().first()
    if not row:
        raise ValueError("sync job not found")
    d = dict(row)
    if not admin_sync_job_is_resumable(d):
        raise ValueError("该同步任务不可放弃")
    db.execute(
        text(
            f"""
            UPDATE admin_sync_jobs
            SET status='FAILED',
                checkpoint_json=NULL,
                result_json=:result_json,
                finished_at={sql_now()},
                message=:m
            WHERE id=:id
            """
        ),
        {
            "m": "已放弃，不可续传",
            "id": job_id,
            "result_json": json.dumps(
                {"resumable": False, "abandoned": True},
                ensure_ascii=False,
                separators=(",", ":"),
            ),
        },
    )


def reconcile_stale_admin_sync_jobs(db: Session, *, do_commit: bool = True) -> None:
    """将长时间无进度更新的 RUNNING/超时 QUEUED 标为 FAILED，便于续传或重新发起。"""
    stale_mins = max(5, int(getattr(settings, "admin_sync_stale_progress_minutes", 30)))
    stale_rows = (
        db.execute(
            text(
                f"""
                SELECT id FROM admin_sync_jobs
                WHERE status='RUNNING'
                  AND COALESCE(progress_at, started_at, created_at)
                      < {sql_minutes_ago(':mins')}
                """
            ),
            {"mins": stale_mins},
        )
        .scalars()
        .all()
    )
    db.execute(
        text(
            f"""
            UPDATE admin_sync_jobs
            SET status='FAILED', finished_at={sql_now()},
                message=COALESCE(message, '') || '（超过 ' || CAST(:mins AS TEXT)
                    || ' 分钟无进度更新，已标失败；可点续传）'
            WHERE status='RUNNING'
              AND COALESCE(progress_at, started_at, created_at)
                  < {sql_minutes_ago(':mins')}
            """
        ),
        {"mins": stale_mins},
    )
    db.execute(
        text(
            f"""
            UPDATE admin_sync_jobs
            SET status='FAILED', finished_at={sql_now()},
                message='排队超时（未在 2 分钟内启动），请重试或续传'
            WHERE status='QUEUED'
              AND created_at < {sql_minutes_ago(':qmins')}
            """
        ),
        {"qmins": 2},
    )
    for job_id in stale_rows:
        try:
            admin_sync_job_bootstrap_checkpoint(db, int(job_id), do_commit=False)
        except Exception:
            _log.exception("admin_sync job %s: bootstrap checkpoint after stale reconcile", job_id)
    if do_commit:
        db.commit()


def import_strategy_files(
    db: Session,
    selected_strategy_ids: list[str] | None = None,
    import_mode: str = "full",
    do_commit: bool = True,
    *,
    skip_strategy_ids: set[str] | None = None,
    sync_job_id: int | None = None,
    strategy_import_job_id: int | None = None,
    resume: bool = False,
) -> dict:
    selected_set = {str(x).strip() for x in (selected_strategy_ids or []) if str(x).strip()}
    skip = {str(x).strip() for x in (skip_strategy_ids or []) if str(x).strip()}

    if strategy_import_job_id:
        job = get_strategy_import_job_row(db, int(strategy_import_job_id))
        if not job:
            return {"imported": 0, "failed": 1, "errors": ["strategy import job not found"]}
        selected_set = set(_json_str_list(job.get("strategy_ids_json")))
        if not resume:
            skip |= set(_json_str_list(job.get("completed_strategy_ids_json")))
        import_mode = str(job.get("import_mode") or import_mode)

    if selected_set:
        quoted = ",".join("'" + s.replace("'", "''") + "'" for s in sorted(selected_set))
        sql = (
            "SELECT strategy_id, file_dir, file_name FROM strategy_configs "
            f"WHERE status='enabled' AND strategy_id IN ({quoted})"
        )
    else:
        sql = "SELECT strategy_id, file_dir, file_name FROM strategy_configs WHERE status='enabled'"
    configs = db.execute(text(sql)).mappings().all()
    imported_new = 0
    failed = 0
    errors: list[str] = []
    completed: set[str] = set(skip)
    verified_strategy_ids: list[str] = []
    verify_rows: dict[str, str] = {}

    if selected_set:
        found_set = {str(c.get("strategy_id") or "").strip() for c in configs}
        missing = sorted(selected_set - found_set)
        for sid in missing:
            failed += 1
            errors.append(f"{sid}: 配置不存在或未启用，未执行导入")

    total_targets = len(selected_set) if selected_set else len(configs)

    for c in configs:
        raise_if_shutting_down()
        file_path = _strategy_excel_path(c.get("file_dir"), c["file_name"])
        sid = str(c.get("strategy_id") or "").strip()
        if not sid:
            failed += 1
            errors.append(f"(空ID) ({c.get('file_name') or ''}): strategy_id 为空，未执行导入")
            continue
        if sid in skip:
            continue
        if not _SID_PATTERN.match(sid):
            failed += 1
            errors.append(
                f"{sid} ({c.get('file_name') or ''}): strategy_id 非法，仅允许字母/数字/下划线/中划线，且需以字母或数字开头"
            )
            continue
        if not os.path.isfile(file_path):
            failed += 1
            errors.append(f"{sid} ({c['file_name']}): 文件不存在: {file_path}")
            continue
        seq = len(completed) + 1
        stream = _strategy_excel_use_streaming(file_path)
        batch_hint = (
            f"流式 每批{_strategy_excel_row_batch_size()}行 列≤{_strategy_excel_read_max_col()}"
            if stream
            else f"pandas 列≤{_strategy_excel_read_max_col()}"
        )
        _log.info(
            "import_strategy %s file=%s streaming=%s sync_job=%s",
            sid,
            c.get("file_name"),
            stream,
            sync_job_id,
        )
        start_msg = (
            f"阶段1/3 [{seq}/{total_targets}] {sid}（{c.get('file_name') or ''}）"
            f"{batch_hint} 开始…"
        )
        if sync_job_id is not None:
            _admin_sync_job_touch(
                sync_job_id, "import", start_msg, db=db, do_commit=do_commit
            )
        if strategy_import_job_id is not None:
            _strategy_import_job_touch(
                db,
                int(strategy_import_job_id),
                message=start_msg,
                do_commit=do_commit,
            )
        try:
            repair_last_period = bool(resume)
            import_stats: dict[str, int] = {}
            for attempt in range(4):
                try:
                    label_meta, import_stats = _import_strategy_holdings_from_excel(
                        db,
                        sid,
                        file_path,
                        import_mode,
                        sync_job_id=sync_job_id,
                        strategy_import_job_id=strategy_import_job_id,
                        repair_last_period=repair_last_period,
                    )
                    if do_commit:
                        db.commit()
                    break
                except OperationalError as oe:
                    if not _mysql_lock_contention(oe) or attempt >= 3:
                        if do_commit:
                            db.rollback()
                        raise
                    if do_commit:
                        db.rollback()
                    time.sleep(0.25 * (2**attempt))

            imported_new += 1
            rows_to_write = None
            gc.collect()
            completed.add(sid)
            if int(import_stats.get("positions_verified") or 0):
                verified_strategy_ids.append(sid)
                ex_n = int(import_stats.get("verify_slice_ex") or import_stats.get("excel_rows") or 0)
                db_n = int(import_stats.get("verify_slice_db") or import_stats.get("rows_after") or 0)
                verify_rows[sid] = f"{db_n}/{ex_n}" if ex_n else str(db_n)
            try:
                from app.strategy_list_metrics import refresh_strategy_list_metrics_one

                refresh_strategy_list_metrics_one(db, sid, do_commit=do_commit)
            except Exception:
                _log.exception("strategy_list_metrics refresh after import sid=%s", sid)
            done_list = sorted(completed)
            db_rows = import_stats.get("rows_after") or _strategy_positions_row_count(db, sid)
            excel_rows = import_stats.get("excel_rows") or 0
            prog_msg = (
                f"阶段1/3 [{len(completed)}/{total_targets}] 已完成 {sid}"
                f"（库内 {db_rows} 行"
                f"{f'，Excel 约 {excel_rows} 行' if excel_rows else ''}"
                f"，累计 {len(completed)}/{total_targets}，失败 {failed}）"
            )
            _import_progress_touch(
                db,
                prog_msg,
                sync_job_id=sync_job_id,
                strategy_import_job_id=strategy_import_job_id,
                do_commit=False,
            )
            if sync_job_id is not None:
                ck_row = (
                    db.execute(
                        text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
                        {"id": sync_job_id},
                    )
                    .mappings()
                    .first()
                )
                cp = _sync_load_checkpoint(ck_row.get("checkpoint_json") if ck_row else None)
                _sync_save_checkpoint(
                    db,
                    sync_job_id,
                    completed_import=done_list,
                    completed_nav=cp.get("completed_nav") or [],
                    completed_update_rb=cp.get("completed_update_rb") or [],
                    nav_progress=_sync_nav_progress_map(cp.get("nav_progress")),
                    stage="import",
                    do_commit=do_commit,
                )
            if strategy_import_job_id is not None:
                _strategy_import_job_touch(
                    db,
                    int(strategy_import_job_id),
                    message=prog_msg,
                    completed_ids=done_list,
                    imported=len(completed),
                    failed=failed,
                    errors=errors,
                    do_commit=do_commit,
                )
        except Exception as ex:
            if do_commit:
                try:
                    db.rollback()
                except Exception:
                    pass
            failed += 1
            hint = ""
            if _mysql_lock_contention(ex):
                hint = "（可能与「数据更新」或其它导入并发，请错开执行或稍后重试）"
            errors.append(f"{sid} ({c['file_name']}): {ex}{hint}")

    expected_done = total_targets
    resumable = len(completed) < expected_done
    if (
        strategy_import_job_id is not None
        and failed == 0
        and expected_done > 0
        and len(completed) >= expected_done
    ):
        _strategy_import_job_touch(
            db,
            int(strategy_import_job_id),
            message=(
                f"阶段1/3 导入已完成（{len(completed)}/{expected_done}，失败 0）"
            ),
            completed_ids=sorted(completed),
            imported=len(completed),
            failed=0,
            errors=[],
            do_commit=do_commit,
        )
    return {
        "imported": len(completed),
        "imported_new": imported_new,
        "failed": failed,
        "errors": errors[:50],
        "completed_strategy_ids": sorted(completed),
        "verified_strategy_ids": verified_strategy_ids,
        "verify_rows": verify_rows,
        "resumable": resumable,
    }


def run_update(
    db: Session,
    job_type: str,
    triggered_by: str,
    full_refresh: bool = False,
    selected_strategy_ids: list[str] | None = None,
    do_commit: bool = True,
    existing_job_id: int | None = None,
    *,
    skip_nav_rebuild: bool = False,
    sync_job_id: int | None = None,
    skip_update_rebalance_dates: set[str] | None = None,
) -> None:
    global _job_running
    if _job_running:
        raise RuntimeError(
            "另一项数据更新正在本进程内执行（例如后台「立即更新」尚未结束，或与定时任务重叠）。"
            "请待其完成后再操作；全量同步最后一步会占用同一互斥锁。"
        )
    _job_running = True

    if existing_job_id is not None:
        job_id = int(existing_job_id)
    else:
        job_id = db.execute(
            text(
                f"""
                INSERT INTO strategy_update_jobs(job_type,status,triggered_by,started_at)
                VALUES (:jt,'RUNNING',:by,{sql_now()})
                """
            ),
            {"jt": job_type, "by": triggered_by},
        ).lastrowid
        if do_commit:
            db.commit()

    wind = None
    wind_merged: dict[str, Any] | None = None
    work_items: list[dict[str, Any] | None] = []

    def prog(msg: str) -> None:
        _job_progress(db, job_id, msg, do_commit=do_commit, sync_job_id=sync_job_id)

    try:
        wind = wind_sql.open_wind(db)
        mtd = wind.execute(text(wind_sql.sql_max_trade_dt())).mappings().first()
        latest_trade = mtd["d"] if mtd else None
        if not latest_trade:
            raise RuntimeError("No trade date in winddb")
        trade_date = _row_sql_date(latest_trade)
        if trade_date is None:
            trade_date = datetime.strptime(_compact_date(latest_trade), "%Y%m%d").date()

        selected_set = {str(x).strip() for x in (selected_strategy_ids or []) if str(x).strip()}
        if selected_set:
            quoted = ",".join("'" + s.replace("'", "''") + "'" for s in sorted(selected_set))
            configs_sql = f"""
                SELECT c.strategy_id, c.benchmark_code, c.benchmark_name
                FROM strategy_configs c
                WHERE c.status = 'enabled'
                  AND c.strategy_id IN ({quoted})
                ORDER BY (
                    SELECT COUNT(DISTINCT p.rebalance_date)
                    FROM strategy_positions p
                    WHERE p.strategy_id = c.strategy_id
                ) ASC,
                c.strategy_id ASC
                """
        else:
            configs_sql = """
                SELECT c.strategy_id, c.benchmark_code, c.benchmark_name
                FROM strategy_configs c
                WHERE c.status = 'enabled'
                ORDER BY (
                    SELECT COUNT(DISTINCT p.rebalance_date)
                    FROM strategy_positions p
                    WHERE p.strategy_id = c.strategy_id
                ) ASC,
                c.strategy_id ASC
                """
        configs = db.execute(text(configs_sql)).mappings().all()

        src = "远程SQLServer(WindDB)"
        mode_text = "全量重算当日快照" if full_refresh else "增量（默认）"
        scope_text = f"指定策略={len(selected_set)}" if selected_set else "全部启用策略"
        incr_rule = ""
        if not full_refresh:
            incr_rule = (
                "；约定=以库末净值为基准、调仓日≤末净值日的最近一期为锚，"
                "净值仅补末净值日之后、持仓仅锚定及之后拉 Wind（非从第 1 期重拉）"
            )
        prog(
            f"Wind源={src} 最新交易日={trade_date} 模式={mode_text} 范围={scope_text}{incr_rule}，"
            f"待处理策略数={len(configs)}"
        )

        for cfg in configs:
            work_items.append(
                _run_update_try_build_work_item(
                    db,
                    cfg,
                    trade_date,
                    full_refresh,
                    job_id,
                    do_commit,
                    sync_job_id=sync_job_id,
                )
            )
        active = [w for w in work_items if w is not None]
        n_active = len(active)
        n_union = len({c for w in active for c in w["stock_codes"]}) if active else 0
        if active and _use_wind_merged_prefetch(n_active, n_union):
            prog(
                f"合并拉取 Wind：{n_active} 个策略、{n_union} 只不重复股票（EOD+行情+指数一次批量）…"
            )
            wind, wind_merged = _run_update_prefetch_wind_merged(
                db, wind, active, str(latest_trade)
            )
        elif active:
            prog(
                f"按策略串行拉 Wind（共 {n_active} 个策略，"
                f"去重 {n_union} 只；合并预拉已关闭，避免多策略占满内存）…"
            )

        i_active = 0
        for wi in work_items:
            if wi is None:
                continue
            i_active += 1
            sid = wi["sid"]
            skip_holdings = bool(wi.get("skip_holdings"))
            bench_code = wi["bench_code"]
            latest_rb = wi["latest_rb"]
            latest_rb_compact = wi["latest_rb_compact"]
            n_rb = wi["n_rb"]
            rb_positions = wi["rb_positions"]
            stock_codes = wi["stock_codes"]
            start_c = wi["start_c"]
            hold_start_idx = int(wi.get("hold_start_idx") or 0)
            anchor_rb_hold = wi.get("anchor_rb_hold")
            hold_scope_note = str(wi.get("hold_scope_note") or "")
            wind_rb_indices: list[int] = list(wi.get("wind_rb_indices") or [])
            n_rb_wind = int(wi.get("n_rb_wind") or len(wind_rb_indices))
            wind_stock_codes: list[str] = list(
                wi.get("wind_stock_codes") or stock_codes
            )
            holding_desc_bars = (
                wind_bulk.holding_eod_desc_max_bars()
                if not full_refresh and wind_rb_indices
                else 280
            )
            if skip_holdings:
                prog(f"[{i_active}/{n_active}] 策略 {sid}：跳过持仓（已齐），仅补净值…")
            else:
                prog(
                    f"[{i_active}/{n_active}] 策略 {sid}：开始处理…"
                    + (f" {hold_scope_note}" if hold_scope_note else "")
                )

            if not skip_nav_rebuild:
                try:
                    plans1 = _batch_nav_mysql_plans(db, [sid], "full")
                    pl1 = plans1.get(sid)
                    if pl1 and not pl1.get("code_set"):
                        raise RuntimeError(
                            f"{sid} 无 strategy_positions 仓位，无法算净值；请先全量导入"
                        )
                    if pl1:
                        last_nav_c = _last_nav_compact_for_update(db, sid)
                        if (
                            last_nav_c
                            and _compact_date(latest_trade) <= last_nav_c
                            and not full_refresh
                        ):
                            prog(
                                f"[{i_active}/{n_active}] {sid}：净值已至 {last_nav_c}，无需补算"
                            )
                        else:
                            lt_c = _compact_date(latest_trade)
                            if full_refresh:
                                nav_hint = f"全量重算 {pl1['start_c']}~{lt_c}（删库后从首调仓）"
                            elif last_nav_c and len(last_nav_c) >= 8:
                                nav_hint = (
                                    f"末净值 {last_nav_c}（nav×本金 bootstrap），"
                                    f"仅补写之后交易日至 {lt_c}；不回放全历史"
                                )
                            else:
                                nav_hint = f"首建全量 {pl1['start_c']}~{lt_c}"
                            prog(
                                f"[{i_active}/{n_active}] {sid}：净值 {nav_hint}"
                                f"（名义本金 {settings.strategy_nav_initial_capital:g} 元）…"
                            )
                            nav_mode = "full" if full_refresh else "incremental"
                            ok_nav, wind = _rebuild_nav_for_strategy(
                                db,
                                wind,
                                sid,
                                nav_mode,
                                latest_trade_c_cached=str(latest_trade),
                                mysql_plan=pl1,
                                wind_bundle=None,
                                nav_full_rebuild=bool(full_refresh),
                                nav_force_reset=bool(full_refresh),
                                sync_job_id=sync_job_id,
                                progress_cb=prog if sync_job_id is None else None,
                            )
                            nav_max_c = _strategy_nav_max_trade_compact(db, sid)
                            lt_c = _compact_date(latest_trade)
                            if not ok_nav:
                                raise RuntimeError(
                                    f"{sid} 净值重建未完成（增量与自动全量回退均失败："
                                    "请确认 Wind 可用；或勾选「全量重算净值」后重试；"
                                    "末净值日缺持仓快照时须先更新持仓）"
                                )
                            if nav_max_c and lt_c and nav_max_c < lt_c:
                                raise RuntimeError(
                                    f"{sid} 净值仅至 {nav_max_c}，未到 Wind 最新 {lt_c}"
                                    "（常见原因：末净值日无持仓快照导致增量尺度回滚；"
                                    "全量更新将删净值重算，或先跑持仓更新）"
                                )
                            prog(
                                f"[{i_active}/{n_active}] {sid}：净值已更新至 {nav_max_c or lt_c}"
                            )
                except Exception as ex_nav:
                    _log.warning("nav rebuild failed for %s: %s", sid, ex_nav)
                    prog(
                        f"[{i_active}/{n_active}] {sid}：净值重算失败：{ex_nav}"[:6000]
                    )
                    raise
                if do_commit:
                    db.commit()

            if skip_holdings:
                from app.strategy_list_metrics import refresh_strategy_list_metrics_safe

                refresh_strategy_list_metrics_safe(
                    db,
                    sid,
                    do_commit=do_commit,
                    stock_count_on_last_date=_effective_rebalance_stock_count(
                        rb_positions, as_of=trade_date
                    ),
                    last_trade_date=trade_date.isoformat(),
                )
                _release_run_update_strategy_memory(
                    eod_by_code={},
                    index_eod_by_code={},
                    quote_map={},
                    eod_local=None,
                )
                continue

            eod_by_code: dict[str, list] = {}
            index_eod_by_code: dict[str, list] = {}
            quote_map: dict[str, Any] = {}
            rb_chunked_eod = (
                wind_merged is None
                and bool(getattr(settings, "update_eod_per_rebalance_chunk", True))
                and (_wind_low_memory_mode() or n_active > 1)
            )

            if wind_merged is not None:
                full_eod = wind_merged["eod"]
                eod_by_code = {c: full_eod[c] for c in stock_codes if c in full_eod}
                if bench_code and bench_code in wind_merged["idx"]:
                    index_eod_by_code[bench_code] = wind_merged["idx"][bench_code]
                full_qu = wind_merged["quote"]
                quote_map = {k: full_qu[k] for k in stock_codes if k in full_qu}
            elif rb_chunked_eod:
                if n_rb_wind <= 0:
                    prog(f"[{i_active}/{n_active}] {sid}：无待拉 Wind 的调仓期")
                else:
                    prog(
                        f"[{i_active}/{n_active}] {sid}：持仓增量，"
                        f"仅 {n_rb_wind} 期拉 EOD（非 {n_rb} 期全历史）…"
                    )
            else:
                eod_codes = wind_stock_codes if wind_stock_codes else stock_codes
                prog(
                    f"[{i_active}/{n_active}] {sid}：串行拉取 A 股 EOD {len(eod_codes)} 只"
                    f"{(' + 指数 1 只' if bench_code else '')} × 区间 {start_c}~{latest_trade} …"
                )
                wind, eod_by_code = wind_bulk.load_eod_by_code(
                    wind, eod_codes, start_c, str(latest_trade), db
                )
                if bench_code:
                    wind, index_eod_by_code = wind_bulk.load_index_eod_by_code(
                        wind, [bench_code], start_c, str(latest_trade), db
                    )
                wind, quote_map = _fetch_wind_quote_map_batched(
                    db, wind, eod_codes, latest_trade
                )

            eod_local = eod_by_code if wind_merged is None and not rb_chunked_eod else None
            skip_rb = _sync_update_rb_done_for_strategy(
                skip_update_rebalance_dates or set(),
                sid,
                allow_legacy_dates=(len(selected_set) <= 1),
            )
            if hold_start_idx > 0 and anchor_rb_hold is not None and not full_refresh:
                prior_td = _holding_prior_trade_date(db, sid, trade_date)
                rb_copy: list[date] = []
                for i_rb_c, (rb_c, _) in enumerate(rb_positions):
                    if i_rb_c < hold_start_idx:
                        rb_copy.append(rb_c)
                    elif i_rb_c not in wind_rb_indices:
                        rb_copy.append(rb_c)
                if prior_td is not None and rb_copy:
                    n_copy = _copy_holding_daily_rebalances(
                        db,
                        sid=sid,
                        from_trade_date=prior_td,
                        to_trade_date=trade_date,
                        rebalance_dates=rb_copy,
                        do_commit=do_commit,
                    )
                    prog(
                        f"[{i_active}/{n_active}] {sid}：持仓 {len(rb_copy)} 期沿用 {prior_td}"
                        f"（{n_copy} 条），仅 {n_rb_wind} 个开放期拉 Wind"
                        f"（EOD 约 {wind_bulk.holding_eod_lookback_calendar_days()} 日回溯）"
                    )
                elif n_rb_wind:
                    prog(
                        f"[{i_active}/{n_active}] {sid}：持仓仅 {n_rb_wind} 个开放期拉 Wind"
                        f"（锚定 {_compact_date(anchor_rb_hold)}，无上一日可沿用）"
                    )
            else:
                prog(
                    f"[{i_active}/{n_active}] {sid}：共 {n_rb} 个调仓期写入持仓"
                    f"（{'全量刷新' if full_refresh else '无末净值'}，逐期拉 Wind）…"
                )

            def _flush_one_rebalance(
                i_rb: int,
                rebalance,
                prepared_rows: list[dict[str, Any]],
                total_weight: float,
                rb_compact: str,
                wind_i: int,
                expected_codes: list[str] | set[str],
            ) -> None:
                n_written = _flush_rebalance_holding_period(
                    db,
                    sid=sid,
                    trade_date=trade_date,
                    rebalance=rebalance,
                    i_rb=i_rb,
                    prepared_rows=prepared_rows,
                    total_weight=total_weight,
                    do_commit=do_commit,
                )
                prog(
                    f"[{i_active}/{n_active}] {sid} Wind {wind_i}/{n_rb_wind} "
                    f"调仓 {rebalance} 已写入 {n_written} 条"
                )
                if sync_job_id is not None and rb_compact:
                    if not _holding_snapshot_complete(
                        db,
                        sid=sid,
                        trade_date=trade_date,
                        rebalance=rebalance,
                        expected_codes=expected_codes,
                    ):
                        raise RuntimeError(
                            f"{sid} 调仓 {rb_compact} 持仓快照未完整写入，未记录续传断点"
                        )
                    _sync_mark_update_rb_done(
                        sync_job_id,
                        sid,
                        rb_compact,
                        db=db,
                        do_commit=False,
                    )

            if rb_chunked_eod:
                # 按调仓期：只拉该期成分股 EOD，凑齐一期即算权重并落库（峰值内存≈单期行数+一小批 K 线）
                for i_rb, (rebalance, positions) in enumerate(rb_positions, start=1):
                    if (i_rb - 1) not in wind_rb_indices:
                        continue
                    next_rebalance = (
                        rb_positions[i_rb][0] if i_rb < len(rb_positions) else None
                    )
                    period_end_c = _holding_period_end_compact(
                        next_rebalance, latest_trade
                    )
                    if not positions:
                        prog(
                            f"[{i_active}/{n_active}] {sid} [{i_rb}/{n_rb}] "
                            f"调仓日 {rebalance} 无仓位，跳过"
                        )
                        continue
                    rb_compact = _compact_date(rebalance)
                    checkpoint_codes = [
                        _wind_code_key(p["stock_code"])
                        for p in positions
                        if p.get("stock_code")
                    ]
                    if (
                        rb_compact
                        and rb_compact in skip_rb
                        and _holding_snapshot_complete(
                            db,
                            sid=sid,
                            trade_date=trade_date,
                            rebalance=rebalance,
                            expected_codes=checkpoint_codes,
                        )
                    ):
                        prog(
                            f"[{i_active}/{n_active}] {sid} [{i_rb}/{n_rb}] 调仓 {rebalance} "
                            f"已跳过（断点已完成）"
                        )
                        continue
                    total_weight = sum(
                        max(float(p["holding_weight"] or 0.0), 0.0) for p in positions
                    )
                    period_codes = sorted(
                        {
                            _wind_code_key(p["stock_code"])
                            for p in positions
                            if p.get("stock_code")
                        }
                    )
                    if not period_codes:
                        prog(
                            f"[{i_active}/{n_active}] {sid} [{i_rb}/{n_rb}] "
                            f"调仓日 {rebalance} 无有效代码，跳过"
                        )
                        continue
                    # 本期收益：调仓日→下一调仓日（末期为最新交易日）；每期仅该期成分×短区间，可一次拉全
                    if full_refresh and period_end_c:
                        period_start_c = wind_bulk.bulk_eod_start_compact(
                            period_end_c, rebalance
                        )
                    else:
                        period_start_c = wind_bulk.holding_eod_start_for_period(
                            trade_date, rebalance, full_refresh=full_refresh
                        )
                    eod_load_end_c = period_end_c or str(latest_trade)
                    wind_i = wind_rb_indices.index(i_rb - 1) + 1
                    prog(
                        f"[{i_active}/{n_active}] {sid} Wind {wind_i}/{n_rb_wind} "
                        f"调仓 {rebalance} EOD {len(period_codes)} 只"
                        f"（{period_start_c}~{eod_load_end_c}）…"
                    )
                    if sync_job_id is not None:
                        _admin_sync_job_touch(
                            sync_job_id,
                            "update",
                            f"[{i_active}/{n_active}] {sid} Wind {wind_i}/{n_rb_wind} "
                            f"调仓 {rebalance} EOD {len(period_codes)} 只…",
                            do_commit=True,
                        )
                    prepared_rows: list[dict[str, Any]] = []
                    quote_td = _holding_quote_td_compact(period_end_c, latest_trade)
                    wind, quote_part = _fetch_wind_quote_map_batched(
                        db, wind, period_codes, quote_td
                    )
                    wind, eod_part = wind_bulk.load_eod_by_code(
                        wind, period_codes, period_start_c, eod_load_end_c, db
                    )
                    for p in positions:
                        wk = _wind_code_key(p["stock_code"])
                        prepared_rows.append(
                            _build_holding_daily_row_from_wind(
                                sid=sid,
                                trade_date=trade_date,
                                rebalance=rebalance,
                                p=p,
                                quote_map=quote_part,
                                eod_series=eod_part.get(wk, []),
                                latest_trade=str(latest_trade),
                                period_end_compact=period_end_c,
                                desc_max_bars=holding_desc_bars,
                            )
                        )
                    quote_part.clear()
                    eod_part.clear()
                    del eod_part
                    gc.collect()
                    _flush_one_rebalance(
                        i_rb,
                        rebalance,
                        prepared_rows,
                        total_weight,
                        rb_compact,
                        wind_i,
                        period_codes,
                    )
                    prepared_rows.clear()
                if sync_job_id is not None and do_commit:
                    db.commit()
                    prog(
                        f"[{i_active}/{n_active}] {sid}：{n_rb_wind} 个调仓期持仓已落库，正在收尾…"
                    )
            else:
                for i_rb, (rebalance, positions) in enumerate(rb_positions, start=1):
                    if (i_rb - 1) not in wind_rb_indices:
                        continue
                    if not positions:
                        prog(
                            f"[{i_active}/{n_active}] {sid} [{i_rb}/{n_rb}] "
                            f"调仓日 {rebalance} 无仓位，跳过"
                        )
                        continue
                    rb_compact = _compact_date(rebalance)
                    checkpoint_codes = [
                        _wind_code_key(p["stock_code"])
                        for p in positions
                        if p.get("stock_code")
                    ]
                    if (
                        rb_compact
                        and rb_compact in skip_rb
                        and _holding_snapshot_complete(
                            db,
                            sid=sid,
                            trade_date=trade_date,
                            rebalance=rebalance,
                            expected_codes=checkpoint_codes,
                        )
                    ):
                        prog(
                            f"[{i_active}/{n_active}] {sid} [{i_rb}/{n_rb}] 调仓 {rebalance} "
                            f"已跳过（断点已完成）"
                        )
                        continue
                    next_rebalance = (
                        rb_positions[i_rb][0] if i_rb < len(rb_positions) else None
                    )
                    period_end_c = _holding_period_end_compact(
                        next_rebalance, latest_trade
                    )

                    prepared_rows = []
                    total_weight = 0.0
                    period_codes_q = sorted(
                        {
                            _wind_code_key(p["stock_code"])
                            for p in positions
                            if p.get("stock_code")
                        }
                    )
                    quote_td = _holding_quote_td_compact(period_end_c, latest_trade)
                    if period_end_c and quote_td != _compact_date(latest_trade):
                        wind, quote_period = _fetch_wind_quote_map_batched(
                            db, wind, period_codes_q, quote_td
                        )
                        qmap = quote_period
                    else:
                        qmap = quote_map
                    for p in positions:
                        total_weight += max(float(p["holding_weight"] or 0.0), 0.0)
                        wk = _wind_code_key(p["stock_code"])
                        prepared_rows.append(
                            _build_holding_daily_row_from_wind(
                                sid=sid,
                                trade_date=trade_date,
                                rebalance=rebalance,
                                p=p,
                                quote_map=qmap,
                                eod_series=eod_by_code.get(wk, []),
                                latest_trade=str(latest_trade),
                                period_end_compact=period_end_c,
                                desc_max_bars=holding_desc_bars,
                            )
                        )
                    if period_end_c and quote_td != _compact_date(latest_trade):
                        quote_period.clear()
                    wind_i = wind_rb_indices.index(i_rb - 1) + 1
                    _flush_one_rebalance(
                        i_rb,
                        rebalance,
                        prepared_rows,
                        total_weight,
                        rb_compact,
                        wind_i,
                        [_wind_code_key(p["stock_code"]) for p in positions if p.get("stock_code")],
                    )
                if sync_job_id is not None and do_commit:
                    db.commit()
                    prog(
                        f"[{i_active}/{n_active}] {sid}：{n_rb_wind} 个调仓期持仓已落库，正在收尾…"
                    )

            if do_commit:
                if sync_job_id is None:
                    prog(f"[{i_active}/{n_active}] {sid}：持仓处理完成")
                db.commit()
            from app.strategy_list_metrics import refresh_strategy_list_metrics_safe

            refresh_strategy_list_metrics_safe(
                db,
                sid,
                do_commit=do_commit,
                stock_count_on_last_date=_effective_rebalance_stock_count(
                    rb_positions, as_of=trade_date
                ),
                last_trade_date=trade_date.isoformat(),
            )
            _release_run_update_strategy_memory(
                eod_by_code=eod_by_code,
                index_eod_by_code=index_eod_by_code,
                quote_map=quote_map,
                eod_local=eod_local,
            )

        if wind_merged is not None:
            _release_wind_memory(wind_merged)
            wind_merged = None

        done_msg = f"全部完成（处理 {len(active)} 个策略，行情日 {trade_date}）"
        if not selected_strategy_ids:
            try:
                from app.strategy_list_metrics import prune_strategy_list_metrics_orphans

                prune_strategy_list_metrics_orphans(db, do_commit=False)
            except Exception:
                _log.exception("strategy_list_metrics prune after run_update failed")
        db.execute(
            text(
                f"""
                UPDATE strategy_update_jobs
                SET status='SUCCESS', finished_at={sql_now()}, message=:m
                WHERE id=:id
                """
            ),
            {"m": done_msg, "id": job_id},
        )
        if do_commit:
            db.commit()
        if sync_job_id is not None:
            _admin_sync_job_touch(
                sync_job_id, "holding_update", done_msg, db=db, do_commit=do_commit
            )
    except Exception as ex:
        if do_commit:
            try:
                db.rollback()
            except Exception:
                pass
        _mark_update_job_failed(db, job_id, _format_update_job_failure_message(ex), do_commit)
        raise
    finally:
        _job_running = False
        if wind_merged is not None:
            _release_wind_memory(wind_merged)
            wind_merged = None
        if wind is not None:
            wind_sql.close_wind_safe(wind, db)
            wind = None
        work_items.clear()
        gc.collect()


# 净值日序列写入：逐条 INSERT 时 MySQL 往返次数 ≈ 交易日数；批量 executemany 可显著降低耗时
_NAV_INSERT_CHUNK = 80

_NAV_INSERT_SQL = text(
    """
    INSERT INTO strategy_nav_daily(
      strategy_id, trade_date, nav_unit, daily_ret,
      benchmark_ret, benchmark_nav, rebalance_date, source_job_id
    ) VALUES (
      :sid, :td, :nav, :ret, :bret, :bnav, :rb, NULL
    )
    ON CONFLICT(strategy_id, trade_date) DO UPDATE SET
      nav_unit=excluded.nav_unit,
      daily_ret=excluded.daily_ret,
      benchmark_ret=excluded.benchmark_ret,
      benchmark_nav=excluded.benchmark_nav,
      rebalance_date=excluded.rebalance_date
    """
)


def _flush_strategy_nav_daily_batch(db: Session, acc: list[dict[str, Any]]) -> None:
    if not acc:
        return
    for i in range(0, len(acc), _NAV_INSERT_CHUNK):
        chunk = acc[i : i + _NAV_INSERT_CHUNK]
        db.execute(_NAV_INSERT_SQL, chunk)


def _flush_strategy_nav_daily_batch_verified(
    db: Session,
    acc: list[dict[str, Any]],
    *,
    sid: str,
    expected_trade_days: set[str],
) -> None:
    if not acc:
        return
    batch_days = {
        _compact_date(row.get("td"))
        for row in acc
        if row.get("td") is not None and _compact_date(row.get("td"))
    }
    if not batch_days:
        _flush_strategy_nav_daily_batch(db, acc)
        return
    start_c = min(batch_days)
    end_c = max(batch_days)
    expected = sum(1 for d in expected_trade_days if start_c <= d <= end_c)
    for attempt in range(2):
        _flush_strategy_nav_daily_batch(db, acc)
        db.commit()
        actual = _strategy_nav_trade_day_count(db, sid, start_c, end_c)
        if actual >= expected:
            return
        _log.warning(
            "nav %s: flush verify short %s..%s actual=%s expected=%s attempt=%s",
            sid,
            start_c,
            end_c,
            actual,
            expected,
            attempt + 1,
        )
    raise RuntimeError(
        f"nav {sid}: persisted rows incomplete for {start_c}..{end_c} "
        f"({actual}/{expected})"
    )


def _strategy_nav_max_trade_compact(db: Session, sid: str) -> str | None:
    """strategy_nav_daily 已有净值的最后交易日（YYYYMMDD compact）。"""
    row = db.execute(
        text(
            f"""
            SELECT {sql_max_date_expr("trade_date")} AS d
            FROM strategy_nav_daily
            WHERE strategy_id=:sid
            """
        ),
        {"sid": sid},
    ).mappings().first()
    if not row or row.get("d") is None:
        return None
    d = _row_sql_date(row["d"])
    if d is None:
        return None
    c = _compact_date(d)
    return c if len(c) >= 8 else None


def _strategy_nav_trade_day_count(
    db: Session,
    sid: str,
    start_c: str,
    end_c: str,
) -> int:
    if not sid or not start_c or not end_c:
        return 0
    row = db.execute(
        text(
            f"""
            SELECT COUNT(DISTINCT {sql_date_compact_expr("trade_date")}) AS n
            FROM strategy_nav_daily
            WHERE strategy_id=:sid
              AND {sql_date_compact_expr("trade_date")} BETWEEN :start_c AND :end_c
            """
        ),
        {"sid": sid, "start_c": start_c[:8], "end_c": end_c[:8]},
    ).mappings().first()
    try:
        return int((row or {}).get("n") or 0)
    except (TypeError, ValueError):
        return 0


def _nav_stored_range_complete(
    db: Session,
    sid: str,
    start_c: str,
    end_c: str,
    trade_days: list[str],
    *,
    label: str,
) -> bool:
    expected = sum(1 for d in trade_days if start_c <= d <= end_c)
    actual = _strategy_nav_trade_day_count(db, sid, start_c, end_c)
    if actual == expected:
        return True
    _log.warning(
        "nav %s: stored range incomplete for %s %s..%s actual=%s expected=%s",
        sid,
        label,
        start_c,
        end_c,
        actual,
        expected,
    )
    return False


def _nav_persist_after_compact(db: Session, sid: str, nav_full_rebuild: bool) -> str | None:
    """
    增量净值：返回已有最后交易日 compact；仅写入该日之后的交易日。
    全量模式返回 None（从首日到最新日均写入）。
    """
    if nav_full_rebuild:
        return None
    return _strategy_nav_max_trade_compact(db, sid)


def _batch_nav_mysql_plans(db: Session, strategy_ids: list[str], _mode_l: str) -> dict[str, dict[str, Any]]:
    """一次查询多策略的持仓与配置，供 rebuild_nav_series 合并 Wind 拉取。"""
    sids = [str(x).strip() for x in strategy_ids if str(x or "").strip()]
    if not sids:
        return {}
    quoted = ",".join("'" + s.replace("'", "''") + "'" for s in sids)
    pos_rows = db.execute(
        text(
            f"""
            SELECT strategy_id, rebalance_date, stock_code, holding_weight
            FROM strategy_positions
            WHERE strategy_id IN ({quoted})
            ORDER BY strategy_id, {sql_order_date_asc("rebalance_date")}, stock_code
            """
        )
    ).mappings().all()
    by_sid: dict[str, list[Any]] = defaultdict(list)
    for r in pos_rows:
        by_sid[str(r["strategy_id"]).strip()].append(r)

    cfg_rows = db.execute(
        text(
            f"""
            SELECT strategy_id, benchmark_code
            FROM strategy_configs
            WHERE strategy_id IN ({quoted})
            """
        )
    ).mappings().all()
    bench_by: dict[str, str] = {}
    for r in cfg_rows:
        sid = str(r["strategy_id"]).strip()
        bench_by[sid] = str(r.get("benchmark_code") or "").strip().upper()

    plans: dict[str, dict[str, Any]] = {}
    for sid in sids:
        rows = by_sid.get(sid) or []
        if not rows:
            continue
        rb_map: dict[date, list[tuple[str, float]]] = {}
        code_set: set[str] = set()
        for r in rows:
            rd = _row_sql_date(r["rebalance_date"])
            if rd is None:
                continue
            sc = str(r["stock_code"]).strip().upper()
            w = float(r.get("holding_weight") or 0.0)
            rb_map.setdefault(rd, []).append((sc, w))
            if sc:
                code_set.add(sc)
        if not code_set:
            continue
        rb_sorted = sorted(rb_map.keys())
        min_rb = rb_sorted[0]
        start_c = _compact_date(min_rb)
        plans[sid] = {
            "rb_map": rb_map,
            "code_set": code_set,
            "bench_code": bench_by.get(sid, ""),
            "start_c": start_c,
            "min_rb": min_rb,
        }
    return plans


def _eod_dict_to_day_map(
    eod_by_code: dict[str, list],
) -> dict[str, dict[str, tuple[float | None, float | None]]]:
    day_map: dict[str, dict[str, tuple[float | None, float | None]]] = {}
    for c, series in eod_by_code.items():
        dct: dict[str, tuple[float | None, float | None]] = {}
        for d, cl, pc, _raw in series:
            clv = None if (isinstance(cl, float) and cl != cl) else float(cl)
            pcv = None if (isinstance(pc, float) and pc != pc) else float(pc)
            dct[_compact_date(d)] = (clv, pcv)
        day_map[c] = dct
    return day_map


def _eod_day_map_has_trade(
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    td_compact: str,
    codes: list[str] | None = None,
) -> bool:
    """day_map 按 wind_code 索引；检查指定 compact 交易日是否至少有一只成分有有效收盘价。"""
    td = str(td_compact or "").strip()[:8]
    if not td or not day_map:
        return False

    def _ok(dct: dict[str, tuple[float | None, float | None]] | None) -> bool:
        if not dct or td not in dct:
            return False
        cl, _ = dct[td]
        return cl is not None and not (isinstance(cl, float) and cl != cl) and cl > 0

    if codes:
        for raw in codes:
            c = str(raw or "").strip().upper()
            if c and _ok(day_map.get(c)):
                return True
        return False
    return any(_ok(dct) for dct in day_map.values())


def _latest_rebalance_stock_count(rb_map: dict[date, list[tuple[str, float]]]) -> int:
    """最新调仓期成分股数（去重），用于动态净值 EOD 分段月数。"""
    if not rb_map:
        return 1
    latest_rb = max(rb_map.keys())
    return len({str(c).strip().upper() for c, _ in rb_map.get(latest_rb, ()) if c})


def resolve_nav_rebuild_eod_months(latest_period_stock_count: int) -> int:
    """
    开算净值前动态月数：latest_n * months <= nav_rebuild_stock_month_budget（至少 1 个月）。
    可选 nav_rebuild_eod_months_max / 旧 nav_rebuild_eod_months>0 作为上限。
    """
    configured_budget = max(1, int(getattr(settings, "nav_rebuild_stock_month_budget", 300) or 300))
    budget_floor = max(0, int(getattr(settings, "nav_rebuild_stock_month_budget_floor", 360) or 0))
    budget = max(configured_budget, budget_floor)
    n = max(1, int(latest_period_stock_count or 1))
    months = max(1, budget // n)
    cap = int(getattr(settings, "nav_rebuild_eod_months_max", 0) or 0)
    if cap <= 0:
        legacy = int(getattr(settings, "nav_rebuild_eod_months", 0) or 0)
        if legacy > 0:
            cap = legacy
    if cap > 0:
        months = min(months, cap)
    return months


def _nav_eod_time_segments(
    start_c: str,
    latest_trade_c: str,
    *,
    step_months: int | None = None,
) -> list[tuple[str, str]]:
    return wind_bulk.eod_range_segments(start_c, latest_trade_c, step_months=step_months)


def _codes_for_nav_segment(
    rb_sorted: list[date],
    rb_map: dict[date, list[tuple[str, float]]],
    shares: dict[str, float],
    seg_st_compact: str,
    seg_ed_compact: str,
) -> list[str]:
    """
    仅拉与本段时间窗有交集的调仓期成分 + 当前模拟持仓股。
    禁止「截至 seg_ed 的全部历史成分」（否则第 13 段也会拉到 676 只）。
    """
    seg_start = datetime.strptime(str(seg_st_compact).strip()[:8], "%Y%m%d").date()
    seg_end = datetime.strptime(str(seg_ed_compact).strip()[:8], "%Y%m%d").date()
    codes: set[str] = {c for c in shares if c}
    for i, rb_d in enumerate(rb_sorted):
        if rb_d > seg_end:
            break
        next_rb = rb_sorted[i + 1] if i + 1 < len(rb_sorted) else None
        # 调仓有效期 [rb_d, next_rb) 与 [seg_start, seg_end] 无交集则跳过
        if next_rb is not None and next_rb <= seg_start:
            continue
        for c, _w in rb_map.get(rb_d, ()):
            if c:
                codes.add(c)
    return sorted(codes)


def _nav_ensure_anchor_eod_in_day_map(
    wind: Any,
    db: Session,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    codes: list[str],
    anchor_c: str,
) -> Any:
    """
    分段 EOD 首段起点可能早于末净值日（如新调仓前移 eod_start）；
    单独拉取锚点日复权价，保证 bootstrap 与全量重算同价体系。
    """
    anchor_c = str(anchor_c or "").strip()[:8]
    if len(anchor_c) < 8 or not codes:
        return wind
    if _eod_day_map_has_trade(day_map, anchor_c, codes):
        return wind
    chunk_sz = wind_bulk.eod_stock_chunk_size()
    for i in range(0, len(codes), chunk_sz):
        part = codes[i : i + chunk_sz]
        if not part:
            continue
        wind, eod_part = wind_bulk.load_eod_by_code(wind, part, anchor_c, anchor_c, db)
        _merge_eod_into_day_map(day_map, eod_part)
        eod_part.clear()
    return wind


def _merge_eod_into_day_map(
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    eod_part: dict[str, list],
) -> None:
    for c, series in eod_part.items():
        dct: dict[str, tuple[float | None, float | None]] = {}
        for d, cl, pc, _raw in series:
            clv = None if (isinstance(cl, float) and cl != cl) else float(cl)
            pcv = None if (isinstance(pc, float) and pc != pc) else float(pc)
            dct[_compact_date(d)] = (clv, pcv)
        day_map[c] = dct


def _load_segment_wind_maps(
    wind: Any,
    db: Session,
    codes: list[str],
    seg_st: str,
    seg_ed: str,
    bench_code: str,
) -> tuple[Any, dict[str, dict[str, tuple[float | None, float | None]]], dict[str, tuple[float | None, float | None]]]:
    """按股票小批拉 EOD 并入 day_map，避免一次 load 全成分。"""
    day_map: dict[str, dict[str, tuple[float | None, float | None]]] = {}
    chunk_sz = wind_bulk.eod_stock_chunk_size()
    for i in range(0, len(codes), chunk_sz):
        raise_if_shutting_down()
        part = codes[i : i + chunk_sz]
        if not part:
            continue
        wind, eod_part = wind_bulk.load_eod_by_code(wind, part, seg_st, seg_ed, db)
        _merge_eod_into_day_map(day_map, eod_part)
        eod_part.clear()
        if _wind_low_memory_mode():
            gc.collect()
    bench_day_map: dict[str, tuple[float | None, float | None]] = {}
    if bench_code:
        wind, index_eod_by_code = wind_bulk.load_index_eod_by_code(
            wind, [bench_code], seg_st, seg_ed, db
        )
        bench_day_map = _index_eod_to_bench_day_map(index_eod_by_code, bench_code)
        index_eod_by_code.clear()
    return wind, day_map, bench_day_map


def _index_eod_to_bench_day_map(
    index_eod_by_code: dict[str, list], bench_code: str
) -> dict[str, tuple[float | None, float | None]]:
    bench_day_map: dict[str, tuple[float | None, float | None]] = {}
    if not bench_code:
        return bench_day_map
    for d, cl, pc, _raw in index_eod_by_code.get(_wind_code_key(bench_code), []):
        clv = None if (isinstance(cl, float) and cl != cl) else float(cl)
        pcv = None if (isinstance(pc, float) and pc != pc) else float(pc)
        bench_day_map[_compact_date(d)] = (clv, pcv)
    return bench_day_map


def _bench_quads_for_code(index_eod_by_code: dict[str, list], bench_code: str) -> list:
    if not bench_code:
        return []
    return list(index_eod_by_code.get(_wind_code_key(bench_code), ()))


def _bench_return_on_trade_day(
    td: str,
    *,
    bench_quads: list | None = None,
    bench_day_map: dict[str, tuple[float | None, float | None]] | None = None,
) -> float | None:
    """用全区间指数序列算日收益（分段拉行情时也能拿到段首日的真实昨收）。"""
    if bench_quads:
        clv, pcv = wind_bulk.index_close_preclose_for_compact_day(bench_quads, td)
        return _safe_return(clv, pcv)
    if bench_day_map:
        bp = bench_day_map.get(td)
        if bp:
            return _safe_return(bp[0], bp[1])
    return None


def _step_benchmark_nav_acc(
    bench_nav_acc: float | None,
    br: float | None,
    *,
    allow_flat: bool,
) -> tuple[float | None, float | None, float | None]:
    """返回 (新累计, benchmark_ret, benchmark_nav)。allow_flat：Wind 有指数但当日算不出收益时按 0 延续。"""
    if bench_nav_acc is None:
        return None, None, None
    if br is not None:
        acc = float(bench_nav_acc) * (1.0 + float(br))
        return acc, float(br), acc
    if allow_flat:
        acc = float(bench_nav_acc)
        return acc, 0.0, acc
    return float(bench_nav_acc), None, None


def _nav_incremental_from_period_enabled() -> bool:
    return bool(getattr(settings, "nav_incremental_from_current_period", True))


def _nav_rb_idx_on_date(rb_sorted: list[date], td_date: date) -> tuple[int, date]:
    """截至 asof 日（含）仍有效的最近一期调仓。"""
    rb_idx = 0
    current_rb = rb_sorted[0]
    for i, rb_d in enumerate(rb_sorted):
        if rb_d <= td_date:
            rb_idx = i
            current_rb = rb_d
        else:
            break
    return rb_idx, current_rb


def _nav_union_codes_between_rebalances(
    rb_map: dict[date, list[tuple[str, float]]],
    rb_sorted: list[date],
    rb_from: date,
    rb_to: date,
) -> list[str]:
    codes: set[str] = set()
    for rb in rb_sorted:
        if rb < rb_from:
            continue
        if rb > rb_to:
            break
        for c, _ in rb_map.get(rb, ()):
            if c:
                codes.add(str(c).strip().upper())
    return sorted(codes)


def _nav_codes_for_incremental(
    rb_map: dict[date, list[tuple[str, float]]],
    rb_sorted: list[date],
    anchor_rb: date,
    current_rb: date,
    last_nav_d: date,
    latest_d: date,
) -> list[str]:
    """
    增量 Wind 成分：默认仅当前调仓期；末净值日与最新日之间若有新调仓才合并多期。
    减少 EOD 拉取股票数与 Turso 无关但缩短 Wind/SQL 窗口。
    """
    has_new_rb = any(last_nav_d < rb <= latest_d for rb in rb_sorted)
    if not has_new_rb:
        codes: set[str] = set()
        for c, _ in rb_map.get(current_rb, ()):
            if c:
                codes.add(str(c).strip().upper())
        if codes:
            return sorted(codes)
    return _nav_union_codes_between_rebalances(
        rb_map, rb_sorted, anchor_rb, current_rb
    )


def _nav_first_trade_on_or_after(rb: date, trade_days: list[str]) -> str | None:
    rb_c = _compact_date(rb)
    for td in trade_days:
        if td >= rb_c:
            return td
    return None


def _nav_notional_from_row(row: Any | None, ic0: float) -> float | None:
    """组合名义市值 = nav_unit × 名义本金（ic0）；锚定/续算均用此尺度，勿用裸 ic0 代替。"""
    if not row or ic0 <= 0:
        return None
    try:
        nu = float(row.get("nav_unit") or 0)
    except (TypeError, ValueError):
        return None
    return nu * ic0 if nu > 0 else None


def _nav_mv_pre_for_rebalance_snap(
    shares: dict[str, float],
    prev_mv: float | None,
    ic0: float,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    td: str,
    last_close_fill: dict[str, float],
) -> float:
    """
    调仓日再平衡前的组合市值。
    已有持仓：按昨持仓×现价；否则用上一日市值 prev_mv（= 末净值×本金），
    仅策略首段首日前才用 ic0（净值≈1）。
    """
    if shares:
        mv = 0.0
        for sc2, sh in list(shares.items()):
            if sh is None or sh <= 0:
                continue
            px = _adj_close_td_ff(day_map, sc2, td, last_close_fill)
            if px is not None:
                mv += sh * px
        return mv
    if prev_mv is not None and prev_mv > 0:
        return prev_mv
    return ic0


def _nav_bootstrap_state_from_append_row(
    db: Session,
    sid: str,
    append_after_c: str,
    rb_sorted: list[date],
    rb_map: dict[date, list[tuple[str, float]]],
    ic0: float,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    bench_code: str,
) -> (
    int,
    date,
    date | None,
    dict[str, float],
    dict[str, float],
    float | None,
    float | None,
) | None:
    """增量/回退重放：以库内末净值日 nav_unit×本金 初始化股数，再只写入之后交易日。"""
    row_last = _nav_fetch_row_on_day(db, sid, append_after_c)
    if not row_last:
        return None
    prev_mv0 = _nav_notional_from_row(row_last, ic0)
    if not prev_mv0 or prev_mv0 <= 0:
        return None
    return _nav_init_state_from_last_row(
        db,
        sid,
        append_after_c,
        rb_sorted,
        rb_map,
        ic0,
        day_map,
        row_last,
        bench_code,
    )


def _nav_align_sim_state_to_db_last(
    append_after_c: str,
    row_last: Any | None,
    ic0: float,
    prev_mv: float | None,
    bench_nav_acc: float | None,
    bench_code: str,
) -> tuple[float | None, float | None]:
    """增量模拟经过末净值日后，用库内净值对齐 prev_mv / 基准净值，保证后续日收益率与列表指标口径连续。"""
    if row_last is None:
        return prev_mv, bench_nav_acc
    aligned = _nav_notional_from_row(row_last, ic0)
    if aligned is not None and aligned > 0:
        prev_mv = aligned
    if bench_code:
        try:
            bn = row_last.get("benchmark_nav")
            if bn is not None and float(bn) > 0:
                bench_nav_acc = float(bn)
        except (TypeError, ValueError):
            pass
    return prev_mv, bench_nav_acc


def _nav_last_good_trade_compact(db: Session, sid: str) -> str | None:
    """尺度断裂时，返回仍与最新净值同尺度的最后一个交易日（compact）。"""
    rows = db.execute(
        text(
            f"""
            SELECT trade_date, nav_unit
            FROM strategy_nav_daily
            WHERE strategy_id=:sid AND nav_unit IS NOT NULL
            ORDER BY {sql_order_date_desc("trade_date")}
            LIMIT 12
            """
        ),
        {"sid": sid},
    ).mappings().all()
    if len(rows) < 2:
        return None
    try:
        top = float(rows[0]["nav_unit"])
    except (TypeError, ValueError):
        return None
    if top <= 0:
        return None
    for i in range(1, len(rows)):
        try:
            old = float(rows[i]["nav_unit"])
        except (TypeError, ValueError):
            continue
        if old <= 0:
            continue
        ratio = top / old
        if ratio < 0.85 or ratio > 1.15:
            d = _row_sql_date(rows[i]["trade_date"])
            return _compact_date(d) if d else None
    d0 = _row_sql_date(rows[0]["trade_date"])
    return _compact_date(d0) if d0 else None


def _nav_scale_break_detected(db: Session, sid: str) -> bool:
    """末净值与数日前净值单位净值尺度突变（如 3.7→1.0），列表指标会出现约 -70% 假跌幅。"""
    rows = db.execute(
        text(
            f"""
            SELECT nav_unit
            FROM strategy_nav_daily
            WHERE strategy_id=:sid AND nav_unit IS NOT NULL
            ORDER BY {sql_order_date_desc("trade_date")}
            LIMIT 8
            """
        ),
        {"sid": sid},
    ).mappings().all()
    if len(rows) < 2:
        return False
    try:
        top = float(rows[0]["nav_unit"])
    except (TypeError, ValueError):
        return False
    if top <= 0:
        return False
    ref = rows[5] if len(rows) > 5 else rows[-1]
    try:
        old = float(ref["nav_unit"])
    except (TypeError, ValueError):
        return False
    if old <= 0:
        return False
    ratio = top / old
    return ratio < 0.85 or ratio > 1.15


def _nav_shares_from_holding_snapshot(
    db: Session,
    sid: str,
    trade_d: date,
    rebalance: date,
    prev_mv: float,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    td_compact: str,
    last_close_fill: dict[str, float],
) -> dict[str, float] | None:
    """
    用末净值日 strategy_holding_daily 快照的 latest_weight 反推股数，
    与当日 nav_unit 尺度一致（删库后增量勿仅用 strategy_positions 当前权重）。
    """
    if prev_mv is None or prev_mv <= 0:
        return None
    td_iso = trade_d.isoformat()
    rb_iso = rebalance.isoformat()
    rb_cmp = rb_iso.replace("-", "")
    td_cmp = _compact_date(trade_d)
    rows = db.execute(
        text(
            f"""
            SELECT stock_code, latest_weight, period_weight, latest_price
            FROM strategy_holding_daily
            WHERE strategy_id=:sid
              AND rebalance_date IN (:rb_iso, :rb_cmp)
              AND (
                trade_date = :td_iso
                OR {sql_date_compact_expr("trade_date")} = :td_cmp
              )
            """
        ),
        {
            "sid": sid,
            "rb_iso": rb_iso,
            "rb_cmp": rb_cmp,
            "td_iso": td_iso,
            "td_cmp": td_cmp,
        },
    ).mappings().all()
    if not rows:
        prior = db.execute(
            text(
                f"""
                SELECT stock_code, latest_weight, period_weight, latest_price
                FROM strategy_holding_daily
                WHERE strategy_id=:sid
                  AND rebalance_date IN (:rb_iso, :rb_cmp)
                  AND {sql_date_compact_expr("trade_date")} <= :td_cmp
                ORDER BY {sql_order_date_desc("trade_date")}
                """
            ),
            {
                "sid": sid,
                "rb_iso": rb_iso,
                "rb_cmp": rb_cmp,
                "td_cmp": td_cmp,
            },
        ).mappings().all()
        if not prior:
            return None
        seen: set[str] = set()
        rows = []
        for r in prior:
            sc = _wind_code_key(r.get("stock_code"))
            if not sc or sc in seen:
                continue
            seen.add(sc)
            rows.append(r)
    shares: dict[str, float] = {}
    tw = 0.0
    for r in rows:
        try:
            tw += max(float(r.get("latest_weight") or 0.0), 0.0)
        except (TypeError, ValueError):
            pass
    for r in rows:
        sc = _wind_code_key(r["stock_code"])
        if not sc:
            continue
        try:
            w = float(r.get("latest_weight") or 0.0)
        except (TypeError, ValueError):
            w = 0.0
        if w <= 0:
            try:
                pw = float(r.get("period_weight") or 0.0)
            except (TypeError, ValueError):
                pw = 0.0
            w = pw / tw if tw > 0 else 0.0
        if w <= 0:
            continue
        px = _adj_close_td_ff(day_map, sc, td_compact, last_close_fill)
        if px is None or px <= 0:
            try:
                px = float(r.get("latest_price") or 0.0)
            except (TypeError, ValueError):
                px = 0.0
        if px is not None and px > 0:
            shares[sc] = prev_mv * w / float(px)
    return shares if shares else None


def _nav_holding_snapshot_on_day(
    db: Session, sid: str, rebalance: date, td_compact: str
) -> bool:
    """末净值日是否有该调仓期的 strategy_holding_daily 快照（可不用 Wind 复权价做 bootstrap）。"""
    td = str(td_compact or "").strip()[:8]
    if not td or len(td) < 8:
        return False
    rb_iso = rebalance.isoformat()
    rb_cmp = rb_iso.replace("-", "")
    td_iso = datetime.strptime(td, "%Y%m%d").date().isoformat()
    n = int(
        db.execute(
            text(
                f"""
                SELECT COUNT(*) FROM strategy_holding_daily
                WHERE strategy_id=:sid
                  AND rebalance_date IN (:rb_iso, :rb_cmp)
                  AND (
                    trade_date = :td_iso
                    OR {sql_date_compact_expr("trade_date")} = :td_cmp
                  )
                """
            ),
            {
                "sid": sid,
                "rb_iso": rb_iso,
                "rb_cmp": rb_cmp,
                "td_iso": td_iso,
                "td_cmp": td,
            },
        ).scalar()
        or 0
    )
    return n > 0


def _nav_rebalance_resume_anchor_compact(
    db: Session,
    sid: str,
    current_rb: date,
    trade_days: list[str],
) -> str | None:
    """调仓期续传锚点：调仓日有净值则用调仓日，否则用调仓后首个有净值的交易日。"""
    rb_cmp = _compact_date(current_rb)
    if _nav_fetch_row_on_day(db, sid, rb_cmp):
        return rb_cmp
    for d in trade_days:
        if d >= rb_cmp and _nav_fetch_row_on_day(db, sid, d):
            return d
    return None


def _resolve_nav_append_anchor(
    db: Session,
    sid: str,
    rb_sorted: list[date],
    trade_days: list[str],
) -> str | None:
    """
    续算锚点：库内末净值日；若在调仓期中间且无持仓快照则回退到最近调仓日
    （或该调仓后首个有净值的交易日）。无净值行时返回 None（从首调仓全量）。
    """
    last_max = _strategy_nav_max_trade_compact(db, sid)
    if not last_max or not rb_sorted:
        return None
    td_cmp = str(last_max).strip().replace("-", "")[:8]
    if len(td_cmp) != 8 or not td_cmp.isdigit():
        return None
    try:
        anchor_d = datetime.strptime(td_cmp, "%Y%m%d").date()
    except ValueError:
        return None
    _, current_rb = _nav_rb_idx_on_date(rb_sorted, anchor_d)
    if anchor_d <= current_rb:
        return td_cmp
    if _nav_holding_snapshot_on_day(db, sid, current_rb, td_cmp):
        return td_cmp
    rollback = _nav_rebalance_resume_anchor_compact(db, sid, current_rb, trade_days)
    if not rollback or rollback == td_cmp:
        return td_cmp
    _log.warning(
        "nav %s: append anchor %s mid rebalance %s without holding; rollback to %s",
        sid,
        td_cmp,
        _compact_date(current_rb),
        rollback,
    )
    return rollback


def _nav_delete_nav_for_rebuild(
    db: Session,
    sid: str,
    append_after_c: str | None,
    *,
    force_reset: bool = False,
) -> None:
    """force_reset 或 append_after_c 为 None：删策略全部净值；否则仅删锚点之后。"""
    if force_reset or not append_after_c:
        db.execute(
            text("DELETE FROM strategy_nav_daily WHERE strategy_id=:sid"),
            {"sid": sid},
        )
    else:
        db.execute(
            text(
                f"""
                DELETE FROM strategy_nav_daily
                WHERE strategy_id=:sid
                  AND {sql_date_compact_expr("trade_date")} > :resume_after
                """
            ),
            {"sid": sid, "resume_after": append_after_c},
        )
    db.commit()


def _nav_incremental_eod_ready(
    db: Session,
    sid: str,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    append_after_c: str,
    rb_on_last_nav: date,
    period_codes: list[str],
    sim_days: list[str],
) -> tuple[bool, str]:
    """
    增量净值 Wind 窗口校验：
    - 末净值日：有 EOD 或有持仓快照即可 bootstrap；
    - 待补写首日：必须有 EOD（否则无法模拟下一交易日）。
    """
    has_snap = _nav_holding_snapshot_on_day(db, sid, rb_on_last_nav, append_after_c)
    has_append_eod = _eod_day_map_has_trade(day_map, append_after_c, period_codes)
    if not has_append_eod and not has_snap:
        return False, f"末净值日 {append_after_c} 无 Wind 行情且无持仓快照"
    if sim_days:
        first_sim = sim_days[0]
        if not _eod_day_map_has_trade(day_map, first_sim, period_codes):
            return False, f"待补写首日 {first_sim} 无 Wind 行情"
    return True, ""


def _nav_rescale_shares_to_notional(
    shares: dict[str, float],
    target_mv: float,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    td_compact: str,
    last_close_fill: dict[str, float],
) -> dict[str, float]:
    """按可用收盘价（Wind 或快照价 last_close_fill）把股数缩放到目标组合市值。"""
    if target_mv <= 0 or not shares:
        return shares
    mv = 0.0
    for sc, sh in shares.items():
        if sh is None or sh <= 0:
            continue
        px = _adj_close_td_ff(day_map, sc, td_compact, last_close_fill)
        if px is not None:
            mv += sh * px
    if mv <= 0:
        return shares
    scale = target_mv / mv
    if abs(scale - 1.0) <= 1e-9:
        return shares
    return {sc: sh * scale for sc, sh in shares.items() if sh and sh > 0}


def _nav_init_state_from_last_row(
    db: Session,
    sid: str,
    append_after_c: str,
    rb_sorted: list[date],
    rb_map: dict[date, list[tuple[str, float]]],
    ic0: float,
    day_map: dict,
    row_last: Any | None,
    bench_code: str,
) -> tuple[
    int,
    date,
    date | None,
    dict[str, float],
    dict[str, float],
    float | None,
    float | None,
]:
    """用库内末净值日持仓与市值初始化，仅用于模拟 append_after_c 之后的交易日。"""
    last_nav_d = datetime.strptime(append_after_c[:8], "%Y%m%d").date()
    rb_idx, current_rb = _nav_rb_idx_on_date(rb_sorted, last_nav_d)
    prev_mv = _nav_notional_from_row(row_last, ic0)
    if not prev_mv or prev_mv <= 0:
        prev_mv = None
    bench_nav_acc: float | None = 1.0 if bench_code else None
    if bench_code and row_last and row_last.get("benchmark_nav") is not None:
        try:
            bench_nav_acc = float(row_last["benchmark_nav"])
        except (TypeError, ValueError):
            bench_nav_acc = 1.0
    last_close_fill: dict[str, float] = {}
    shares: dict[str, float] | None = None
    if prev_mv is not None and prev_mv > 0:
        shares = _nav_shares_from_holding_snapshot(
            db,
            sid,
            last_nav_d,
            current_rb,
            prev_mv,
            day_map,
            append_after_c,
            last_close_fill,
        )
    if not shares:
        holdings0 = rb_map.get(current_rb, [])
        if prev_mv is None or prev_mv <= 0:
            return rb_idx, current_rb, last_nav_d, {}, last_close_fill, None, bench_nav_acc
        shares = _nav_snap_shares_from_holdings(
            holdings0, prev_mv, day_map, append_after_c, last_close_fill
        )
    prev_mv, bench_nav_acc = _nav_align_sim_state_to_db_last(
        append_after_c, row_last, ic0, prev_mv, bench_nav_acc, bench_code
    )
    if shares and prev_mv is not None and prev_mv > 0:
        shares = _nav_rescale_shares_to_notional(
            shares, prev_mv, day_map, append_after_c, last_close_fill
        )
    return rb_idx, current_rb, last_nav_d, shares, last_close_fill, prev_mv, bench_nav_acc


def _nav_init_matches_last_row(
    row_last: Any | None,
    shares: dict[str, float],
    day_map: dict,
    append_after_c: str,
    last_close_fill: dict[str, float],
    ic0: float,
    *,
    tol: float = 0.03,
) -> bool:
    """增量初始化后，组合市值折算的 nav_unit 须与库内末净值日一致，否则勿写入（防删库后断尺）。"""
    if not row_last or ic0 <= 0 or not shares:
        return bool(shares)
    try:
        nu_db = float(row_last.get("nav_unit") or 0)
    except (TypeError, ValueError):
        return True
    if nu_db <= 0:
        return True
    mv = 0.0
    for sc, sh in shares.items():
        if sh is None or sh <= 0:
            continue
        px = _adj_close_td_ff(day_map, sc, append_after_c, last_close_fill)
        if px is not None:
            mv += sh * px
    if mv <= 0:
        return False
    nu_calc = mv / ic0
    return abs(nu_calc - nu_db) / nu_db <= tol


def _nav_init_mismatch_detail(
    row_last: Any | None,
    shares: dict[str, float],
    day_map: dict,
    append_after_c: str,
    last_close_fill: dict[str, float],
    ic0: float,
) -> str:
    if not row_last or not shares or ic0 <= 0:
        return "empty row/shares"
    try:
        nu_db = float(row_last.get("nav_unit") or 0)
    except (TypeError, ValueError):
        return "nav_unit invalid"
    mv = 0.0
    priced = 0
    for sc, sh in shares.items():
        if sh is None or sh <= 0:
            continue
        px = _adj_close_td_ff(day_map, sc, append_after_c, last_close_fill)
        if px is not None:
            mv += sh * px
            priced += 1
    if mv <= 0:
        return f"nu_db={nu_db:.6f} priced_stocks=0/{len(shares)}"
    nu_calc = mv / ic0
    rel = abs(nu_calc - nu_db) / nu_db if nu_db > 0 else 0.0
    return (
        f"nu_db={nu_db:.6f} nu_calc={nu_calc:.6f} rel_err={rel:.4f} "
        f"priced={priced}/{len(shares)}"
    )


def _nav_fetch_row_on_day(db: Session, sid: str, td_compact: str) -> Any | None:
    return (
        db.execute(
            text(
                f"""
                SELECT nav_unit, benchmark_nav, rebalance_date
                FROM strategy_nav_daily
                WHERE strategy_id=:sid
                  AND {sql_date_compact_expr("trade_date")} = :td_cmp
                LIMIT 1
                """
            ),
            {"sid": sid, "td_cmp": td_compact},
        )
        .mappings()
        .first()
    )


def _nav_snap_shares_from_holdings(
    holdings: list[tuple[str, float]],
    mv_pre: float,
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    td: str,
    last_close_fill: dict[str, float],
) -> dict[str, float]:
    total_weight = sum(max(float(w or 0.0), 0.0) for _, w in holdings)
    new_shares: dict[str, float] = {}
    if total_weight <= 0 or mv_pre <= 0:
        return new_shares
    valid: list[tuple[str, float, float]] = []
    for sc, w0 in holdings:
        w0p = max(float(w0 or 0.0), 0.0)
        if w0p <= 0:
            continue
        px = _adj_close_td_ff(day_map, sc, td, last_close_fill)
        if px is None or px <= 0:
            continue
        valid.append((sc, w0p, px))
    tw2 = sum(w for _, w, _ in valid)
    if tw2 > 0:
        for sc, w0p, px in valid:
            new_shares[sc] = mv_pre * (w0p / tw2) / px
    return new_shares


def _nav_process_one_trade_day(
    *,
    td: str,
    td_date: date,
    rb_sorted: list[date],
    rb_map: dict[date, list[tuple[str, float]]],
    rb_idx: int,
    current_rb: date,
    prev_td_date: date | None,
    shares: dict[str, float],
    last_close_fill: dict[str, float],
    day_map: dict[str, dict[str, tuple[float | None, float | None]]],
    ic0: float,
    prev_mv: float | None,
    bench_code: str,
    bench_quads: list,
    bench_day_map: dict[str, tuple[float | None, float | None]],
    bench_nav_acc: float | None,
) -> tuple[
    int,
    date,
    date | None,
    dict[str, float],
    dict[str, float],
    float | None,
    float | None,
    dict[str, Any],
]:
    rb_idx_at_start = rb_idx
    while rb_idx + 1 < len(rb_sorted) and rb_sorted[rb_idx + 1] <= td_date:
        rb_idx += 1
        current_rb = rb_sorted[rb_idx]
    snap_rebalance = (
        (rb_idx != rb_idx_at_start)
        or (td_date == current_rb)
        or (
            prev_td_date is not None
            and prev_td_date < current_rb <= td_date
        )
        or (prev_td_date is None and current_rb <= td_date)
    )
    holdings = rb_map.get(current_rb, [])
    if snap_rebalance:
        mv_pre = _nav_mv_pre_for_rebalance_snap(
            shares, prev_mv, ic0, day_map, td, last_close_fill
        )
        shares = _nav_snap_shares_from_holdings(
            holdings, mv_pre, day_map, td, last_close_fill
        )
    mv_eod = 0.0
    for sc2, sh in shares.items():
        if sh is None or sh <= 0:
            continue
        px = _adj_close_td_ff(day_map, sc2, td, last_close_fill)
        if px is not None:
            mv_eod += sh * px
    nav = mv_eod / ic0 if ic0 > 0 else 1.0
    if prev_mv is not None and prev_mv > 0:
        day_ret = mv_eod / prev_mv - 1.0
    else:
        day_ret = (mv_eod / ic0 - 1.0) if ic0 > 0 else 0.0
    prev_mv = mv_eod
    prev_td_date = td_date
    bench_ret_ins = None
    bench_nav_ins = None
    if bench_code and bench_nav_acc is not None:
        br = _bench_return_on_trade_day(
            td, bench_quads=bench_quads, bench_day_map=bench_day_map
        )
        bench_nav_acc, bench_ret_ins, bench_nav_ins = _step_benchmark_nav_acc(
            bench_nav_acc, br, allow_flat=True
        )
    row = {
        "td": td_date,
        "nav": nav,
        "ret": day_ret,
        "bret": bench_ret_ins,
        "bnav": bench_nav_ins,
        "rb": current_rb,
    }
    return (
        rb_idx,
        current_rb,
        prev_td_date,
        shares,
        last_close_fill,
        prev_mv,
        bench_nav_acc,
        row,
    )


def _rebuild_nav_forward_from_anchor(
    db: Session,
    wind: Any,
    sid: str,
    rb_map: dict[date, list[tuple[str, float]]],
    bench_code: str,
    start_c: str,
    latest_trade_c: str,
    ic0: float,
    append_after_c: str | None,
    trade_days: list[str],
    *,
    sync_job_id: int | None = None,
    progress_cb: Callable[[str], None] | None = None,
) -> tuple[bool, Any]:
    """
    统一前向净值：从锚点日 bootstrap 后仅写入之后各日；append_after_c=None 时从 start_c 起全量模拟。
    Wind/EOD 按区间窄窗拉取，长区间自动分段（与日常增量同一引擎）。
    """
    from_start = not append_after_c
    if not from_start and not _nav_incremental_from_period_enabled():
        _log.info("nav forward %s: disabled (NAV_INCREMENTAL_FROM_CURRENT_PERIOD)", sid)
        return False, wind
    rb_sorted = sorted(rb_map.keys())
    if not rb_sorted or not trade_days:
        return False, wind
    latest_d = datetime.strptime(latest_trade_c[:8], "%Y%m%d").date()
    if from_start:
        sim_days = [d for d in trade_days if d >= start_c]
        if not sim_days:
            return False, wind
        if not _nav_trade_days_continuous(
            sid,
            sim_days,
            label="full-forward",
            sync_job_id=sync_job_id,
            progress_cb=progress_cb,
            db=db,
        ):
            return False, wind
        period_codes = sorted({c for lst in rb_map.values() for c, _ in lst if c})
        eod_start_c = start_c
        anchor_rb = rb_sorted[0]
        last_nav_d = datetime.strptime(start_c[:8], "%Y%m%d").date()
    else:
        last_nav_d = datetime.strptime(append_after_c[:8], "%Y%m%d").date()
        _, anchor_rb = _nav_rb_idx_on_date(rb_sorted, last_nav_d)
        sim_days = [d for d in trade_days if d > append_after_c]
        if not sim_days:
            return True, wind
        if not _nav_trade_days_continuous(
            sid,
            [append_after_c, *sim_days],
            label="forward",
            sync_job_id=sync_job_id,
            progress_cb=progress_cb,
            db=db,
        ):
            return False, wind
        _, current_rb = _nav_rb_idx_on_date(rb_sorted, latest_d)
        period_codes = _nav_codes_for_incremental(
            rb_map, rb_sorted, anchor_rb, current_rb, last_nav_d, latest_d
        )
        eod_start_c = wind_bulk.nav_incremental_eod_start(
            append_after_c, rb_sorted, last_nav_d, latest_d, trade_days
        )
    if not period_codes:
        return False, wind

    inc_max_days = max(1, int(getattr(settings, "nav_incremental_max_sim_days", 31) or 31))
    use_nav_segments = bool(getattr(settings, "nav_rebuild_year_segments", True))
    use_seg = use_nav_segments and (
        _wind_low_memory_mode()
        or len(period_codes) > 250
        or len(sim_days) > inc_max_days
        or from_start
    )
    nav_persist_chunk = max(50, int(getattr(settings, "nav_rebuild_persist_chunk", 400)))
    nav_accum: list[dict[str, Any]] = []
    row_last = (
        None if from_start else _nav_fetch_row_on_day(db, sid, append_after_c)
    )
    trade_days_expected_set = set(trade_days)
    use_verified_flush = from_start or len(sim_days) > inc_max_days

    from app.db import SessionLocalFactory, turso_stream_lock, uses_remote_turso_only

    turso_remote = uses_remote_turso_only()

    def _locked_db_op(fn):
        if turso_remote:
            with turso_stream_lock():
                sess = SessionLocalFactory()
                try:
                    return fn(sess)
                finally:
                    sess.close()
        return fn(db)

    def _flush_accum_batch(
        batch: list[dict[str, Any]],
        *,
        last_td: str | None,
        td_i: int,
        seg_i: int | None = None,
        seg_total: int | None = None,
    ) -> None:
        if not batch:
            return

        def _do(sess: Session) -> None:
            if use_verified_flush:
                _flush_strategy_nav_daily_batch_verified(
                    sess,
                    batch,
                    sid=sid,
                    expected_trade_days=trade_days_expected_set,
                )
            else:
                _flush_strategy_nav_daily_batch(sess, batch)

        _locked_db_op(_do)
        if last_td and (sync_job_id is not None or progress_cb):
            seg_hint = (
                f"段 {seg_i}/{seg_total}，"
                if seg_i is not None and seg_total
                else ""
            )
            _nav_progress_touch(
                f"阶段2/3 {sid}：净值已写入至 {last_td}（{seg_hint}第 {td_i} 日）",
                sync_job_id=sync_job_id,
                progress_cb=progress_cb,
                db=None if turso_remote else db,
                turso_remote=turso_remote,
            )
        _log_runtime_progress(
            f"nav:{sid}",
            f"nav {sid}: flushed {len(batch)} rows through {last_td or '?'}",
            force=True,
        )

    bench_quads: list = []
    bench_day_map: dict[str, tuple[float | None, float | None]] = {}
    if bench_code:

        def _load_bench(sess: Session):
            nonlocal wind
            w, idx_all = wind_bulk.load_index_eod_by_code(
                wind, [bench_code], eod_start_c, latest_trade_c, sess
            )
            return w, idx_all

        if turso_remote:
            with turso_stream_lock():
                bdb = SessionLocalFactory()
                try:
                    wind, idx_all = _load_bench(bdb)
                finally:
                    bdb.close()
        else:
            wind, idx_all = _load_bench(db)
        bench_quads = _bench_quads_for_code(idx_all, bench_code)
        bench_day_map = _index_eod_to_bench_day_map(idx_all, bench_code)
        idx_all.clear()

    if sync_job_id is not None or progress_cb:
        if from_start:
            nav_hint = f"全量自 {start_c} 至 {latest_trade_c}，共 {len(sim_days)} 日"
        else:
            nav_hint = (
                f"锚定调仓 {_compact_date(anchor_rb)} 末净值 {append_after_c} "
                f"EOD {eod_start_c}→{latest_trade_c} 补 {len(sim_days)} 日"
            )
        _nav_progress_touch(
            f"{sid}：{nav_hint} 成分约 {len(period_codes)} 只…",
            sync_job_id=sync_job_id,
            progress_cb=progress_cb,
            db=None if turso_remote else db,
            turso_remote=turso_remote,
        )

    if use_seg:
        latest_n = len(period_codes)
        eod_step_months = resolve_nav_rebuild_eod_months(latest_n)
        time_segs = _nav_eod_time_segments(
            eod_start_c, latest_trade_c, step_months=eod_step_months
        )
        if sync_job_id is not None or progress_cb:
            nav_mode = (
                f"全量 {start_c}~{latest_trade_c}"
                if from_start
                else f"续算自 {append_after_c} 至 {latest_trade_c}"
            )
            _nav_progress_touch(
                f"阶段2/3 {sid}：{nav_mode}；动态分段 {eod_step_months} 月/段"
                f"（最新期约 {latest_n} 只），共 {len(time_segs)} 段…",
                sync_job_id=sync_job_id,
                progress_cb=progress_cb,
                db=None if turso_remote else db,
                turso_remote=turso_remote,
            )
        sim_rb_idx = 0
        sim_current_rb = rb_sorted[0]
        prev_td_date: date | None = None
        shares: dict[str, float] = {}
        last_close_fill: dict[str, float] = {}
        prev_mv: float | None = None
        bench_nav_acc: float | None = None
        state_inited = False
        td_i = 0
        for seg_i, (seg_st, seg_ed) in enumerate(time_segs, start=1):
            raise_if_shutting_down()
            seg_days = [d for d in sim_days if seg_st <= d <= seg_ed]
            if not seg_days:
                continue
            seg_codes = _codes_for_nav_segment(
                rb_sorted, rb_map, shares, seg_st, seg_ed
            )
            if not seg_codes:
                continue
            _log_runtime_progress(
                f"nav:{sid}:seg",
                (
                    f"nav {sid}: segment {seg_i}/{len(time_segs)} load "
                    f"{seg_st}..{seg_ed} days={len(seg_days)} codes={len(seg_codes)}"
                ),
                force=True,
            )
            if sync_job_id is not None or progress_cb:
                _nav_progress_touch(
                    f"阶段2/3 {sid}：行情段 {seg_i}/{len(time_segs)}"
                    f"（{seg_st}~{seg_ed}，{len(seg_codes)} 只股票）…",
                    sync_job_id=sync_job_id,
                    progress_cb=progress_cb,
                    db=None if turso_remote else db,
                    turso_remote=turso_remote,
                )

            def _load_maps(sess: Session):
                return _load_segment_wind_maps(
                    wind, sess, seg_codes, seg_st, seg_ed, bench_code
                )

            if turso_remote:
                with turso_stream_lock():
                    seg_db = SessionLocalFactory()
                    try:
                        wind, day_map, seg_bench_dm = _load_maps(seg_db)
                    finally:
                        seg_db.close()
            else:
                wind, day_map, seg_bench_dm = _load_maps(db)
            if bench_code and seg_bench_dm:
                bench_day_map.update(seg_bench_dm)
            _log_runtime_progress(
                f"nav:{sid}:seg",
                (
                    f"nav {sid}: segment {seg_i}/{len(time_segs)} loaded "
                    f"day_map_codes={len(day_map)} pending_rows={len(nav_accum)}"
                ),
                force=True,
            )
            if not state_inited:
                if from_start:
                    sim_rb_idx = 0
                    sim_current_rb = rb_sorted[0]
                    prev_td_date = None
                    shares = {}
                    last_close_fill = {}
                    prev_mv = None
                    bench_nav_acc = 1.0 if bench_code else None
                    state_inited = True
                else:
                    wind = _nav_ensure_anchor_eod_in_day_map(
                        wind, db, day_map, seg_codes, append_after_c
                    )
                    ok_eod, eod_reason = _nav_incremental_eod_ready(
                        db,
                        sid,
                        day_map,
                        append_after_c,
                        anchor_rb,
                        seg_codes,
                        sim_days,
                    )
                    if not ok_eod:
                        _log.warning("nav forward %s: %s", sid, eod_reason)
                        return False, wind
                    (
                        sim_rb_idx,
                        sim_current_rb,
                        prev_td_date,
                        shares,
                        last_close_fill,
                        prev_mv,
                        bench_nav_acc,
                    ) = _nav_init_state_from_last_row(
                        db,
                        sid,
                        append_after_c,
                        rb_sorted,
                        rb_map,
                        ic0,
                        day_map,
                        row_last,
                        bench_code,
                    )
                    if not _nav_init_matches_last_row(
                        row_last, shares, day_map, append_after_c, last_close_fill, ic0
                    ):
                        _log.warning(
                            "nav forward %s: init nav mismatch on %s (%s)",
                            sid,
                            append_after_c,
                            _nav_init_mismatch_detail(
                                row_last,
                                shares,
                                day_map,
                                append_after_c,
                                last_close_fill,
                                ic0,
                            ),
                        )
                        return False, wind
                    state_inited = True
            for td in seg_days:
                if td_i % 15 == 0:
                    raise_if_shutting_down()
                td_i += 1
                if sync_job_id is not None or progress_cb:
                    if td_i % 15 == 0:
                        _nav_progress_touch(
                            f"阶段2/3 {sid}：净值计算 {td}"
                            f"（段 {seg_i}/{len(time_segs)}，第 {td_i} 日）…",
                            sync_job_id=sync_job_id,
                            progress_cb=progress_cb,
                            db=None if turso_remote else db,
                            turso_remote=turso_remote,
                        )
                if td_i % 30 == 0:
                    _log_runtime_progress(
                        f"nav:{sid}",
                        (
                            f"nav {sid}: calc td={td} day={td_i} "
                            f"segment={seg_i}/{len(time_segs)} "
                            f"pending_rows={len(nav_accum)} shares={len(shares)}"
                        ),
                    )
                td_date = datetime.strptime(td, "%Y%m%d").date()
                (
                    sim_rb_idx,
                    sim_current_rb,
                    prev_td_date,
                    shares,
                    last_close_fill,
                    prev_mv,
                    bench_nav_acc,
                    row,
                ) = _nav_process_one_trade_day(
                    td=td,
                    td_date=td_date,
                    rb_sorted=rb_sorted,
                    rb_map=rb_map,
                    rb_idx=sim_rb_idx,
                    current_rb=sim_current_rb,
                    prev_td_date=prev_td_date,
                    shares=shares,
                    last_close_fill=last_close_fill,
                    day_map=day_map,
                    ic0=ic0,
                    prev_mv=prev_mv,
                    bench_code=bench_code,
                    bench_quads=bench_quads,
                    bench_day_map=bench_day_map,
                    bench_nav_acc=bench_nav_acc,
                )
                nav_accum.append(
                    {
                        "sid": sid,
                        "td": row["td"],
                        "nav": row["nav"],
                        "ret": row["ret"],
                        "bret": row["bret"],
                        "bnav": row["bnav"],
                        "rb": row["rb"],
                    }
                )
                if len(nav_accum) >= nav_persist_chunk:
                    batch = nav_accum
                    nav_accum = []
                    _flush_accum_batch(
                        batch,
                        last_td=td,
                        td_i=td_i,
                        seg_i=seg_i,
                        seg_total=len(time_segs),
                    )
            day_map.clear()
            if _wind_low_memory_mode():
                gc.collect()

                def _reopen_wind(sess: Session):
                    nonlocal wind
                    try:
                        wind_sql.close_wind(wind, sess)
                    except Exception:
                        pass
                    wind = wind_sql.open_wind(sess)

                _locked_db_op(_reopen_wind)
        if not state_inited:
            return False, wind
        if nav_accum:
            tail_td = None
            td_val = nav_accum[-1].get("td")
            if td_val is not None:
                tail_td = _compact_date(td_val)
            _flush_accum_batch(
                nav_accum,
                last_td=tail_td,
                td_i=td_i,
                seg_total=len(time_segs),
            )
            nav_accum.clear()
        _log.info(
            "nav forward %s: %s..%s%s",
            sid,
            sim_days[0],
            latest_trade_c,
            "" if from_start else f" after {append_after_c}",
        )
        if _nav_scale_break_detected(db, sid):
            good_c = (
                _nav_last_good_trade_compact(db, sid)
                or append_after_c
                or start_c
            )
            db.execute(
                text(
                    f"""
                    DELETE FROM strategy_nav_daily
                    WHERE strategy_id=:sid
                      AND {sql_date_compact_expr("trade_date")} > :good_c
                    """
                ),
                {"sid": sid, "good_c": good_c},
            )
            db.commit()
            _log.warning(
                "nav forward %s: scale break after write, deleted nav after %s",
                sid,
                good_c,
            )
            return False, wind
        if from_start:
            if not _nav_stored_range_complete(
                db,
                sid,
                start_c,
                latest_trade_c,
                trade_days,
                label="full-forward-final",
            ):
                return False, wind
        else:
            nav_max_c = _strategy_nav_max_trade_compact(db, sid)
            if nav_max_c and latest_trade_c and nav_max_c < latest_trade_c:
                return False, wind
        return True, wind

    wind, eod_by_code = wind_bulk.load_eod_by_code(
        wind, period_codes, eod_start_c, latest_trade_c, db
    )
    day_map = _eod_dict_to_day_map(eod_by_code)
    eod_by_code.clear()
    if from_start:
        sim_rb_idx = 0
        sim_current_rb = rb_sorted[0]
        prev_td_date = None
        shares = {}
        last_close_fill = {}
        prev_mv = None
        bench_nav_acc = 1.0 if bench_code else None
    else:
        ok_eod, eod_reason = _nav_incremental_eod_ready(
            db,
            sid,
            day_map,
            append_after_c,
            anchor_rb,
            period_codes,
            sim_days,
        )
        if not ok_eod:
            _log.warning("nav forward %s: %s", sid, eod_reason)
            day_map.clear()
            return False, wind
        (
            sim_rb_idx,
            sim_current_rb,
            prev_td_date,
            shares,
            last_close_fill,
            prev_mv,
            bench_nav_acc,
        ) = _nav_init_state_from_last_row(
            db,
            sid,
            append_after_c,
            rb_sorted,
            rb_map,
            ic0,
            day_map,
            row_last,
            bench_code,
        )
        if not _nav_init_matches_last_row(
            row_last, shares, day_map, append_after_c, last_close_fill, ic0
        ):
            _log.warning(
                "nav forward %s: init nav mismatch on %s (%s)",
                sid,
                append_after_c,
                _nav_init_mismatch_detail(
                    row_last, shares, day_map, append_after_c, last_close_fill, ic0
                ),
            )
            day_map.clear()
            return False, wind
    td_i = 0
    for td in sim_days:
        td_i += 1
        if sync_job_id is not None or progress_cb:
            if td_i % 15 == 0:
                _nav_progress_touch(
                    f"阶段2/3 {sid}：净值计算 {td}（第 {td_i}/{len(sim_days)} 日）…",
                    sync_job_id=sync_job_id,
                    progress_cb=progress_cb,
                    db=None if turso_remote else db,
                    turso_remote=turso_remote,
                )
        td_date = datetime.strptime(td, "%Y%m%d").date()
        (
            sim_rb_idx,
            sim_current_rb,
            prev_td_date,
            shares,
            last_close_fill,
            prev_mv,
            bench_nav_acc,
            row,
        ) = _nav_process_one_trade_day(
            td=td,
            td_date=td_date,
            rb_sorted=rb_sorted,
            rb_map=rb_map,
            rb_idx=sim_rb_idx,
            current_rb=sim_current_rb,
            prev_td_date=prev_td_date,
            shares=shares,
            last_close_fill=last_close_fill,
            day_map=day_map,
            ic0=ic0,
            prev_mv=prev_mv,
            bench_code=bench_code,
            bench_quads=bench_quads,
            bench_day_map=bench_day_map,
            bench_nav_acc=bench_nav_acc,
        )
        nav_accum.append(
            {
                "sid": sid,
                "td": row["td"],
                "nav": row["nav"],
                "ret": row["ret"],
                "bret": row["bret"],
                "bnav": row["bnav"],
                "rb": row["rb"],
            }
        )
        if len(nav_accum) >= nav_persist_chunk:
            batch = nav_accum
            nav_accum = []
            _flush_accum_batch(batch, last_td=td, td_i=td_i)
    day_map.clear()
    if nav_accum:
        tail_td = _compact_date(nav_accum[-1]["td"]) if nav_accum else None
        _flush_accum_batch(nav_accum, last_td=tail_td, td_i=td_i)
        nav_accum.clear()
    if _wind_low_memory_mode():
        gc.collect()
    _log.info(
        "nav forward %s: %s..%s%s",
        sid,
        sim_days[0],
        latest_trade_c,
        "" if from_start else f" after {append_after_c}",
    )
    if _nav_scale_break_detected(db, sid):
        good_c = (
            _nav_last_good_trade_compact(db, sid) or append_after_c or start_c
        )
        db.execute(
            text(
                f"""
                DELETE FROM strategy_nav_daily
                WHERE strategy_id=:sid
                  AND {sql_date_compact_expr("trade_date")} > :good_c
                """
            ),
            {"sid": sid, "good_c": good_c},
        )
        db.commit()
        _log.warning(
            "nav forward %s: scale break after write, deleted nav after %s",
            sid,
            good_c,
        )
        return False, wind
    if from_start:
        if not _nav_stored_range_complete(
            db,
            sid,
            start_c,
            latest_trade_c,
            trade_days,
            label="full-forward-final",
        ):
            return False, wind
    else:
        nav_max_c = _strategy_nav_max_trade_compact(db, sid)
        if nav_max_c and latest_trade_c and nav_max_c < latest_trade_c:
            return False, wind
    return True, wind


_rebuild_nav_incremental_from_current_period = _rebuild_nav_forward_from_anchor


def _rebuild_nav_for_strategy(
    db: Session,
    wind: Any,
    sid: str,
    _mode_l: str,
    *,
    latest_trade_c_cached: str | None = None,
    mysql_plan: dict[str, Any] | None = None,
    wind_bundle: dict[str, Any] | None = None,
    sync_job_id: int | None = None,
    nav_full_rebuild: bool = True,
    nav_force_reset: bool = False,
    progress_cb: Callable[[str], None] | None = None,
) -> tuple[bool, Any]:
    """
    单策略净值重建。固定股数、调仓日按收盘复权价满仓再平衡。
    调仓自然日可为非交易日：在序列中首个交易日即按「当前调仓期」持仓用该日收盘价建仓（不跳过该期调仓）。

    nav_unit = 收盘组合市值 / 名义本金（settings.strategy_nav_initial_capital，默认 1 亿元）；
    daily_ret = 当日市值 / 上一交易日市值 - 1。

    nav_force_reset=True：用户显式全量重算，删库后从首调仓模拟至最新交易日。
    否则以库内末净值（或回退调仓日）为锚，统一前向补算至最新日（与日常增量同一引擎）。

    mysql_plan / wind_bundle 由 rebuild_nav_series 批量预取时传入，避免重复查库与重复拉 Wind。
    """
    cfg_row: Any | None = None
    if mysql_plan is not None:
        rb_map = mysql_plan["rb_map"]
        code_set = mysql_plan["code_set"]
        bench_code = mysql_plan["bench_code"]
        min_rb = mysql_plan["min_rb"]
        if not code_set:
            return False, wind
    else:
        rb_rows = db.execute(
            text(
                f"""
                SELECT DISTINCT rebalance_date
                FROM strategy_positions
                WHERE strategy_id=:sid
                ORDER BY {sql_order_date_asc("rebalance_date")}
                """
            ),
            {"sid": sid},
        ).mappings().all()
        if not rb_rows:
            return False, wind
        min_rb = _row_sql_date(rb_rows[0]["rebalance_date"])
        if min_rb is None:
            return False, wind
        pos_rows = db.execute(
            text(
                """
                SELECT rebalance_date, stock_code, holding_weight
                FROM strategy_positions
                WHERE strategy_id=:sid
                """
            ),
            {"sid": sid},
        ).mappings().all()
        rb_map: dict[date, list[tuple[str, float]]] = {}
        code_set: set[str] = set()
        for r in pos_rows:
            rd = _row_sql_date(r["rebalance_date"])
            if rd is None:
                continue
            sc = str(r["stock_code"]).strip().upper()
            w = float(r.get("holding_weight") or 0.0)
            rb_map.setdefault(rd, []).append((sc, w))
            if sc:
                code_set.add(sc)
        if not code_set:
            return False, wind
        cfg_row = db.execute(
            text(
                """
                SELECT benchmark_code, benchmark_name
                FROM strategy_configs
                WHERE strategy_id=:sid
                LIMIT 1
                """
            ),
            {"sid": sid},
        ).mappings().first()
        bench_code_raw = (cfg_row or {}).get("benchmark_code")
        bench_code = str(bench_code_raw or "").strip().upper() if bench_code_raw else ""

    start_c = _compact_date(min_rb)
    ic0 = _strategy_nav_notional_capital()

    if latest_trade_c_cached:
        latest_trade_c = str(latest_trade_c_cached).strip()
    else:
        mtd_nav = wind.execute(text(wind_sql.sql_max_trade_dt())).mappings().first()
        latest_trade = mtd_nav["d"] if mtd_nav else None
        if not latest_trade:
            raise RuntimeError("No trade date in winddb")
        latest_trade_c = _compact_date(latest_trade)
        if len(latest_trade_c) < 8:
            latest_trade_c = str(latest_trade).strip().replace("-", "")[:8]

    rb_sorted = sorted(rb_map.keys())
    if wind_bundle is not None:
        td_all = wind_bundle["td"]
        trade_days = [d for d in td_all if d >= start_c]
    else:
        wind, trade_days = wind_bulk.fetch_trade_date_compacts(
            wind, db, start_c, latest_trade_c
        )
    if not trade_days:
        return False, wind

    if nav_force_reset:
        _nav_delete_nav_for_rebuild(db, sid, None, force_reset=True)
        return _rebuild_nav_forward_from_anchor(
            db,
            wind,
            sid,
            rb_map,
            bench_code,
            start_c,
            latest_trade_c,
            ic0,
            None,
            trade_days,
            sync_job_id=sync_job_id,
            progress_cb=progress_cb,
        )

    append_after_c = _resolve_nav_append_anchor(db, sid, rb_sorted, trade_days)
    if append_after_c and append_after_c >= latest_trade_c:
        if _nav_stored_range_complete(
            db,
            sid,
            start_c,
            latest_trade_c,
            trade_days,
            label="complete",
        ):
            _log.info("nav %s: already through %s", sid, append_after_c)
            return True, wind

    if not nav_force_reset and append_after_c and _nav_scale_break_detected(db, sid):
        good_c = _nav_last_good_trade_compact(db, sid) or append_after_c
        _nav_delete_nav_for_rebuild(db, sid, good_c, force_reset=False)
        append_after_c = good_c
        _log.warning(
            "nav %s: scale break; deleted nav after %s, re-forward",
            sid,
            good_c,
        )

    if append_after_c:
        _nav_delete_nav_for_rebuild(db, sid, append_after_c, force_reset=False)

    ok, wind = _rebuild_nav_forward_from_anchor(
        db,
        wind,
        sid,
        rb_map,
        bench_code,
        start_c,
        latest_trade_c,
        ic0,
        append_after_c,
        trade_days,
        sync_job_id=sync_job_id,
        progress_cb=progress_cb,
    )
    if ok:
        return True, wind
    if append_after_c:
        rollback = _resolve_nav_append_anchor(db, sid, rb_sorted, trade_days)
        _log.warning(
            "nav %s: forward from %s failed; retry from resolved anchor %s",
            sid,
            append_after_c,
            rollback,
        )
        if rollback:
            _nav_delete_nav_for_rebuild(db, sid, rollback, force_reset=False)
        return _rebuild_nav_forward_from_anchor(
            db,
            wind,
            sid,
            rb_map,
            bench_code,
            start_c,
            latest_trade_c,
            ic0,
            rollback,
            trade_days,
            sync_job_id=sync_job_id,
            progress_cb=progress_cb,
        )
    return False, wind


def rebuild_nav_series(
    db: Session | None,
    wind,
    strategy_ids: list[str],
    mode: str = "incremental",
    do_commit: bool = True,
    *,
    skip_strategy_ids: set[str] | None = None,
    sync_job_id: int | None = None,
) -> tuple[dict, Any]:
    """
    导入后同步重建净值：固定股数、收盘调仓，按交易日生成 strategy_nav_daily。返回 (结果字典, 当前 Wind 连接)。

    低内存模式（wind_low_memory_mode）：按策略串行拉 Wind EOD/指数/日历，算完即释放，适合 Render 免费档。
    否则合并多策略一次性拉 EOD（本机大内存、减少 SQL Server 往返）。
    导入后默认全量重写净值；日常 run_update 为自库中最后净值日起逐日补至最新交易日。
    """
    mode_l = str(mode or "incremental").strip().lower()
    if mode_l not in ("incremental", "full"):
        raise ValueError(f"invalid nav rebuild mode: {mode}")
    from app.db import SessionLocalFactory, turso_stream_lock, uses_remote_turso_only

    turso_remote = uses_remote_turso_only()

    sids = [str(x).strip() for x in strategy_ids if str(x or "").strip()]
    if not sids:
        if do_commit and db is not None:
            db.commit()
        return {"rebuilt": 0, "failed": 0, "errors": []}, wind

    if sync_job_id is not None:
        _admin_sync_job_touch(
            sync_job_id,
            "nav",
            "阶段2/3：查询 Wind 最新交易日…",
            db=None if turso_remote else db,
            do_commit=True,
        )
    mtd_nav = wind.execute(text(wind_sql.sql_max_trade_dt())).mappings().first()
    latest_trade = mtd_nav["d"] if mtd_nav else None
    if not latest_trade:
        if do_commit and db is not None:
            db.commit()
        return {"rebuilt": 0, "failed": len(sids), "errors": ["No trade date in winddb"]}, wind
    latest_trade_c = _compact_date(latest_trade)
    if len(latest_trade_c) < 8:
        latest_trade_c = str(latest_trade).strip().replace("-", "")[:8]

    if turso_remote:
        with turso_stream_lock():
            plan_db = SessionLocalFactory()
            try:
                plans = _batch_nav_mysql_plans(plan_db, sids, mode_l)
            finally:
                plan_db.close()
    else:
        if db is None:
            raise ValueError("rebuild_nav_series requires db session")
        plans = _batch_nav_mysql_plans(db, sids, mode_l)
    low_mem = _wind_low_memory_mode()
    wind_bundle: dict[str, Any] | None = None
    union_codes: set[str] = set()
    union_bench: set[str] = set()
    global_st: str | None = None
    for _sid, pl in plans.items():
        for c in pl["code_set"]:
            union_codes.add(c)
        bc = str(pl.get("bench_code") or "").strip().upper()
        if bc:
            union_bench.add(bc)
        st = pl["start_c"]
        if global_st is None or st < global_st:
            global_st = st
    merge_nav_prefetch = (not low_mem) and _use_wind_merged_prefetch(len(sids), len(union_codes))
    if merge_nav_prefetch and global_st and union_codes:
            if turso_remote:
                with turso_stream_lock():
                    bundle_db = SessionLocalFactory()
                    try:
                        wind, eod_all = wind_bulk.load_eod_by_code(
                            wind, sorted(union_codes), global_st, latest_trade_c, bundle_db
                        )
                        if union_bench:
                            wind, idx_all = wind_bulk.load_index_eod_by_code(
                                wind, sorted(union_bench), global_st, latest_trade_c, bundle_db
                            )
                        else:
                            idx_all = {}
                        wind, td_all = wind_bulk.fetch_trade_date_compacts(
                            wind, bundle_db, global_st, latest_trade_c
                        )
                    finally:
                        bundle_db.close()
            else:
                wind, eod_all = wind_bulk.load_eod_by_code(
                    wind, sorted(union_codes), global_st, latest_trade_c, db
                )
                if union_bench:
                    wind, idx_all = wind_bulk.load_index_eod_by_code(
                        wind, sorted(union_bench), global_st, latest_trade_c, db
                    )
                else:
                    idx_all = {}
                wind, td_all = wind_bulk.fetch_trade_date_compacts(wind, db, global_st, latest_trade_c)
            wind_bundle = {"eod": eod_all, "idx": idx_all, "td": td_all}
    elif not low_mem and global_st and union_codes and len(sids) > 1:
        _log.info(
            "nav rebuild: skip merged EOD prefetch (strategies=%s union_codes=%s); per-strategy load",
            len(sids),
            len(union_codes),
        )

    skip = {str(x).strip() for x in (skip_strategy_ids or []) if str(x).strip()}
    if mode_l == "full" and skip:
        _log.warning("nav full rebuild ignores completed_nav checkpoints: %s", sorted(skip))
        skip = set()
    done = len(skip)
    failed = 0
    errors: list[str] = []
    completed_nav = sorted(skip)
    total_nav = len(sids)
    for sid in sids:
        raise_if_shutting_down()
        if sid in skip:
            continue
        nav_seq = len(completed_nav) + 1
        if sync_job_id is not None:
            _admin_sync_job_touch(
                sync_job_id,
                "nav",
                f"阶段2/3 净值 [{nav_seq}/{total_nav}] 正在处理 {sid}…",
                db=None if turso_remote else db,
                do_commit=bool(turso_remote),
            )
        per_bundle: dict[str, Any] | None = None
        try:
            # Turso 远程 Hrana 不支持 SQLAlchemy SAVEPOINT（sa_savepoint_*），勿用 begin_nested。
            pl = plans.get(sid)

            def _rebuild_one(sdb: Session) -> tuple[bool, Any]:
                if pl and low_mem:
                    return _rebuild_nav_for_strategy(
                        sdb,
                        wind,
                        sid,
                        mode_l,
                        latest_trade_c_cached=latest_trade_c,
                        mysql_plan=pl,
                        wind_bundle=None,
                        sync_job_id=sync_job_id,
                    )
                if pl:
                    return _rebuild_nav_for_strategy(
                        sdb,
                        wind,
                        sid,
                        mode_l,
                        latest_trade_c_cached=latest_trade_c,
                        mysql_plan=pl,
                        wind_bundle=wind_bundle,
                        sync_job_id=sync_job_id,
                    )
                return _rebuild_nav_for_strategy(
                    sdb,
                    wind,
                    sid,
                    mode_l,
                    latest_trade_c_cached=latest_trade_c,
                    sync_job_id=sync_job_id,
                )

            if turso_remote:
                # 勿整段持锁：净值重建内 _locked_db_op 已按 SQL 批次加锁，Wind 在锁外执行
                sdb = SessionLocalFactory()
                try:
                    if wind is None:
                        wind = wind_sql.open_wind(sdb)
                    counted, wind = _rebuild_one(sdb)
                    if counted:
                        done += 1
                        completed_nav.append(sid)
                        if sync_job_id is not None:
                            with turso_stream_lock():
                                ck_row = (
                                    sdb.execute(
                                        text(
                                            "SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"
                                        ),
                                        {"id": sync_job_id},
                                    )
                                    .mappings()
                                    .first()
                                )
                                cp = _sync_load_checkpoint(
                                    ck_row.get("checkpoint_json") if ck_row else None
                                )
                                _sync_save_checkpoint(
                                    sdb,
                                    sync_job_id,
                                    completed_import=cp.get("completed_import") or [],
                                    completed_nav=(
                                        [] if mode_l == "full" else sorted(set(completed_nav))
                                    ),
                                    completed_update_rb=cp.get("completed_update_rb") or [],
                                    nav_progress=_sync_nav_progress_map(cp.get("nav_progress")),
                                    stage="nav",
                                    do_commit=False,
                                )
                                _admin_sync_job_touch(
                                    sync_job_id,
                                    "nav",
                                    f"阶段2/3 净值 [{len(completed_nav)}/{total_nav}] 已完成 {sid}",
                                    db=sdb,
                                    do_commit=False,
                                )
                    if do_commit:
                        with turso_stream_lock():
                            sdb.commit()
                    if counted:
                        try:
                            from app.strategy_list_metrics import (
                                refresh_strategy_list_metrics_one,
                            )

                            refresh_strategy_list_metrics_one(
                                sdb, sid, do_commit=True
                            )
                        except Exception:
                            _log.exception(
                                "strategy_list_metrics refresh after nav sid=%s", sid
                            )
                    else:
                        failed += 1
                        errors.append(f"{sid}: nav rebuild incomplete")
                        if sync_job_id is not None:
                            _admin_sync_job_touch(
                                sync_job_id,
                                "nav",
                                f"阶段2/3 净值 [{len(completed_nav)}/{total_nav}] {sid} 未完成",
                                db=sdb,
                                do_commit=do_commit,
                            )
                finally:
                    if low_mem:
                        try:
                            wind_sql.close_wind(wind, sdb)
                        except Exception:
                            pass
                        wind = None
                    sdb.close()
            else:
                counted, wind = _rebuild_one(db)
                if counted:
                    done += 1
                    completed_nav.append(sid)
                    if sync_job_id is not None:
                        ck_row = (
                            db.execute(
                                text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
                                {"id": sync_job_id},
                            )
                            .mappings()
                            .first()
                        )
                        cp = _sync_load_checkpoint(ck_row.get("checkpoint_json") if ck_row else None)
                        _sync_save_checkpoint(
                            db,
                            sync_job_id,
                            completed_import=cp.get("completed_import") or [],
                            completed_nav=(
                                [] if mode_l == "full" else sorted(set(completed_nav))
                            ),
                            completed_update_rb=cp.get("completed_update_rb") or [],
                            nav_progress=_sync_nav_progress_map(cp.get("nav_progress")),
                            stage="nav",
                            do_commit=False,
                        )
                        _admin_sync_job_touch(
                            sync_job_id,
                            "nav",
                            f"阶段2/3 净值 [{len(completed_nav)}/{total_nav}] 已完成 {sid}",
                            db=db,
                            do_commit=False,
                        )
                if do_commit:
                    db.commit()
                if counted:
                    try:
                        from app.strategy_list_metrics import (
                            refresh_strategy_list_metrics_one,
                        )

                        refresh_strategy_list_metrics_one(db, sid, do_commit=True)
                    except Exception:
                        _log.exception(
                            "strategy_list_metrics refresh after nav sid=%s", sid
                        )
                else:
                    failed += 1
                    errors.append(f"{sid}: nav rebuild incomplete")
                    if sync_job_id is not None:
                        _admin_sync_job_touch(
                            sync_job_id,
                            "nav",
                            f"阶段2/3 净值 [{len(completed_nav)}/{total_nav}] {sid} 未完成",
                            db=db,
                            do_commit=do_commit,
                        )
        except Exception as ex:
            failed += 1
            errors.append(f"{sid}: {ex}")
            if sync_job_id is not None:
                _admin_sync_job_touch(
                    sync_job_id,
                    "nav",
                    f"阶段2/3 净值 [{len(completed_nav)}/{total_nav}] {sid} 失败：{ex}"[:6000],
                    db=None if turso_remote else db,
                    do_commit=bool(turso_remote),
                )
            if do_commit and db is not None:
                try:
                    db.rollback()
                except Exception:
                    pass
            if _wind_sql_transient_disconnect(ex):
                try:
                    wind_sql.close_wind(wind, db)
                except Exception:
                    pass
                try:
                    wind = wind_sql.open_wind(db)
                except Exception:
                    pass
        finally:
            _release_wind_memory(per_bundle)
            per_bundle = None
            gc.collect()
    _release_wind_memory(wind_bundle)
    wind_bundle = None
    gc.collect()
    if do_commit and db is not None:
        db.commit()
    return {
        "rebuilt": done,
        "failed": failed,
        "errors": errors[:50],
        "completed_nav_ids": sorted(set(completed_nav)),
        "resumable": len(completed_nav) < len(sids),
    }, wind


def _admin_sync_mark_running(job_id: int, *, resume: bool = False) -> bool:
    from app.db import SessionLocalFactory, turso_stream_lock

    start_msg = "续传已启动，正在从断点继续…" if resume else "后台任务已启动，正在执行…"
    with turso_stream_lock():
        db = SessionLocalFactory()
        try:
            cur = db.execute(
                text(
                    f"""
                    UPDATE admin_sync_jobs
                    SET status='RUNNING', started_at={sql_now()}, stage='start',
                        message=:m, progress_at={sql_now()}
                    WHERE id=:id AND status <> 'ABANDONED'
                    """
                ),
                {"m": start_msg, "id": job_id},
            )
            db.commit()
            return int(getattr(cur, "rowcount", 0) or 0) > 0
        finally:
            db.close()


def _admin_sync_job_touch(
    job_id: int,
    stage: str,
    message: str,
    *,
    db: Session | None = None,
    do_commit: bool = True,
) -> None:
    """更新 admin_sync_jobs 进度。传入 db 时与导入/净值同事务，避免 Turso 双流冲突。"""
    from app.db import SessionLocalFactory, turso_stream_lock

    def _run(sess: Session) -> None:
        sess.execute(
            text(
                f"""
                UPDATE admin_sync_jobs
                SET stage=:st, message=:msg, progress_at={sql_now()}
                WHERE id=:id AND status='RUNNING'
                """
            ),
            {"st": (stage or "")[:64], "msg": (message or "")[:6000], "id": job_id},
        )
        if do_commit:
            sess.commit()

    if db is not None:
        _run(db)
        return
    with turso_stream_lock():
        own = SessionLocalFactory()
        try:
            _run(own)
        finally:
            own.close()


def _clear_update_running_flag_if_stale() -> bool:
    """Clear the process-local update lock when no update job is RUNNING in DB."""
    global _job_running
    if not _job_running:
        return False
    from app.db import SessionLocalFactory

    db = SessionLocalFactory()
    try:
        row = db.execute(
            text("SELECT id FROM strategy_update_jobs WHERE status='RUNNING' LIMIT 1")
        ).first()
        if row is not None:
            return False
        _log.warning("clearing stale in-process update lock: no RUNNING strategy_update_jobs")
        _job_running = False
        return True
    finally:
        db.close()


def _finalize_admin_sync_job(job_id: int, result: dict) -> None:
    from app.db import SessionLocalFactory, turso_stream_lock

    ok = bool(result.get("ok"))
    status = "SUCCESS" if ok else "FAILED"
    errs = result.get("errors") or []
    summary = (
        ("成功：" + json.dumps({k: result.get(k) for k in ("imported", "nav_rebuilt", "stage") if k in result}, ensure_ascii=False))
        if ok
        else ("失败：" + ("；".join(str(x) for x in errs)[:1800]))
    )
    if (not ok) and bool(result.get("resumable")) and str(result.get("stage") or "") in ("shutdown", "interrupted"):
        summary = "\u540e\u53f0\u4efb\u52a1\u88ab\u4e2d\u65ad\uff0c\u53ef\u70b9\u300c\u7eed\u4f20\u300d\u7ee7\u7eed\u3002"
    body = json.dumps(result, ensure_ascii=False)
    with turso_stream_lock():
        db = SessionLocalFactory()
        try:
            if (not ok) and bool(result.get("resumable")):
                try:
                    admin_sync_job_bootstrap_checkpoint(db, job_id, do_commit=False)
                except Exception:
                    _log.exception("admin_sync job %s: bootstrap checkpoint on finalize failed", job_id)
            db.execute(
                text(
                    f"""
                    UPDATE admin_sync_jobs
                    SET status=:st,
                        finished_at={sql_now()},
                        stage=:sg,
                        message=:sm,
                        result_json=:rj
                    WHERE id=:id AND status <> 'ABANDONED'
                    """
                ),
                {
                    "st": status,
                    "sg": str(result.get("stage") or "")[:64],
                    "sm": summary[:4000],
                    "rj": body,
                    "id": job_id,
                },
            )
            db.commit()
        finally:
            db.close()


def execute_admin_sync_pipeline(
    username: str,
    selected_ids: list[str],
    import_mode: str,
    *,
    sync_job_id: int | None = None,
    resume: bool = False,
) -> dict:
    """
    策略配置「净值+EOD」一体化：跳过 Excel 导入 → 重建净值 → 全量更新持仓快照。
    与原先 admin_sync 同步逻辑一致；可选 sync_job_id 写入 admin_sync_jobs 进度。
    """
    ids = [str(x).strip() for x in selected_ids if str(x or "").strip()]
    if not ids:
        return {"ok": False, "stage": "validate", "failed": 0, "errors": ["strategy_ids 为空"]}

    _log.info(
        "admin_sync pipeline start job=%s ids=%s resume=%s mode=%s",
        sync_job_id,
        ids,
        resume,
        import_mode,
    )

    from app.db import SessionLocalFactory, turso_stream_lock

    def p(stage: str, msg: str, *, detached: bool = False) -> None:
        if sync_job_id is not None:
            _admin_sync_job_touch(
                sync_job_id,
                stage,
                msg,
                db=None if detached else db,
                do_commit=True,
            )

    wind = None
    imp: dict | None = None
    nav_ret: dict | None = None
    completed_import: set[str] = set()
    completed_nav: set[str] = set()
    completed_update_rb: set[str] = set()
    ids_set = set(ids)
    skip_excel_import = True
    import_pending = False

    with turso_stream_lock():
        db = SessionLocalFactory()
        try:
            p("precheck", "检查僵尸 RUNNING、数据更新任务互斥…")
            if sync_job_id is not None and resume:
                ck_row = (
                    db.execute(
                        text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
                        {"id": sync_job_id},
                    )
                    .mappings()
                    .first()
                )
                cp = _sync_load_checkpoint(ck_row.get("checkpoint_json") if ck_row else None)
                completed_import = set(cp.get("completed_import") or [])
                completed_nav = set(cp.get("completed_nav") or [])
                completed_update_rb = _sync_normalize_update_rb_done(
                    cp.get("completed_update_rb") or []
                )
                if str(import_mode or "").strip().lower() == "full" and completed_nav:
                    # Full resume must rebuild NAV again. Holding snapshot
                    # checkpoints are kept, but each skipped rebalance is
                    # verified against the target snapshot before it is trusted.
                    _log.warning(
                        "admin_sync job %s: ignore completed_nav checkpoint on full resume: %s",
                        sync_job_id,
                        sorted(completed_nav),
                    )
                    completed_nav = set()
                if completed_import:
                    p(
                        "resume",
                        f"续传：净值 {len(completed_nav)} 策略，"
                        f"持仓快照调仓期 {len(completed_update_rb)} 个已完成",
                    )
            stale_mins = max(1, int(getattr(settings, "stale_running_update_job_minutes", 240)))
            db.execute(
                text(
                    f"""
                    UPDATE strategy_update_jobs
                    SET status='FAILED', finished_at={sql_now()},
                        message=COALESCE(message, '') || '（僵尸RUNNING：已超过 ' || :mins || ' 分钟未结束，同步前自动标记失败；若确为长跑任务请调大 STALE_RUNNING_UPDATE_JOB_MINUTES）'
                    WHERE status='RUNNING'
                      AND started_at < {sql_minutes_ago(':mins')}
                    """
                ),
                {"mins": stale_mins},
            )
            db.commit()
            running = db.execute(
                text(
                    """
                    SELECT id, started_at
                    FROM strategy_update_jobs
                    WHERE status='RUNNING'
                    ORDER BY id DESC
                    LIMIT 1
                    """
                )
            ).first()
            if running:
                rid, rst = running[0], running[1]
                return {
                    "ok": False,
                    "stage": "blocked",
                    "imported": 0,
                    "nav_rebuilt": 0,
                    "failed": len(ids),
                    "errors": [
                        f"已有进行中的数据更新任务 id={rid}（开始于 {rst}），请待其完成后再执行同步。"
                        f"若确认无进程在跑，可将该条 status 改为 FAILED，或等待超过 {stale_mins} 分钟后重试。"
                    ],
                }
            if skip_excel_import:
                completed_import = set(ids_set)
                import_pending = False
            else:
                import_pending = not (ids_set <= completed_import)
            if sync_job_id is not None:
                admin_sync_job_bootstrap_checkpoint(
                    db,
                    sync_job_id,
                    strategy_ids=ids,
                    import_mode=import_mode,
                    skip_excel_import=skip_excel_import,
                    do_commit=True,
                )
        except Exception as ex:
            try:
                db.rollback()
            except Exception:
                pass
            _log.exception("admin_sync job %s precheck failed", sync_job_id)
            return {
                "ok": False,
                "stage": "precheck",
                "imported": 0,
                "nav_rebuilt": 0,
                "failed": len(ids),
                "errors": ["?????????????????"],
            }
        finally:
            db.close()

    if import_pending:
        _log.info("admin_sync job %s: stage1 import begin", sync_job_id)
        db = SessionLocalFactory()
        try:
            p(
                "import",
                f"阶段1/3：从 Excel 导入（{len(ids)} 个策略，{import_mode}"
                f"{f'，续传跳过 {len(completed_import)} 个' if completed_import else ''}）…",
            )
            imp = import_strategy_files(
                db,
                selected_strategy_ids=ids,
                import_mode=import_mode,
                do_commit=True,
                skip_strategy_ids=completed_import,
                sync_job_id=sync_job_id,
                resume=resume,
            )
            completed_import = set(imp.get("completed_strategy_ids") or [])
            pre_verified = set(imp.get("verified_strategy_ids") or [])
            verify_errors: list[str] = []
            if completed_import:
                for sid in sorted(completed_import):
                    if sid in pre_verified:
                        continue
                    try:
                        fp = _strategy_config_excel_path(db, sid)
                        _verify_strategy_full_import_row_count(
                            db,
                            sid,
                            fp,
                            import_mode=import_mode,
                            heavy=True,
                        )
                    except ValueError as ve:
                        completed_import.discard(sid)
                        verify_errors.append(str(ve))
                if verify_errors:
                    imp = {
                        **imp,
                        "failed": max(int(imp.get("failed") or 0), len(verify_errors)),
                        "errors": list(imp.get("errors") or []) + verify_errors,
                        "completed_strategy_ids": sorted(completed_import),
                        "resumable": True,
                    }
            if int(imp.get("failed") or 0) > 0 and not ids_set <= completed_import:
                return {
                    "ok": False,
                    "stage": "import",
                    "resumable": True,
                    **imp,
                }
        except Exception as ex:
            try:
                db.rollback()
            except Exception:
                pass
            _log.exception("admin_sync job %s stage1 import failed", sync_job_id)
            return {
                "ok": False,
                "stage": "import",
                "imported": (imp or {}).get("imported", 0),
                "nav_rebuilt": 0,
                "failed": len(ids),
                "errors": ["\u540e\u53f0\u4efb\u52a1\u88ab\u4e2d\u65ad\uff0c\u8bf7\u70b9\u300c\u7eed\u4f20\u300d\u7ee7\u7eed\u3002"],
            }
        finally:
            db.close()
        _log.info(
            "admin_sync job %s: stage1 import done imported=%s failed=%s",
            sync_job_id,
            len(completed_import),
            int((imp or {}).get("failed") or 0),
        )
        raise_if_shutting_down()
    else:
        imp = {
            "imported": 0,
            "failed": 0,
            "errors": [],
            "completed_strategy_ids": sorted(ids_set),
        }
        p("import", f"阶段1/3：已跳过 Excel 导入，直接进入净值计算（{len(ids_set)} 个策略）", detached=True)

    t1 = float(settings.admin_sync_sleep_after_import_seconds or 0.0)
    if t1 > 0 and not skip_excel_import:
        p("import", f"导入完成，休眠 {t1}s 后进入净值…", detached=True)
        time.sleep(t1)
    gc.collect()

    try:
        raise_if_shutting_down()
        if ids_set <= completed_nav:
            nav_ret = {
                "rebuilt": len(completed_nav),
                "failed": 0,
                "errors": [],
                "completed_nav_ids": sorted(completed_nav),
            }
            p("nav", f"阶段2/3：净值已跳过（{len(completed_nav)} 个策略已于断点完成）", detached=True)
        else:
            _log.info("admin_sync job %s: stage2 nav rebuild begin", sync_job_id)
            p(
                "nav",
                f"阶段2/3：重建净值序列（Wind，{len(ids)} 个策略"
                f"{f'，续传跳过 {len(completed_nav)} 个' if completed_nav else ''}）…",
                detached=True,
            )
            p("nav", "阶段2/3：连接 Wind SQL Server…", detached=True)
            wind = wind_sql.open_wind(None)
            nav_ret, wind = rebuild_nav_series(
                None,
                wind,
                ids,
                mode=import_mode,
                do_commit=True,
                skip_strategy_ids=completed_nav,
                sync_job_id=sync_job_id,
            )
            completed_nav = set(nav_ret.get("completed_nav_ids") or [])
        if not ids_set <= completed_nav:
            nav_errors = list(nav_ret.get("errors") or [])
            missing_nav = sorted(ids_set - completed_nav)
            if missing_nav:
                nav_errors.append("净值未完成策略：" + ",".join(missing_nav))
            return {
                "ok": False,
                "stage": "nav",
                "resumable": True,
                "imported": imp.get("imported", 0),
                "nav_rebuilt": nav_ret.get("rebuilt", 0),
                "failed": max(int(nav_ret.get("failed") or 0), len(missing_nav)),
                "errors": nav_errors,
            }
        if sync_job_id is not None:
            with turso_stream_lock():
                db_ck = SessionLocalFactory()
                try:
                    ck_row = (
                        db_ck.execute(
                            text("SELECT checkpoint_json FROM admin_sync_jobs WHERE id=:id"),
                            {"id": sync_job_id},
                        )
                        .mappings()
                        .first()
                    )
                    cp = _sync_load_checkpoint(ck_row.get("checkpoint_json") if ck_row else None)
                    _sync_save_checkpoint(
                        db_ck,
                        sync_job_id,
                        completed_import=sorted(completed_import),
                        completed_nav=(
                            []
                            if str(import_mode or "").strip().lower() == "full"
                            else sorted(completed_nav)
                        ),
                        completed_update_rb=sorted(completed_update_rb),
                        nav_progress=_sync_nav_progress_map(cp.get("nav_progress")),
                        stage="update",
                        do_commit=True,
                    )
                finally:
                    db_ck.close()
        t2 = float(settings.admin_sync_sleep_after_nav_seconds or 0.0)
        if t2 > 0:
            p("nav", f"净值完成，休眠 {t2}s…", detached=True)
            time.sleep(t2)
        if wind is not None:
            wind_sql.close_wind(wind, None)
            wind = None
        gc.collect()
    except Exception as ex:
        if wind is not None:
            try:
                wind_sql.close_wind(wind, None)
            except Exception:
                pass
        return {
            "ok": False,
            "stage": "nav",
            "resumable": True,
            "imported": (imp or {}).get("imported", 0),
            "nav_rebuilt": (nav_ret or {}).get("rebuilt", 0),
            "failed": len(ids),
            "errors": [str(ex)],
        }

    cap = int(getattr(settings, "admin_sync_wait_idle_update_seconds", 180) or 0)
    step = 3
    raise_if_shutting_down()
    p("wait_idle", "阶段3/3 前：等待进程内数据更新互斥锁释放…", detached=True)
    if cap > 0:
        t0 = time.monotonic()
        last_touch = 0.0
        while _job_running and (time.monotonic() - t0) < cap:
            if _clear_update_running_flag_if_stale():
                p("wait_idle", "阶段3/3 前：已清理残留的数据更新占用标记，继续写入持仓快照…", detached=True)
                break
            elapsed = int(time.monotonic() - t0)
            if elapsed - last_touch >= 9:
                last_touch = float(elapsed)
                p("wait_idle", f"等待其它数据更新结束… 已等待 {elapsed}s / 上限 {cap}s", detached=True)
            raise_if_shutting_down()
            time.sleep(step)
    if _job_running:
        wait_hint = f"已等待 {cap} 秒" if cap > 0 else "未配置等待（ADMIN_SYNC_WAIT_IDLE_UPDATE_SECONDS=0）"
        return {
            "ok": False,
            "stage": "update",
            "imported": (imp or {}).get("imported", 0),
            "nav_rebuilt": (nav_ret or {}).get("rebuilt", 0),
            "failed": len(ids),
            "resumable": True,
            "errors": [
                f"其它数据更新仍占用进程（{wait_hint}）。请待「立即更新」或定时任务结束后再点同步；导入与净值已提交。"
            ],
        }

    p("holding_update", "阶段3/3：写入最新交易日持仓快照（可能较久）…", detached=True)
    raise_if_shutting_down()
    db2: Session | None = None
    try:
        db2 = SessionLocalFactory()
        run_update(
            db2,
            "MANUAL",
            username,
            full_refresh=True,
            selected_strategy_ids=ids,
            do_commit=True,
            skip_nav_rebuild=True,
            sync_job_id=sync_job_id,
            skip_update_rebalance_dates=completed_update_rb,
        )
        p("holding_update", "阶段3/3：持仓快照已完成，正在写入同步结果…", detached=True)
        return {
            "ok": True,
            "stage": "all_success",
            "imported": (imp or {}).get("imported", 0),
            "nav_rebuilt": (nav_ret or {}).get("rebuilt", 0),
            "failed": 0,
            "errors": [],
        }
    except Exception as ex:
        if db2 is not None:
            try:
                db2.rollback()
            except Exception:
                pass
        return {
            "ok": False,
            "stage": "update",
            "resumable": True,
            "imported": (imp or {}).get("imported", 0),
            "nav_rebuilt": (nav_ret or {}).get("rebuilt", 0),
            "failed": len(ids),
            "errors": [str(ex)],
        }
    finally:
        if db2 is not None:
            db2.close()


def run_admin_sync_background_task(
    job_id: int,
    username: str,
    selected_ids: list[str],
    import_mode: str,
    *,
    resume: bool = False,
) -> None:
    """供 FastAPI BackgroundTasks 调用：先标 RUNNING，再跑管道并落库终态。"""
    try:
        if not _admin_sync_mark_running(job_id, resume=resume):
            return
        ret = execute_admin_sync_pipeline(
            username, selected_ids, import_mode, sync_job_id=job_id, resume=resume
        )
        _finalize_admin_sync_job(job_id, ret)
    except ShutdownRequested as ex:
        _log.warning("admin_sync job %s paused for shutdown: %s", job_id, ex)
        _finalize_admin_sync_job(
            job_id,
            {
                "ok": False,
                "stage": "interrupted",
                "resumable": True,
                "imported": 0,
                "nav_rebuilt": 0,
                "failed": len(selected_ids),
                "errors": [str(ex)],
            },
        )
    except Exception as ex:
        _log.exception("admin_sync job %s failed", job_id)
        _finalize_admin_sync_job(
            job_id,
            {
                "ok": False,
                "stage": "exception",
                "resumable": True,
                "imported": 0,
                "nav_rebuilt": 0,
                "failed": len(selected_ids),
                "errors": [str(ex)],
            },
        )


def run_strategy_import_background_task(
    job_id: int,
    *,
    resume: bool = False,
) -> None:
    """守护线程中执行；按批短持 Turso 流锁，避免整段导入阻塞 API 轮询 90 秒。"""
    _run_strategy_import_background_task_impl(job_id, resume=resume)


def _run_strategy_import_background_task_impl(
    job_id: int,
    *,
    resume: bool = False,
) -> None:
    from app.db import SessionLocalFactory

    db = SessionLocalFactory()
    ret: dict[str, Any] | None = None
    ids: list[str] = []
    import_mode = "full"
    try:
        job = get_strategy_import_job_row(db, job_id)
        if not job:
            return
        if str(job.get("status") or "").upper() == "ABANDONED":
            return
        ids = _json_str_list(job.get("strategy_ids_json"))
        import_mode = str(job.get("import_mode") or "full")
        skip = set(_json_str_list(job.get("completed_strategy_ids_json"))) if resume else set()
        db.execute(
            text(
                f"""
                UPDATE strategy_import_jobs
                SET status='RUNNING', started_at={sql_now()}, message=:m, progress_at={sql_now()}
                WHERE id=:id AND status <> 'ABANDONED'
                """
            ),
            {
                "m": "续传执行中…" if resume else "后台导入执行中…",
                "id": job_id,
            },
        )
        db.commit()
        ret = import_strategy_files(
            db,
            selected_strategy_ids=ids,
            import_mode=import_mode,
            do_commit=True,
            skip_strategy_ids=skip,
            strategy_import_job_id=job_id,
            resume=resume,
        )
        _finalize_strategy_import_job(db, job_id, ret, import_mode, ids)
        db.commit()
    except ShutdownRequested as ex:
        _log.warning("strategy_import job %s paused for shutdown: %s", job_id, ex)
        try:
            db.execute(
                text(
                    f"""
                    UPDATE strategy_import_jobs
                    SET status='FAILED', finished_at={sql_now()}, message=:m
                    WHERE id=:id AND status <> 'ABANDONED'
                    """
                ),
                {"m": "\u540e\u53f0\u4efb\u52a1\u88ab\u4e2d\u65ad\uff0c\u8bf7\u70b9\u300c\u7eed\u4f20\u300d\u7ee7\u7eed\u3002", "id": job_id},
            )
            db.commit()
        except Exception:
            db.rollback()
    except Exception as ex:
        _log.exception("strategy_import job %s failed", job_id)
        all_ids = {str(x).strip() for x in ids if str(x).strip()}
        done = set((ret or {}).get("completed_strategy_ids") or [])
        if ret and all_ids <= done and int(ret.get("failed") or 0) == 0:
            try:
                _finalize_strategy_import_job(db, job_id, ret, import_mode, ids)
                db.commit()
                _log.info(
                    "strategy_import job %s: finalize recovered after %s",
                    job_id,
                    type(ex).__name__,
                )
                return
            except Exception:
                _log.exception("strategy_import job %s: finalize recovery failed", job_id)
        try:
            db.execute(
                text(
                    f"""
                    UPDATE strategy_import_jobs
                    SET status='FAILED', finished_at={sql_now()}, message=:m
                    WHERE id=:id AND status <> 'ABANDONED'
                    """
                ),
                {"m": f"异常：{str(ex)[:5900]}", "id": job_id},
            )
            db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()
